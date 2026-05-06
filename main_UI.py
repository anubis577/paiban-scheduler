#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
排班管理软件 - 西南空管局版本（模块化重构版）
"""

import sys
import os
import re
import math
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QTableWidget, QTableWidgetItem, QPushButton, QLabel,
    QLineEdit, QComboBox, QCheckBox, QGroupBox, QFormLayout, QMessageBox,
    QSpinBox, QDoubleSpinBox, QTextEdit, QStatusBar, QDialog, QDialogButtonBox,
    QFrame, QGridLayout, QStyledItemDelegate, QScrollArea, QHeaderView,
    QSizePolicy
)
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont, QColor

# 导入models模块
from models.database import Database
from models.person import Person
from models.seat import Seat
from models.scheduler import (
    ShiftScheduler,
    check_all_rules,
    format_warnings,
    get_exclusions,
    ScheduleData
)

# 脚本目录
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))


class PersonEditDialog(QDialog):
    """人员编辑/添加对话框"""
    def __init__(self, parent=None, person_data=None):
        super().__init__(parent)
        self.person_data = person_data
        self.setup_ui()
        
    def setup_ui(self):
        self.setWindowTitle("编辑人员信息" if self.person_data else "添加人员")
        self.setMinimumWidth(350)
        self.setStyleSheet("""
            QLineEdit, QComboBox {
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 3px;
            }
        """)
        
        layout = QFormLayout(self)
        layout.setSpacing(12)
        
        self.name_edit = QLineEdit()
        if self.person_data:
            self.name_edit.setText(self.person_data['name'])
        layout.addRow("姓名 *", self.name_edit)
        
        self.level_combo = QComboBox()
        self.level_combo.addItems(["C1", "C2", "C3", "I", "S"])
        if self.person_data:
            self.level_combo.setCurrentText(self.person_data['level'])
        self.level_combo.currentTextChanged.connect(self.on_level_changed)
        layout.addRow("等级 *", self.level_combo)
        
        score_layout = QHBoxLayout()
        self.score_spin = QDoubleSpinBox()
        self.score_spin.setRange(0.1, 10.0)
        self.score_spin.setSingleStep(0.1)
        self.score_spin.setDecimals(1)
        level_scores = {"C1": 1, "C2": 2, "C3": 3, "I": 3, "S": 3}
        if self.person_data:
            self.score_spin.setValue(float(self.person_data['score']))
            self._user_edited_score = True  # 标记用户已编辑，避免on_level_changed覆盖
        else:
            self.score_spin.setValue(float(level_scores.get(self.level_combo.currentText(), 1)))
            self._user_edited_score = False
        self.score_spin.setFixedWidth(80)
        score_layout.addWidget(self.score_spin)
        score_layout.addWidget(QLabel("分"))
        score_layout.addStretch()
        layout.addRow("分数 *", score_layout)
        
        self.active_checkbox = QCheckBox("参与排班")
        self.active_checkbox.setChecked(True)
        if self.person_data:
            self.active_checkbox.setChecked(bool(self.person_data['active']))
        layout.addRow("", self.active_checkbox)
        
        self.on_level_changed(self.level_combo.currentText())
        
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        buttons.accepted.connect(self.validate_input)
        buttons.rejected.connect(self.reject)
        layout.addRow("", buttons)
    
    def on_level_changed(self, level):
        level_scores = {"C1": 1, "C2": 2, "C3": 3, "I": 3, "S": 3}
        default_score = level_scores.get(level, 1)
        if not hasattr(self, '_user_edited_score') or not self._user_edited_score:
            self.score_spin.setValue(float(default_score))
    
    def validate_input(self):
        name = self.name_edit.text().strip()
        if not name:
            CustomMessageBox.warning(self, "输入错误", "请输入姓名")
            return
        if len(name) < 2:
            CustomMessageBox.warning(self, "输入错误", "姓名至少需要2个字符")
            return
        if len(name) > 20:
            CustomMessageBox.warning(self, "输入错误", "姓名不能超过20个字符")
            return
        if not re.match(r'^[\u4e00-\u9fa5a-zA-Z0-9_]+$', name):
            CustomMessageBox.warning(self, "输入错误", "姓名只能包含中英文、数字和下划线")
            return
        score = self.score_spin.value()
        if score < 0.1 or score > 10.0:
            CustomMessageBox.warning(self, "输入错误", "分数必须在0.1-10.0之间")
            return
        self.accept()
    
    def get_data(self):
        return {
            'name': self.name_edit.text().strip(),
            'level': self.level_combo.currentText(),
            'score': self.score_spin.value(),
            'active': 1 if self.active_checkbox.isChecked() else 0
        }


class PersonDetailDialog(QDialog):
    """人员详情对话框"""
    def __init__(self, parent=None, person_data=None):
        super().__init__(parent)
        self.person_data = person_data
        self.setup_ui()
    
    def setup_ui(self):
        self.setWindowTitle("人员详细信息")
        self.setMinimumSize(350, 300)
        layout = QVBoxLayout(self)
        
        info_text = QTextEdit()
        info_text.setReadOnly(True)
        info_text.setStyleSheet("background-color: #21262d; color: #c9d1d9;")
        
        p = self.person_data
        status = "✓ 参与排班" if p['active'] else "✗ 不参与排班"
        level_color = "#e74c3c" if p['level'] == "C1" else ("#f39c12" if p['level'] == "C2" else "#27ae60")
        
        html = f"""
        <h2 style="color: {level_color};">{p['name']}</h2>
        <p><b>等级：</b><span style="color: {level_color};">{p['level']}</span></p>
        <p><b>分数：</b><span style="color: #1565c0; font-weight: bold;">{p['score']}分</span></p>
        <p><b>状态：</b>{status}</p>
        <hr>
        <p><small>添加时间：{p.get('created_at', '未知')}</small></p>
        <p><small>更新时间：{p.get('updated_at', '未知')}</small></p>
        """
        info_text.setHtml(html)
        layout.addWidget(info_text)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        close_btn = QPushButton("关闭")
        close_btn.setStyleSheet("background-color: #3498db; color: white; padding: 6px 20px; border: none; border-radius: 4px;")
        close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)


class CustomMessageBox(QDialog):
    """自定义消息框"""
    def __init__(self, parent=None, title="", message="", icon_type="info"):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumWidth(350)
        self.icon_type = icon_type
        self.setup_ui(message)

    def setup_ui(self, message):
        self.setStyleSheet("""
            QDialog { background-color: #0d1117; color: #c9d1d9; }
            QLabel { color: #c9d1d9; }
        """)
        
        layout = QVBoxLayout(self)
        layout.setSpacing(15)
        
        h_layout = QHBoxLayout()
        icon_label = QLabel()
        if self.icon_type == "warning":
            icon_label.setText("⚠️")
        elif self.icon_type == "question":
            icon_label.setText("❓")
        else:
            icon_label.setText("✅")
        icon_label.setStyleSheet("font-size: 32px;")
        icon_label.setFixedWidth(50)
        h_layout.addWidget(icon_label)
        
        msg_label = QLabel(message)
        msg_label.setStyleSheet("color: #c9d1d9; font-size: 14px;")
        msg_label.setWordWrap(True)
        h_layout.addWidget(msg_label)
        layout.addLayout(h_layout)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        if self.icon_type == "question":
            self.yes_btn = QPushButton("是")
            self.yes_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 8px 25px; border: none; border-radius: 4px; font-weight: bold;")
            self.yes_btn.clicked.connect(self.accept)
            self.no_btn = QPushButton("否")
            self.no_btn.setStyleSheet("background-color: #95a5a6; color: white; padding: 8px 25px; border: none; border-radius: 4px;")
            self.no_btn.clicked.connect(self.reject)
            btn_layout.addWidget(self.yes_btn)
            btn_layout.addWidget(self.no_btn)
        else:
            self.ok_btn = QPushButton("确定")
            self.ok_btn.setStyleSheet("background-color: #3498db; color: white; padding: 8px 25px; border: none; border-radius: 4px; font-weight: bold;")
            self.ok_btn.clicked.connect(self.accept)
            btn_layout.addWidget(self.ok_btn)
        
        layout.addLayout(btn_layout)
    
    def exec_(self):
        return QDialog.exec_(self)
    
    @staticmethod
    def information(parent, title, message):
        dialog = CustomMessageBox(parent, title, message, "info")
        dialog.exec_()
    
    @staticmethod
    def warning(parent, title, message):
        dialog = CustomMessageBox(parent, title, message, "warning")
        dialog.exec_()
    
    @staticmethod
    def question(parent, title, message):
        dialog = CustomMessageBox(parent, title, message, "question")
        result = dialog.exec_()
        return result == QDialog.Accepted


class RuleEditDialog(QDialog):
    """规则编辑对话框"""
    def __init__(self, parent=None, rule_data=None):
        super().__init__(parent)
        self.rule_data = rule_data
        self.setup_ui()

    def setup_ui(self):
        self.setWindowTitle("编辑规则" if self.rule_data else "添加规则")
        self.setMinimumWidth(350)
        self.setStyleSheet("""
            QDialog { background-color: #161b22; }
            QLabel { color: #c9d1d9; font-size: 12px; }
            QLineEdit { padding: 8px; border: 1px solid #30363d; border-radius: 5px; background: #21262d; color: #c9d1d9; }
            QSpinBox { padding: 8px; border: 1px solid #30363d; border-radius: 5px; background: #21262d; color: #c9d1d9; }
            QCheckBox { padding: 8px; color: #c9d1d9; }
        """)

        layout = QFormLayout(self)
        layout.setSpacing(15)

        # 规则名称
        self.name_edit = QLineEdit()
        self.name_edit.setText(self.rule_data.get('name', '') if self.rule_data else '')
        layout.addRow("规则名称", self.name_edit)

        # 描述
        self.desc_edit = QLineEdit()
        self.desc_edit.setText(self.rule_data.get('description', '') if self.rule_data else '')
        layout.addRow("描述", self.desc_edit)

        # 分数修改
        self.score_spin = QSpinBox()
        self.score_spin.setRange(-10, 10)
        self.score_spin.setValue(int(self.rule_data.get('score_modifier', 0)) if self.rule_data else 0)
        layout.addRow("分数修改", self.score_spin)

        # 启用状态
        self.active_check = QCheckBox("启用规则")
        self.active_check.setChecked(bool(self.rule_data.get('active', 1)) if self.rule_data else True)
        layout.addRow("", self.active_check)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        ok_btn = QPushButton("确定")
        ok_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 8px 25px; border: none; border-radius: 4px;")
        ok_btn.clicked.connect(self.accept)
        btn_layout.addWidget(ok_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet("background-color: #7f8c8d; color: white; padding: 8px 25px; border: none; border-radius: 4px;")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addRow("", btn_layout)

    def get_data(self):
        return {
            'name': self.name_edit.text(),
            'description': self.desc_edit.text(),
            'score_modifier': self.score_spin.value(),
            'active': 1 if self.active_check.isChecked() else 0
        }


class SeatEditDialog(QDialog):
    """席位编辑对话框"""
    def __init__(self, parent=None, seat_data=None):
        super().__init__(parent)
        self.seat_data = seat_data
        self.setup_ui()
    
    def setup_ui(self):
        self.setWindowTitle("编辑席位")
        self.setMinimumWidth(350)
        self.setStyleSheet("""
            QDialog { background-color: #161b22; }
            QLabel { color: #c9d1d9; font-size: 12px; }
            QSpinBox { padding: 8px; border: 1px solid #30363d; border-radius: 5px; background: #21262d; color: #c9d1d9; }
            QComboBox { padding: 8px; border: 1px solid #30363d; border-radius: 5px; background: #21262d; color: #c9d1d9; }
        """)
        
        layout = QFormLayout(self)
        layout.setSpacing(15)
        
        # 席位名称（只读）
        self.app_label = QLabel(seat_data.get('app_name', '') if seat_data else '')
        self.app_label.setStyleSheet("color: #58a6ff; font-weight: bold;")
        layout.addRow("席位名称", self.app_label)
        
        # 人数
        self.count_spin = QSpinBox()
        self.count_spin.setRange(2, 6)
        self.count_spin.setValue(seat_data.get('persons_count', 3) if seat_data else 3)
        layout.addRow("人数", self.count_spin)
        
        # 要求分数
        self.score_spin = QSpinBox()
        self.score_spin.setRange(4, 20)
        self.score_spin.setValue(seat_data.get('required_score', 5) if seat_data else 5)
        layout.addRow("要求分数", self.score_spin)
        
        # 可用状态
        self.available_check = QCheckBox("启用席位")
        self.available_check.setChecked(bool(seat_data.get('available', 1)) if seat_data else True)
        layout.addRow("", self.available_check)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 8px 25px; border: none; border-radius: 4px;")
        ok_btn.clicked.connect(self.accept)
        btn_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet("background-color: #7f8c8d; color: white; padding: 8px 25px; border: none; border-radius: 4px;")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        layout.addRow("", btn_layout)
    
    def get_data(self):
        return {
            'available': 1 if self.available_check.isChecked() else 0,
            'persons_count': self.count_spin.value(),
            'required_score': self.score_spin.value()
        }


class TemplateEditDialog(QDialog):
    """席位模版编辑对话框"""
    def __init__(self, parent=None, template_data=None):
        super().__init__(parent)
        self.template_data = template_data
        self._closed = False
        self._build_ui()

    def done(self, r):
        if not self._closed:
            self._closed = True
            super().done(r)

    def _build_ui(self):
        self.setWindowTitle("编辑模版" if self.template_data else "添加模版")
        self.setMinimumWidth(350)
        self.setStyleSheet("""
            QDialog { background-color: #161b22; }
            QLabel { color: #c9d1d9; font-size: 12px; }
            QLineEdit { padding: 8px; border: 1px solid #30363d; border-radius: 5px; background: #21262d; color: #c9d1d9; }
            QSpinBox { padding: 8px; border: 1px solid #30363d; border-radius: 5px; background: #21262d; color: #c9d1d9; }
            QComboBox { padding: 8px; border: 1px solid #30363d; border-radius: 5px; background: #21262d; color: #c9d1d9; }
            QCheckBox { padding: 8px; color: #c9d1d9; }
        """)
        
        layout = QFormLayout(self)
        layout.setSpacing(15)
        
        self.id_edit = QLineEdit()
        if self.template_data:
            self.id_edit.setText(str(self.template_data.get('id', '')))
            self.id_edit.setReadOnly(True)
            self.id_edit.setStyleSheet("background-color: #161b22; color: #58a6ff;")
        else:
            parent = self.parent()
            if parent and hasattr(parent, 'db'):
                next_id = parent._get_next_template_id()
                self.id_edit.setText(str(next_id))
                self.id_edit.setReadOnly(True)
                self.id_edit.setStyleSheet("background-color: #161b22; color: #58a6ff;")
        layout.addRow("ID", self.id_edit)

        self.name_edit = QLineEdit()
        if self.template_data:
            self.name_edit.setText(self.template_data.get('name', ''))
        layout.addRow("名称 *", self.name_edit)

        self.desc_edit = QLineEdit()
        if self.template_data:
            self.desc_edit.setText(self.template_data.get('description', ''))
        layout.addRow("描述", self.desc_edit)

        self.available_check = QCheckBox("启用此模版")
        self.available_check.setChecked(bool(self.template_data.get('available', 1)) if self.template_data else True)
        layout.addRow("", self.available_check)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        self.ok_btn = QPushButton("确定")
        self.ok_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 8px 25px; border: none; border-radius: 4px;")
        self.ok_btn.clicked.connect(self.accept)
        btn_layout.addWidget(self.ok_btn)
        
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.setStyleSheet("background-color: #7f8c8d; color: white; padding: 8px 25px; border: none; border-radius: 4px;")
        self.cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(self.cancel_btn)
        
        layout.addRow("", btn_layout)
    
    def accept(self):
        if not self.name_edit.text():
            CustomMessageBox.warning(self, "错误", "模版名称不能为空")
            return
        super().accept()
    
    def get_data(self):
        return {
            'id': self.id_edit.text(),
            'name': self.name_edit.text(),
            'description': self.desc_edit.text(),
            'available': 1 if self.available_check.isChecked() else 0
        }


class PersonRuleSelectDialog(QDialog):
    """人员规则选择对话框"""
    def __init__(self, parent=None, person_id=None, db=None):
        super().__init__(parent)
        self.person_id = person_id
        self.db = db
        self.setup_ui()
        self.load_rules()
    
    def setup_ui(self):
        self.setWindowTitle("选择生效规则")
        self.setMinimumWidth(450)
        self.setStyleSheet("""
            QDialog { background-color: #161b22; }
            QLabel { color: #c9d1d9; font-size: 14px; }
            QCheckBox { color: #c9d1d9; font-size: 12px; padding: 5px; }
            QCheckBox::indicator:checked { background-color: #238636; border-color: #238636; }
            QCheckBox::indicator:unchecked { background-color: #21262d; border: 1px solid #30363d; }
        """)
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel("勾选要生效的规则（分数将根据规则计算）"))
        
        self.rule_widget = QWidget()
        self.rule_widget.setStyleSheet("background-color: #21262d; border-radius: 5px;")
        self.rule_layout = QVBoxLayout(self.rule_widget)
        self.rule_layout.setSpacing(5)
        self.rule_layout.setContentsMargins(5, 5, 5, 5)
        self.checkboxes = {}
        
        scroll = QScrollArea()
        scroll.setWidget(self.rule_widget)
        scroll.setWidgetResizable(True)
        scroll.setMinimumHeight(200)
        layout.addWidget(scroll)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        ok_btn = QPushButton("确定")
        ok_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 8px 25px; border: none; border-radius: 4px;")
        ok_btn.clicked.connect(self.save_and_close)
        btn_layout.addWidget(ok_btn)
        
        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet("background-color: #7f8c8d; color: white; padding: 8px 25px; border: none; border-radius: 4px;")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        layout.addLayout(btn_layout)
    
    def load_rules(self):
        all_rules = self.db.get_active_rules()
        person_rules = self.db.get_person_rules(self.person_id)
        
        rule_status = {}
        for r in person_rules:
            r_dict = dict(r)
            rule_status[r_dict['id']] = r_dict.get('enabled', 0) == 1
        
        for rule in all_rules:
            rule_dict = dict(rule)
            desc = rule_dict.get('description', '') or ''
            cb = QCheckBox(f"{rule_dict['name']} ({rule_dict['score_modifier']:+.2f}分) - {desc}")
            cb.setChecked(rule_status.get(rule_dict['id'], False))
            self.checkboxes[rule_dict['id']] = cb
            self.rule_layout.addWidget(cb)
    
    def save_and_close(self):
        for rule_id, cb in self.checkboxes.items():
            self.db.set_person_rule(self.person_id, rule_id, 1 if cb.isChecked() else 0)
        self.accept()


class DutyEditDialog(QDialog):
    """执勤数据编辑对话框"""
    def __init__(self, person, seat, current_count, parent=None):
        super().__init__(parent)
        self.person = person
        self.seat = seat
        self.current_count = current_count
        self.setup_ui()
    
    def setup_ui(self):
        self.setWindowTitle(f"编辑 {self.person['name']} 在 {self.seat['app_name']} 的执勤数据")
        self.setMinimumWidth(400)
        self.setStyleSheet("""
            QDialog { background-color: #161b22; }
            QLabel { color: #c9d1d9; font-size: 14px; }
            QLineEdit, QSpinBox { padding: 8px; border: 1px solid #30363d; border-radius: 4px; font-size: 14px; background: #21262d; color: #c9d1d9; }
            QComboBox { padding: 8px; border: 1px solid #30363d; border-radius: 4px; font-size: 14px; background: #21262d; color: #c9d1d9; }
        """)
        
        layout = QVBoxLayout(self)
        layout.addWidget(QLabel(f"人员：{self.person['name']} ({self.person['level']})"))
        layout.addWidget(QLabel(f"席位：{self.seat['app_name']}"))  # 显示物理席位APP编号
        
        count_layout = QHBoxLayout()
        count_layout.addWidget(QLabel("执勤次数："))
        self.count_input = QSpinBox()
        self.count_input.setRange(0, 9999)
        self.count_input.setValue(int(self.current_count) if str(self.current_count).isdigit() else 0)
        count_layout.addWidget(self.count_input)
        layout.addLayout(count_layout)
        
        last_seat_layout = QHBoxLayout()
        last_seat_layout.addWidget(QLabel("上次席位："))
        self.last_seat_combo = QComboBox()
        self.last_seat_combo.addItem("无", None)
        seats = self.parent().db.get_all_seats() if self.parent() else []
        for s in seats:
            self.last_seat_combo.addItem(f"{s['app_name']}", s['app_name'])  # 使用app_name作为userData
        if dict(self.person).get('last_seat_app_name'):
            idx = self.last_seat_combo.findData(self.person['last_seat_app_name'])
            if idx >= 0:
                self.last_seat_combo.setCurrentIndex(idx)
        last_seat_layout.addWidget(self.last_seat_combo)
        layout.addLayout(last_seat_layout)
        
        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("确定")
        ok_btn.setStyleSheet("background-color: #2ecc71; color: white; padding: 10px 30px; border: none; border-radius: 5px;")
        ok_btn.clicked.connect(self.accept)
        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet("background-color: #95a5a6; color: white; padding: 10px 30px; border: none; border-radius: 5px;")
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(ok_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
    
    def get_count(self):
        return self.count_input.value()
    
    def get_last_seat(self):
        return self.last_seat_combo.currentData()


class CenterAlignDelegate(QStyledItemDelegate):
    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)
        option.displayAlignment = Qt.AlignCenter


class SeatWidget(QWidget):
    """单个席位设置控件"""
    available_changed = pyqtSignal(str, bool)
    count_changed = pyqtSignal(str, int)
    score_changed = pyqtSignal(str, int)
    app_name_changed = pyqtSignal(str, str)
    
    def __init__(self, seat_data, db=None):
        super().__init__()
        self.seat_id = seat_data['app_name']  # 使用app_name作为主键
        self.seat_data = seat_data
        self.db = db
        self.setup_ui(seat_data)
    
    def setup_ui(self, seat_data):
        self.setStyleSheet("""
            SeatWidget { 
                background-color: #161b22; 
                border: 2px solid #30363d; 
                border-radius: 10px; 
            }
            QComboBox { 
                background-color: #21262d; 
                color: #c9d1d9; 
                border: 1px solid #30363d; 
                padding: 5px; 
                border-radius: 5px; 
            }
            QComboBox QAbstractItemView { 
                background-color: #161b22; 
                color: #c9d1d9; 
                selection-background-color: #1f6feb; 
            }
            QSpinBox { 
                background-color: #21262d; 
                color: #c9d1d9; 
                border: 1px solid #30363d; 
                border-radius: 5px; 
            }
            QCheckBox { 
                color: #c9d1d9; 
            }
            QLabel { 
                color: #58a6ff; 
                font-weight: bold; 
            }
        """)
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        
        name_label = QLabel(seat_data.get('name', seat_data.get('app_name', '空')))
        name_label.setStyleSheet("font-weight: bold; font-size: 14px; min-width: 60px;")
        layout.addWidget(name_label)
        
        app_label = QLabel("  PS:")
        layout.addWidget(app_label)
        
        self.app_combo = QComboBox()
        
        # 获取可用模版，填充下拉框
        template_items = ["空"]
        if self.db:
            templates = self.db.get_all_templates()
            for t in templates:
                if t.get('available', 1):
                    template_items.append(f"{t['id']}:{t['name']}")
        
        self.app_combo.addItems(template_items)
        
        # 设置当前选中的模版
        current_template_id = seat_data.get('template_id')
        if current_template_id:
            # 查找对应的模版名称，注意ID类型可能不同
            if self.db:
                templates = self.db.get_all_templates()
                current_str = str(current_template_id)
                found = False
                for t in templates:
                    if str(t['id']) == current_str:
                        self.app_combo.setCurrentText(f"{t['id']}:{t['name']}")
                        found = True
                        break
                if not found:
                    self.app_combo.setCurrentText("空")
        else:
            self.app_combo.setCurrentText("空")
        self.app_combo.setFixedWidth(120)
        self.app_combo.currentTextChanged.connect(lambda v: self.app_name_changed.emit(self.seat_id, v))
        layout.addWidget(self.app_combo)
        
        available_label = QLabel("  可用:")
        layout.addWidget(available_label)
        
        self.available_switch = QCheckBox()
        self.available_switch.setText("启用" if seat_data.get('available', 1) else "禁用")
        self.available_switch.setChecked(bool(seat_data.get('available', 1)))
        self.available_switch.stateChanged.connect(lambda state: self.on_available_changed(state))
        layout.addWidget(self.available_switch)
        
        count_label = QLabel("  人数:")
        layout.addWidget(count_label)
        
        self.count_spin = QSpinBox()
        self.count_spin.setRange(1, 6)
        self.count_spin.setValue(seat_data.get('persons_count', 3))
        self.count_spin.setFixedWidth(50)
        self.count_spin.valueChanged.connect(lambda v: self.count_changed.emit(self.seat_id, v))
        layout.addWidget(self.count_spin)
        
        score_label = QLabel("  分数:")
        layout.addWidget(score_label)
        
        self.score_combo = QComboBox()
        self.score_combo.addItems([str(i) for i in range(4, 13)])  # 4-12分
        self.score_combo.setCurrentText(str(seat_data.get('required_score', 5)))
        self.score_combo.setEditable(True)
        self.score_combo.setFixedWidth(60)
        self.score_combo.currentTextChanged.connect(lambda v: self.score_changed.emit(self.seat_id, int(v) if v.isdigit() else 0))
        layout.addWidget(self.score_combo)
        
        self.status_label = QLabel("🟢 已启用" if seat_data.get('available', 1) else "🔴 已禁用")
        self.status_label.setStyleSheet("margin-left: 10px;")
        layout.addWidget(self.status_label)
        
        layout.addStretch()
        self.update_style(seat_data.get('available', 1))
    
    def on_available_changed(self, state):
        available = state == Qt.Checked
        self.available_switch.setText("启用" if available else "禁用")
        self.status_label.setText("🟢 已启用" if available else "🔴 已禁用")
        self.available_changed.emit(self.seat_id, available)
        self.update_style(available)
    
    def update_style(self, available):
        if available:
            self.setStyleSheet("""
                SeatWidget { background-color: #161b22; border: 2px solid #238636; border-radius: 8px; }
            """)
            # 启用其他控件，但保留switch可点击
            for w in [self.app_combo, self.count_spin, self.score_combo]:
                w.setEnabled(True)
            self.available_switch.setChecked(True)
        else:
            self.setStyleSheet("""
                SeatWidget { background-color: #21262d; border: 2px solid #484f58; border-radius: 8px; }
            """)
            # 禁用其他控件
            for w in [self.app_combo, self.count_spin, self.score_combo]:
                w.setEnabled(False)
            self.available_switch.setChecked(False)

    def refresh_template_combo(self):
        """刷新模版下拉框（模版可用状态变更后调用）"""
        current_text = self.app_combo.currentText()

        # 重新构建可用模版列表
        template_items = ["空"]
        if self.db:
            templates = self.db.get_all_templates()
            for t in templates:
                if t.get('available', 1):
                    template_items.append(f"{t['id']}:{t['name']}")

        # 阻止信号防止触发app_name_changed
        self.app_combo.blockSignals(True)
        self.app_combo.clear()
        self.app_combo.addItems(template_items)

        # 恢复当前选中
        if current_text in template_items:
            self.app_combo.setCurrentText(current_text)
        else:
            self.app_combo.setCurrentText("空")
        self.app_combo.blockSignals(False)


class SeatResultWidget(QWidget):
    """排班结果显示控件"""
    seat_changed = pyqtSignal()
    
    def __init__(self, seat, parent=None):
        super().__init__(parent)
        self.seat = seat
        self.person_combos = []
        self.setup_ui()
    
    def refresh_ui(self):
        """从数据层刷新UI"""
        main_window = self.window()
        if hasattr(main_window, 'schedule_data') and main_window.schedule_data:
            seat_id_str = self.seat.app_name
            if seat_id_str in main_window.schedule_data.seats:
                available = main_window.schedule_data.seats[seat_id_str].get('available', True)
                self.set_available(available)
                self.update_count()
    
    def setup_ui(self):
        self.setMinimumHeight(55)
        self.setStyleSheet("""
            SeatResultWidget { background-color: #161b22; border: 2px solid #238636; border-radius: 8px; margin: 4px 0; }
        """)
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(15, 10, 15, 10)
        layout.setSpacing(15)
        
        left_layout = QVBoxLayout()
        left_layout.setSpacing(4)
        left_layout.setContentsMargins(0, 0, 0, 0)
        
        self.title_label = QLabel(f"{self.seat.app_name}/未安排")
        self.title_label.setStyleSheet("font-weight: bold; font-size: 16px; color: #58a6ff; background-color: transparent; padding-right: 80px;")
        left_layout.addWidget(self.title_label)

        # 删除独立的count_label，合并到title_label中
        
        layout.addLayout(left_layout)
        
        separator = QFrame()
        separator.setFrameShape(QFrame.VLine)
        separator.setFixedWidth(2)
        separator.setStyleSheet("color: #30363d;")
        layout.addWidget(separator)
        
        self.persons_container = QWidget()
        self.persons_layout = QHBoxLayout(self.persons_container)
        self.persons_layout.setSpacing(6)
        self.persons_layout.setContentsMargins(0, 0, 0, 0)
        
        # 固定生成所有下拉框，根据席位状态控制启用/禁用
        for i in range(self.seat.persons_count):
            combo = QComboBox()
            combo.setFixedWidth(100)
            # 可用席位：绿色边框；禁用席位：灰色边框
            if self.seat.available and self.seat.app_name and self.seat.app_name != "空":
                combo.setStyleSheet("""
                    QComboBox { padding: 5px 8px; border: 2px solid #238636; border-radius: 5px; background-color: #21262d; color: #c9d1d9; font-size: 14px; font-weight: bold; }
                    QComboBox:hover { border-color: #2ea043; }
                    QComboBox::drop-down { border: none; width: 20px; }
                    QComboBox QAbstractItemView { background-color: #161b22; color: #c9d1d9; selection-background-color: #58a6ff; }
                """)
                combo.setEnabled(True)
            else:
                combo.setStyleSheet("""
                    QComboBox { padding: 5px 8px; border: 2px solid #484f58; border-radius: 5px; background-color: #21262d; color: #484f58; font-size: 14px; font-weight: bold; }
                    QComboBox::drop-down { border: none; width: 20px; }
                    QComboBox QAbstractItemView { background-color: #161b22; color: #484f58; }
                """)
                combo.setEnabled(False)
            combo.currentTextChanged.connect(self.on_selection_changed)
            self.person_combos.append(combo)
            self.persons_layout.addWidget(combo)
        
        self.persons_layout.addStretch()
        # 始终添加人员容器，由setEnabled控制可用性
        layout.addWidget(self.persons_container)
    
    def on_selection_changed(self, text):
        # 更新样式
        combo = self.sender()
        if combo:
            has_selection = text and text != "未选择"
            self._set_combo_style(combo, has_selection)
        
        self.update_count()
        self.seat_changed.emit()
        # 边框样式由 MainWindow 统一刷新，不再单独检查
    
    def set_available(self, available):
        """根据席位可用状态更新UI（被动刷新，从数据层读取）"""
        # 先设置combo的禁用状态
        for combo in self.person_combos:
            combo.blockSignals(True)
            if available:
                combo.setStyleSheet("""
                    QComboBox { padding: 5px 8px; border: 2px solid #238636; border-radius: 5px; background-color: #21262d; color: #c9d1d9; font-size: 14px; font-weight: bold; }
                    QComboBox:hover { border-color: #2ea043; }
                    QComboBox::drop-down { border: none; width: 20px; }
                    QComboBox QAbstractItemView { background-color: #161b22; color: #c9d1d9; selection-background-color: #58a6ff; }
                """)
                combo.setEnabled(True)
            else:
                combo.setStyleSheet("""
                    QComboBox { padding: 5px 8px; border: 2px solid #484f58; border-radius: 5px; background-color: #21262d; color: #484f58; font-size: 14px; font-weight: bold; }
                    QComboBox::drop-down { border: none; width: 20px; }
                    QComboBox QAbstractItemView { background-color: #161b22; color: #484f58; }
                """)
                combo.setEnabled(False)
                combo.setCurrentText("")  # 清空选择
            combo.blockSignals(False)
        
        # 更新整个widget背景色：禁用时变灰
        if available:
            self.setStyleSheet("""
                SeatResultWidget { 
                    background-color: #161b22; 
                    border: 2px solid #30363d; 
                    border-radius: 10px; 
                }
            """)
        else:
            self.setStyleSheet("""
                SeatResultWidget { 
                    background-color: #21262d; 
                    border: 2px solid #484f58; 
                    border-radius: 10px; 
                }
            """)
        
        # 更新人数显示
        self.update_count()
    
    def _check_and_update_border_style(self):
        """检查规则，不满足时边框变红，返回违规信息"""
        # 临时阻塞信号避免循环
        for combo in self.person_combos:
            combo.blockSignals(True)
        
        self._check_and_update_border_style_core()
        
        # 恢复信号
        for combo in self.person_combos:
            combo.blockSignals(False)
    
    def _check_and_update_border_style_noemit(self):
        """检查规则（不阻塞信号，用于恢复选择后）"""
        self._check_and_update_border_style_core()
    
    def _check_and_update_border_style_core(self):
        """从数据层检查规则并刷新UI"""
        from models.person import Person as PersonModel
        from models.scheduler import check_c1_c2_rules
        
        # 先检查席位是否可用，如果不可用则使用禁用样式
        main_window = self.window()
        seat_id_str = self.seat.app_name
        if hasattr(main_window, 'schedule_data') and main_window.schedule_data:
            available = main_window.schedule_data.seats.get(seat_id_str, {}).get('available', True)
            if not available:
                # 席位被禁用，使用禁用样式，不检查其他规则
                self.setStyleSheet("""
                    SeatResultWidget { 
                        background-color: #21262d; 
                        border: 2px solid #484f58; 
                        border-radius: 10px; 
                    }
                """)
                for combo in self.person_combos:
                    combo.setStyleSheet("""
                        QComboBox { padding: 5px 8px; border: 2px solid #484f58; border-radius: 5px; background-color: #21262d; color: #484f58; font-size: 14px; font-weight: bold; }
                        QComboBox::drop-down { border: none; width: 20px; }
                        QComboBox QAbstractItemView { background-color: #161b22; color: #484f58; }
                    """)
                    combo.setEnabled(False)
                return
        
        # 从数据层读取选择
        seat_id_str = self.seat.app_name
        main_window = self.window()
        
        selections = []
        if hasattr(main_window, 'schedule_data') and main_window.schedule_data:
            sd = main_window.schedule_data
            if seat_id_str in sd.selections:
                selections = sd.selections[seat_id_str]
        
        # 收集人员信息
        selected_people = []
        persons_map = {}
        if hasattr(main_window, 'schedule_data') and main_window.schedule_data:
            for pid, pdata in main_window.schedule_data.persons.items():
                persons_map[pdata['name']] = pdata
        
        for i, name in enumerate(selections):
            if name and name in persons_map:
                pdata = persons_map[name]
                level = pdata.get('level', '')
                score = pdata.get('score', 0)
                score_modifier = pdata.get('score_modifier', 0)
                effective = pdata.get('effective_score', score + score_modifier)
                selected_people.append({
                    'name': name,
                    'level': level,
                    'score': score,
                    'score_modifier': score_modifier,
                    'effective_score': effective,
                    'index': i
                })
        
        # 检查规则
        violation_msg = None
        
        # 4人席位：分数规则分别检查ABC和BCD组合
        if self.seat.persons_count == 4 and len(selected_people) >= 3 and self.seat.required_score > 0:
            # ABC组合（位置0,1,2）
            abc_people = [p for p in selected_people if p['index'] in [0, 1, 2]]
            if len(abc_people) >= 3:
                abc_score = sum(p['effective_score'] for p in abc_people)
                if abc_score < self.seat.required_score:
                    violation_msg = f"{self.seat.app_name}: ABC组合分数不足({self.seat.required_score}分)"
            
            # BCD组合（位置1,2,3）
            if not violation_msg:
                bcd_people = [p for p in selected_people if p['index'] in [1, 2, 3]]
                if len(bcd_people) >= 3:
                    bcd_score = sum(p['effective_score'] for p in bcd_people)
                    if bcd_score < self.seat.required_score:
                        violation_msg = f"{self.seat.app_name}: BCD组合分数不足({self.seat.required_score}分)"
        elif len(selected_people) >= 3 and self.seat.required_score > 0:
            # 非4人席位：直接检查总分
            total_score = sum(p['effective_score'] for p in selected_people)
            if total_score < self.seat.required_score:
                violation_msg = f"{self.seat.app_name}: 人员分数不足({self.seat.required_score}分)"
        
        # 检查C1/C2时间重叠规则
        # 使用统一规则检查函数，只针对当前席位
        if not violation_msg and hasattr(main_window, 'template_slots'):
            # 构造当前席位的selections字典
            seat_selections = {seat_id_str: selections}
            # 构造seats字典
            seats_dict = {seat_id_str: {'name': self.seat.app_name, 'count': self.seat.persons_count}}
            # 构造persons字典
            persons_dict = {}
            for p in selected_people:
                persons_dict[p['name']] = {'name': p['name'], 'level': p['level'], 'score': p.get('score', 0), 'score_modifier': p.get('score_modifier', 0), 'effective_score': p['effective_score']}
            
            # 调用统一规则检查
            template_slots = main_window.template_slots
            warnings = check_c1_c2_rules(seat_selections, seats_dict, persons_dict, template_slots)
            if warnings:
                violation_msg = warnings[0].message
        
        # 检查重复选择（跨席位）
        if not violation_msg:
            # 获取所有席位的选择
            all_selections = []
            if hasattr(main_window, 'schedule_data') and main_window.schedule_data:
                for sid, snames in main_window.schedule_data.selections.items():
                    all_selections.extend([n for n in snames if n])
            
            from collections import Counter
            counts = Counter(all_selections)
            duplicates = [name for name, count in counts.items() if count > 1]
            if duplicates:
                violation_msg = f"人员重复选择: {', '.join(duplicates)}"
        
        # 检查C1/C2列限制
        if not violation_msg:
            from models.scheduler import check_c1c2_column_limit
            if hasattr(main_window, 'schedule_data') and main_window.schedule_data:
                column_warnings = check_c1c2_column_limit(
                    main_window.schedule_data.selections,
                    main_window.schedule_data.seats,
                    main_window.schedule_data.persons
                )
                if column_warnings:
                    violation_msg = column_warnings[0].message
        
        # 根据结果刷新UI
        if violation_msg:
            # 违规：红色边框
            for i in range(len(self.person_combos)):
                if i < len(selections) and selections[i]:
                    self.person_combos[i].setStyleSheet("""
                        QComboBox { padding: 5px 8px; border: 2px solid #e74c3c; border-radius: 5px; background-color: #21262d; color: #c9d1d9; font-size: 14px; font-weight: bold; }
                        QComboBox:hover { border: 2px solid #ff6b6b; }
                        QComboBox::drop-down { border: none; width: 20px; }
                        QComboBox QAbstractItemView { background-color: #161b22; color: #c9d1d9; selection-background-color: #e74c3c; border: 1px solid #30363d; }
                    """)
        else:
            # 正常：蓝色边框
            for i in range(len(self.person_combos)):
                has_sel = i < len(selections) and selections[i]
                self.person_combos[i].setStyleSheet("""
                    QComboBox { padding: 5px 8px; border: 2px solid #58a6ff; border-radius: 5px; background-color: #21262d; color: #c9d1d9; font-size: 14px; font-weight: bold; }
                    QComboBox:hover { border: 2px solid #79c0ff; }
                    QComboBox::drop-down { border: none; width: 20px; }
                    QComboBox QAbstractItemView { background-color: #161b22; color: #c9d1d9; selection-background-color: #58a6ff; border: 1px solid #30363d; }
""")
    
    def _set_all_combos_style(self, style_type):
        """统一设置所有下拉框的边框样式（red/blue）"""
        for combo in self.person_combos:
            if style_type == 'red':
                combo.setStyleSheet("""
                    QComboBox { padding: 5px 8px; border: 2px solid #e74c3c; border-radius: 5px; background-color: #21262d; color: #c9d1d9; font-size: 14px; font-weight: bold; }
                    QComboBox:hover { border: 2px solid #ff6b6b; }
                    QComboBox::drop-down { border: none; width: 20px; }
                    QComboBox QAbstractItemView { background-color: #161b22; color: #c9d1d9; selection-background-color: #e74c3c; border: 1px solid #30363d; }
                """)
            else:  # blue
                combo.setStyleSheet("""
                    QComboBox { padding: 5px 8px; border: 2px solid #58a6ff; border-radius: 5px; background-color: #21262d; color: #c9d1d9; font-size: 14px; font-weight: bold; }
                    QComboBox:hover { border: 2px solid #79c0ff; }
                    QComboBox::drop-down { border: none; width: 20px; }
                    QComboBox QAbstractItemView { background-color: #161b22; color: #c9d1d9; selection-background-color: #58a6ff; border: 1px solid #30363d; }
                """)
    
    def update_count(self):
        # 从数据层获取选择
        seat_id_str = self.seat.app_name
        main_window = self.window()

        selections = []
        if hasattr(main_window, 'schedule_data') and main_window.schedule_data:
            sd = main_window.schedule_data
            if seat_id_str in sd.selections:
                selections = sd.selections[seat_id_str]
        
        selected = sum(1 for name in selections if name)
        total = self.seat.persons_count
        
        # 合并到title_label中显示
        self.title_label.setText(f"{seat_id_str}/{selected}/{total}人")

        # 检查分数是否满足要求（从数据层读取）
        # 只有当人员数>=3时才检查分数限制
        if selected >= 3:
            total_score = 0
            persons_map = {}
            if hasattr(main_window, 'schedule_data') and main_window.schedule_data:
                for pid, pdata in main_window.schedule_data.persons.items():
                    persons_map[pdata['name']] = pdata
            
            for name in selections:
                if name and name in persons_map:
                    pdata = persons_map[name]
                    total_score += pdata.get('effective_score', pdata.get('score', 0))
            
            if total_score < self.seat.required_score:
                self.title_label.setStyleSheet("font-weight: bold; font-size: 16px; color: #e74c3c; background-color: transparent; padding-right: 80px;")
            else:
                self.title_label.setStyleSheet("font-weight: bold; font-size: 16px; color: #58a6ff; background-color: transparent; padding-right: 80px;")
        else:
            # 不足3人，恢复正常颜色
            self.title_label.setStyleSheet("font-weight: bold; font-size: 16px; color: #58a6ff; background-color: transparent; padding-right: 80px;")
    
    def update_persons_list(self, persons_list, excluded_names=None, current_combo_index=None):
        """更新下拉框中的人员列表
        excluded_names: 其他席位已选择的人员列表
        current_combo_index: 当前正在操作的下拉框索引（None表示批量更新）
        """
        if excluded_names is None:
            excluded_names = []

        # 保存当前选择（从 schedule_data 获取，保证数据源一致）
        current_selections = []
        max_combos = len(self.person_combos)
        seat_id_str = self.seat.app_name
        
        for i in range(max_combos):
            if hasattr(self, 'window') and self.window():
                sd = self.window().schedule_data
                if sd and sd.selections:
                    if seat_id_str in sd.selections and i < len(sd.selections[seat_id_str]):
                        current_selections.append(sd.selections[seat_id_str][i])
                    else:
                        current_selections.append(None)
                else:
                    current_selections.append(None)
            else:
                current_selections.append(None)
        
        # 阻止所有信号，避免循环触发
        for combo in self.person_combos:
            combo.blockSignals(True)
        
        # 重建所有下拉框
        for i, combo in enumerate(self.person_combos):
            combo.clear()
            combo.addItem("未选择", None)
            
            # 先添加被排除但当前席位选择的人员（保证恢复选择能找到）
            if current_selections[i]:
                name = current_selections[i]
                # 获取人员等级
                level = ''
                for p in persons_list:
                    try:
                        if p['name'] == name:
                            level = p['level']
                            break
                    except:
                        if p.name == name:
                            level = p.level
                            break
                if level:
                    combo.addItem(f"{name}({level})", name)
            
            for p in persons_list:
                # 支持字典、sqlite3.Row和对象三种格式
                try:
                    name = p['name']
                except:
                    name = p.name
                try:
                    level = p['level']
                except:
                    level = p.level
                
                # 跳过当前选择（已添加）
                if name == current_selections[i]:
                    continue
                
                # 规则1：排除其他席位的选择（跨席位互斥）
                if name in excluded_names:
                    continue
                
                combo.addItem(f"{name}({level})", name)
        
        # 恢复选择（此时信号仍被阻塞，不会触发）
        for i, selection in enumerate(current_selections):
            if selection:
                idx = self.person_combos[i].findData(selection)
                if idx >= 0:
                    self.person_combos[i].setCurrentIndex(idx)
        # 恢复后统一检查规则样式（会覆盖上面的 _set_combo_style）
        self._check_and_update_border_style_noemit()
        # 恢复完成后解除信号阻止
        for combo in self.person_combos:
            combo.blockSignals(False)
        
        self.update_count()
    
    def get_selected_persons(self):
        selected = []
        for combo in self.person_combos:
            name = combo.currentData()
            if name:
                selected.append(name)
        return selected
    
    def set_selected_persons(self, names):
        for combo in self.person_combos:
            combo.blockSignals(True)
        for i, name in enumerate(names):
            if i < len(self.person_combos):
                index = self.person_combos[i].findData(name)
                if index >= 0:
                    self.person_combos[i].setCurrentIndex(index)
        for combo in self.person_combos:
            combo.blockSignals(False)
        self.update_count()
    
    def clear_selection(self):
        for combo in self.person_combos:
            combo.blockSignals(True)
            combo.setCurrentIndex(0)
            combo.blockSignals(False)
        self.update_count()
    
    def set_available(self, available):
        self.seat.available = available
        if available:
            self.persons_container.show()
            for combo in self.person_combos:
                combo.setEnabled(True)
                self._set_combo_style(combo, combo.currentIndex() >= 0)
            # 启用时检查规则样式
            self._check_and_update_border_style()
        else:
            self.clear_selection()
            self.persons_container.show()  # 显示但不隐藏
            for combo in self.person_combos:
                combo.setEnabled(False)  # 不可选
                # 设置禁用状态样式
                combo.setStyleSheet("""
                    QComboBox { padding: 5px 8px; border: 2px solid #30363d; border-radius: 5px; background-color: #161b22; color: #484f58; font-size: 14px; font-weight: bold; }
                    QComboBox::drop-down { border: none; width: 20px; }
                    QComboBox QAbstractItemView { background-color: #161b22; color: #484f58; }
                """)
    
    def _set_combo_style(self, combo, has_selection):
        """根据是否有选择设置下拉框样式（暗黑科技风格）"""
        # 统一使用强调色边框样式，背景色保持一致
        combo.setStyleSheet("""
            QComboBox { padding: 5px 8px; border: 2px solid #58a6ff; border-radius: 5px; background-color: #21262d; color: #c9d1d9; font-size: 14px; font-weight: bold; }
            QComboBox:hover { border: 2px solid #79c0ff; }
            QComboBox::drop-down { border: none; width: 20px; }
            QComboBox QAbstractItemView { background-color: #161b22; color: #c9d1d9; selection-background-color: #58a6ff; border: 1px solid #30363d; }
        """)
    
    def recreate_combos(self, new_count):
        current_selections = self.get_selected_persons()
        
        while self.persons_layout.count():
            item = self.persons_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        self.person_combos = []
        self.seat.persons_count = new_count
        # count_label已合并到title_label中，这里不再需要单独更新
        
        for i in range(new_count):
            combo = QComboBox()
            combo.setFixedWidth(100)
            # 暗黑科技风格
            combo.setStyleSheet("""
                QComboBox { padding: 5px 8px; border: 2px solid #58a6ff; border-radius: 5px; background-color: #21262d; color: #c9d1d9; font-size: 14px; font-weight: bold; }
                QComboBox:hover { border: 2px solid #79c0ff; }
                QComboBox::drop-down { border: none; width: 20px; }
                QComboBox QAbstractItemView { background-color: #161b22; color: #c9d1d9; selection-background-color: #58a6ff; border: 1px solid #30363d; }
            """)
            combo.currentTextChanged.connect(self.on_selection_changed)
            self.person_combos.append(combo)
            self.persons_layout.addWidget(combo)
        
        self.persons_layout.addStretch()
        
        try:
            db = self.window().db
            if db:
                active_persons = db.get_active_persons()
                self.update_persons_list(active_persons)
        except:
            pass
        
        for i, name in enumerate(current_selections[:new_count]):
            idx = self.person_combos[i].findData(name)
            if idx >= 0:
                self.person_combos[i].blockSignals(True)
                self.person_combos[i].setCurrentIndex(idx)
                self.person_combos[i].blockSignals(False)
                self._set_combo_style(self.person_combos[i], True)
        
        # 确保所有下拉框都有正确的样式
        for i, combo in enumerate(self.person_combos):
            has_selection = i < len(current_selections) and current_selections[i] is not None
            self._set_combo_style(combo, has_selection)
        
        # 重建后检查规则样式
        self._check_and_update_border_style()


class MainWindow(QMainWindow):
    """主窗口"""
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.seat_widgets = []
        self.seat_result_widgets = []
        self.current_schedule = None
        
        # 初始化template_slots为空字典
        self.template_slots = {}
        
        # 数据层（来自 scheduler_rules）
        self.schedule_data = None
        self._init_schedule_data()
        
        # 安装事件过滤器禁用下拉框滚轮
        from PyQt5.QtWidgets import QApplication
        QApplication.instance().installEventFilter(self)
        
        self.setup_ui()
        self.load_data()
        self.init_schedule_combos()
        self.refresh_ui()  # 从数据层刷新UI
    
    def eventFilter(self, obj, event):
        """禁用下拉框的滚轮事件"""
        from PyQt5.QtCore import QEvent
        from PyQt5.QtWidgets import QComboBox
        if event.type() == QEvent.Wheel and isinstance(obj, QComboBox):
            obj.wheelEvent = lambda e: None
            return True
        return super().eventFilter(obj, event)
    
    def _init_schedule_data(self):
        """初始化数据层"""
        self.schedule_data = ScheduleData()
        self._load_schedule_data()
    
    def on_tab_changed(self, index):
        """Tab切换时刷新排班预览"""
        # 排班预览Tab的索引是3（根据添加顺序：规则设置0, 人员管理1, 排班设置2, 排班预览3, 执勤数据4）
        if index == 3:  # 排班预览
            self.refresh_preview()
    
    def _load_schedule_data(self):
        """加载席位信息（使用所有席位，确保与 UI widget 一致）"""
        seats = self.db.get_all_seats()
        for s in seats:
            self.schedule_data.seats[s['app_name']] = {
                'name': s['app_name'],
                'app': s.get('app_name', ''),
                'available': s.get('available', 1),
                'count': s.get('persons_count', 3),
                'required': s.get('required_score', 5),
                'template_id': s.get('template_id'),
            }
        
        # 构建所有席位的时段配置
        self._build_template_slots()
        
        # 加载人员信息
        persons = self.db.get_active_persons()
        for p in persons:
            score_mod = self.db.get_person_score_modifier(p['id'])
            self.schedule_data.persons[p['name']] = {
                'id': p['id'],
                'name': p['name'],
                'level': p['level'],
                'score': p['score'],
                'score_modifier': score_mod,
                'effective_score': p['score'] + score_mod
            }
    
    def init_schedule_combos(self):
        """初始化排班席位"""
        active_persons = self.db.get_active_persons()
        
        for widget in self.seat_result_widgets:
            widget.update_persons_list(active_persons, [], None)
        
        # 连接信号
        for widget in self.seat_result_widgets:
            for idx in range(len(widget.person_combos)):
                widget.person_combos[idx].currentTextChanged.connect(
                    lambda text, i=idx, w=widget: self.on_seat_selection_changed(w, i)
                )

    def on_seat_selection_changed(self, changed_widget, changed_combo_index):
        """下拉框变化 - 统一更新数据层后刷新UI"""
        # 1. 更新 schedule_data.selections（单一数据源）
        seat_id = changed_widget.seat.app_name
        seat_id_str = seat_id  # 直接使用，已经是字符串
        position = changed_combo_index
        new_name = changed_widget.person_combos[position].currentData()
        
        # 确保格式为 [A, B, C, D]（固定4个位置，None表示未选）
        if seat_id_str not in self.schedule_data.selections:
            self.schedule_data.selections[seat_id_str] = [None, None, None, None]
        
        # 更新对应位置
        self.schedule_data.selections[seat_id_str][position] = new_name
        
        # 2. 检查规则
        warnings = check_all_rules(
            self.schedule_data.selections,
            self.schedule_data.seats,
            self.schedule_data.persons,
            self.template_slots
        )
        
        # 3. 更新结果显示框
        if warnings:
            msg = format_warnings(warnings)
            self.schedule_result_display.setText(msg)
            self.schedule_result_display.setStyleSheet("""
                QTextEdit {
                    border: 2px solid #30363d;
                    border-radius: 6px;
                    background-color: #161b22;
                    color: #e74c3c;
                    padding: 8px;
                    font-size: 12px;
                }
            """)
        else:
            # 清除结果显示框
            self.schedule_result_display.clear()
            self.schedule_result_display.setStyleSheet("""
                QTextEdit {
                    border: 1px solid #30363d;
                    border-radius: 6px;
                    background-color: #161b22;
                    color: #c9d1d9;
                    padding: 8px;
                    font-size: 12px;
                }
            """)
        
        # 4. 刷新所有下拉框（排除已选人员，基于统一数据源）
        self._refresh_dropdowns()
        
        # 5. 统一刷新所有席位的边框样式
        self._refresh_all_border_styles(warnings)
        
        # 6. 刷新排班预览（更新显示的人员位置）
        self.refresh_preview()
    
    def _refresh_all_border_styles(self, warnings):
        """统一刷新所有席位的边框样式"""
        from models.scheduler import RuleType
        
        # 收集所有有违规的席位ID
        violated_seats = set()
        duplicate_names = set()
        
        for w in warnings:
            if w.rule_type == RuleType.DUPLICATE:
                # 提取重复的人员名
                import re
                matches = re.findall(r'[^,\s]+', w.message)
                for m in matches:
                    if m not in ('人员重复选择', '人员', '重复选择'):
                        duplicate_names.add(m)
            elif w.seat_id:
                violated_seats.add(w.seat_id)
        
        # 特殊处理列限制违规 - 列限制涉及多个席位
        # 如果有列限制违规，所有席位都需要变红
        has_column_violation = any(
            w.rule_type == RuleType.COLUMN_LIMIT for w in warnings
        )
        
        # 收集重复选择的人员名对应的席位
        duplicate_seats = set()
        if duplicate_names:
            for seat_id, names in self.schedule_data.selections.items():
                for name in names:
                    if name in duplicate_names:
                        duplicate_seats.add(seat_id)
        
        # 合并所有需要红色边框的席位
        all_violated_seats = violated_seats | duplicate_seats
        
        # 如果有列限制违规，所有席位都变红
        if has_column_violation:
            all_violated_seats = {w.seat.app_name for w in self.seat_result_widgets}
        
        # 更新所有席位的边框样式
        for widget in self.seat_result_widgets:
            seat_id_str = widget.seat.app_name
            
            # 先检查席位是否可用
            if seat_id_str in self.schedule_data.seats:
                available = self.schedule_data.seats[seat_id_str].get('available', True)
                if not available:
                    # 席位被禁用，使用禁用样式
                    widget.set_available(False)
                    continue
            
            if seat_id_str in all_violated_seats:
                # 违规：红色边框
                widget._set_all_combos_style('red')
            else:
                # 正常：蓝色边框
                widget._set_all_combos_style('blue')
    
    def _refresh_dropdowns(self):
        """刷新所有下拉框"""
        active_persons = self.db.get_active_persons()
        
        for widget in self.seat_result_widgets:
            if widget.seat.available:
                seat_id = widget.seat.app_name
                # 计算排除列表
                excluded = get_exclusions(
                    self.schedule_data.selections,
                    widget.seat.app_name
                )
                # 更新下拉框（排除已选人员）
                widget.update_persons_list(active_persons, excluded, None)
    
    def run_schedule(self):
        """开始排班"""
        self._run_schedule_original()
    
    def check_c1c2_column_limit(self):
        """检查ABCD每列的C1/C2数量是否超限"""
        # 统计3人及以上席位数（3人+4人席位都参与计算）
        active_seats = []
        for widget in self.seat_result_widgets:
            if widget.seat.available and widget.seat.persons_count >= 3:
                active_seats.append(widget)
        
        if not active_seats:
            return  # 没有3人及以上席位，无需检查
        
        active_seat_count = len(active_seats)
        max_c_count = math.floor(active_seat_count * 0.5)  # C1/C2上限，取整
        
        # 统计每列的C1/C2数量（按位置A/B/C/D）
        column_c_count = [0, 0, 0, 0]  # A,B,C,D列的C1/C2数量
        
        # 人员等级缓存
        person_levels = {}
        for person in self.db.get_active_persons():
            person_levels[person['name']] = person['level']
        
        for widget in active_seats:
            # 根据席位人数确定检查的位置数
            max_positions = min(4, widget.seat.persons_count)
            for i, combo in enumerate(widget.person_combos):
                if i < max_positions:
                    name = combo.currentData()
                    if name:
                        level = person_levels.get(name, '')
                        if level in ('C1', 'C2'):
                            column_c_count[i] += 1
        
        # 检查是否超限
        for i, count in enumerate(column_c_count):
            if count > max_c_count:
                position = ['A', 'B', 'C', 'D'][i]
                CustomMessageBox.warning(
                    self, "C1/C2限制提醒",
                    f"{position}列C1/C2数量为{count}人，超过了{active_seat_count}个3人及以上席位的50%限制（{max_c_count}人）"
                )
                return

    def setup_ui(self):
        self.setWindowTitle("排班管理系统")
        self.setMinimumSize(950, 750)
        self.setStyleSheet("""
            QMainWindow { background-color: #0d1117; }
            QLabel { color: #c9d1d9; }
            QMessageBox { background-color: #161b22; }
            QMessageBox QLabel { color: #c9d1d9; font-size: 14px; }
        """)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        
        tabs = QTabWidget()
        tabs.setStyleSheet("""
            QTabBar::tab {
                min-width: 120px;
                padding: 10px 20px;
                font-size: 14px;
                font-weight: bold;
                color: #8b949e;
                background-color: #21262d;
                border: 1px solid #30363d;
                border-bottom: none;
                margin-right: 2px;
            }
            QTabBar::tab:selected {
                color: #ffffff;
                background-color: #0d1117;
                border: 1px solid #58a6ff;
                border-bottom: none;
            }
            QTabBar::tab:hover {
                color: #c9d1d9;
                background-color: #30363d;
                border-color: #484f58;
            }
            QTabWidget::pane {
                border: 1px solid #58a6ff;
                background-color: #0d1117;
                margin-top: -1px;
            }
        """)
        tabs.addTab(self.create_rules_tab(), "规则设置")
        tabs.addTab(self.create_person_tab(), "人员管理")
        tabs.addTab(self.create_schedule_tab(), "排班设置")
        tabs.addTab(self.create_preview_tab(), "排班预览")
        tabs.addTab(self.create_duty_data_tab(), "执勤数据")

        # 切换到排班预览Tab时刷新
        tabs.currentChanged.connect(self.on_tab_changed)
        
        main_layout.addWidget(tabs)
        
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("就绪")
    
    def create_rules_tab(self):
        # 使用滚动区域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("""
            QScrollArea {
                border: none;
                background-color: #0d1117;
            }
            QScrollBar:vertical {
                background-color: #21262d;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background-color: #484f58;
                border-radius: 5px;
                min-height: 30px;
            }
            QScrollBar::handle:vertical:hover {
                background-color: #58a6ff;
            }
        """)
        
        widget = QWidget()
        widget.setStyleSheet("background-color: #0d1117;")
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        
        info_label = QLabel("规则设置：定义各种分数调整规则（如生病扣分、培训扣分等），然后在人员管理中为每个人员选择生效的规则")
        info_label.setStyleSheet("color: #8b949e; font-size: 12px; padding: 10px; background-color: #161b22; border-radius: 5px;")
        layout.addWidget(info_label)
        
        # ===== 席位模版设置部分 =====
        templates_title = QLabel("席位模版设置")
        templates_title.setStyleSheet("font-weight: bold; font-size: 16px; color: #58a6ff; padding: 10px 0;")
        layout.addWidget(templates_title)
        
        templates_toolbar = QHBoxLayout()
        
        add_tpl_btn = QPushButton("添加模版")
        add_tpl_btn.setStyleSheet("""
            QPushButton {
                background-color: #21262d;
                color: #3fb950;
                border: 1px solid #3fb950;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #30363d;
                border-color: #3fb950;
            }
        """)
        add_tpl_btn.clicked.connect(self.add_template)
        templates_toolbar.addWidget(add_tpl_btn)
        
        edit_tpl_btn = QPushButton("编辑模版")
        edit_tpl_btn.setStyleSheet("""
            QPushButton {
                background-color: #21262d;
                color: #58a6ff;
                border: 1px solid #58a6ff;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #30363d;
                border-color: #58a6ff;
            }
        """)
        preview_tpl_btn = QPushButton("预览模版")
        preview_tpl_btn.setStyleSheet("""
            QPushButton {
                background-color: #21262d;
                color: #f39c12;
                border: 1px solid #f39c12;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
                min-width: 80px;
            }
            QPushButton:hover {
                background-color: #30363d;
                border-color: #f39c12;
            }
        """)
        preview_tpl_btn.clicked.connect(self.preview_template)
        templates_toolbar.addWidget(preview_tpl_btn)
        
        delete_tpl_btn = QPushButton("删除模版")
        delete_tpl_btn.setStyleSheet("""
            QPushButton {
                background-color: #21262d;
                color: #f85149;
                border: 1px solid #f85149;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #30363d;
                border-color: #f85149;
            }
        """)
        delete_tpl_btn.clicked.connect(self.delete_template)
        templates_toolbar.addWidget(delete_tpl_btn)
        
        templates_toolbar.addStretch()
        layout.addLayout(templates_toolbar)
        
        self.templates_table = QTableWidget()
        self.templates_table.setColumnCount(4)
        self.templates_table.setHorizontalHeaderLabels(["ID", "名称", "描述", "可用状态"])
        self.templates_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.templates_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.templates_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.templates_table.setAlternatingRowColors(True)
        self.templates_table.setStyleSheet("""
            QTableWidget { border: 1px solid #30363d; background-color: #0d1117; color: #c9d1d9; }
            QTableWidget::item { padding: 5px; text-align: center; }
            QHeaderView::section { background: #21262d; font-weight: bold; padding: 6px; color: #58a6ff; }
        """)
        self.templates_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.templates_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.templates_table)
        
        self.load_templates()
        
        # 分隔线
        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setStyleSheet("background-color: #30363d; padding: 20px 0;")
        layout.addWidget(separator)
        
        # 分隔线
        separator = QFrame()
        separator.setFrameShape(QFrame.HLine)
        separator.setStyleSheet("background-color: #30363d; padding: 20px 0;")
        layout.addWidget(separator)
        
        # ===== 规则设置部分 =====
        rules_title = QLabel("规则设置")
        rules_title.setStyleSheet("font-weight: bold; font-size: 16px; color: #58a6ff; padding: 10px 0;")
        layout.addWidget(rules_title)
        
        toolbar = QHBoxLayout()
        
        add_btn = QPushButton("添加规则")
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #21262d;
                color: #3fb950;
                border: 1px solid #3fb950;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #30363d;
                border-color: #3fb950;
            }
        """)
        add_btn.clicked.connect(self.add_rule)
        toolbar.addWidget(add_btn)
        
        edit_btn = QPushButton("编辑规则")
        edit_btn.setStyleSheet("""
            QPushButton {
                background-color: #21262d;
                color: #58a6ff;
                border: 1px solid #58a6ff;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #30363d;
                border-color: #58a6ff;
            }
        """)
        edit_btn.clicked.connect(self.edit_rule)
        toolbar.addWidget(edit_btn)
        
        delete_btn = QPushButton("删除规则")
        delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #21262d;
                color: #f85149;
                border: 1px solid #f85149;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #30363d;
                border-color: #f85149;
            }
        """)
        delete_btn.clicked.connect(self.delete_rule)
        toolbar.addWidget(delete_btn)
        
        toolbar.addStretch()
        layout.addLayout(toolbar)
        
        self.rules_table = QTableWidget()
        self.rules_table.setColumnCount(5)
        self.rules_table.setHorizontalHeaderLabels(["ID", "规则名称", "描述", "分数修改", "启用状态"])
        self.rules_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.rules_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.rules_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.rules_table.setAlternatingRowColors(True)
        self.rules_table.doubleClicked.connect(self.on_rules_double_clicked)
        # 设置列宽比例：ID窄(50)，规则名称宽(150)，描述最宽(200)，分数修改(80)，启用状态(80)
        self.rules_table.setColumnWidth(0, 50)
        self.rules_table.setColumnWidth(1, 150)
        self.rules_table.setColumnWidth(2, 200)
        self.rules_table.setColumnWidth(3, 80)
        self.rules_table.setColumnWidth(4, 80)
        self.rules_table.horizontalHeader().setStretchLastSection(True)
        self.rules_table.setStyleSheet("""
            QTableWidget { border: 1px solid #30363d; background-color: #0d1117; color: #c9d1d9; }
            QTableWidget::item { padding: 5px; text-align: center; }
            QHeaderView::section { background: #21262d; font-weight: bold; padding: 6px; color: #58a6ff; }
        """)
        # 设置表格高度策略
        self.rules_table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.rules_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.rules_table)
        
        # 让规则列表拉伸填满剩余空间
        layout.setStretch(0, 0)  # info_label
        layout.setStretch(1, 0)  # toolbar
        layout.setStretch(2, 1)  # rules_table 拉伸
        layout.setStretch(3, 0)  # stats_label

        self.rules_stats_label = QLabel()
        self.rules_stats_label.setStyleSheet("font-weight: bold; font-size: 14px; padding: 10px;")
        layout.addWidget(self.rules_stats_label)
        
        self.load_rules()
        
        scroll.setWidget(widget)
        return scroll
    
    def load_templates(self):
        """加载席位模版列表"""
        templates = self.db.get_all_templates()
        self.templates_table.setRowCount(len(templates))
        
        for i, tpl in enumerate(templates):
            tpl = dict(tpl)
            self.templates_table.setItem(i, 0, QTableWidgetItem(str(tpl['id'])))
            self.templates_table.item(i, 0).setTextAlignment(Qt.AlignCenter)
            self.templates_table.setItem(i, 1, QTableWidgetItem(tpl['name']))
            self.templates_table.item(i, 1).setTextAlignment(Qt.AlignCenter)
            # 描述
            self.templates_table.setItem(i, 2, QTableWidgetItem(tpl.get('description', '')))
            self.templates_table.item(i, 2).setTextAlignment(Qt.AlignCenter)
            # 可用状态 - 点击切换
            active_text = "✓ 可用" if tpl.get('available', 1) else "✗ 禁用"
            active_item = QTableWidgetItem(active_text)
            active_item.setTextAlignment(Qt.AlignCenter)
            active_item.setForeground(QColor(39, 174, 96) if tpl.get('available', 1) else QColor(149, 165, 166))
            active_item.setData(Qt.UserRole, tpl['id'])  # 存储ID用于点击时切换
            self.templates_table.setItem(i, 3, active_item)
        
        # 双击切换可用状态（防重复连接）
        try:
            self.templates_table.cellDoubleClicked.disconnect()
        except:
            pass
        self.templates_table.cellDoubleClicked.connect(self.on_template_available_clicked)

        # 刷新所有席位的模版下拉框
        for widget in self.seat_widgets:
            widget.refresh_template_combo()
    
    def _build_template_slots(self):
        """构建所有席位的时段配置，用于C1/C2规则检查"""
        position_to_index = {'A': 0, 'B': 1, 'C': 2, 'D': 3}
        self.template_slots = {}  # {seat_id: [slot_dict, ...]}
        
        for seat_id, seat_info in self.schedule_data.seats.items():
            template_id = seat_info.get('template_id')
            if not template_id:
                continue
            
            slots = self.db.get_template_time_slots(template_id)
            if slots:
                self.template_slots[seat_id] = slots

    def _get_next_template_id(self):
        """获取下一个模版ID（数字）"""
        return self.db.get_next_template_id()
    
    def add_template(self):
        """添加席位模版"""
        dialog = TemplateEditDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            success, msg = self.db.add_template(data['id'], data['name'], data.get('description', ''), data.get('time_start', '08:40'), data.get('time_end', '10:30'), data.get('available', 1))
            if success:
                # 设置默认位置
                positions = ['A', 'B', 'C'] if data.get('persons_count', 3) == 3 else ['A', 'B', 'C', 'D']
                self.db.set_template_positions(data['id'], positions)
                CustomMessageBox.information(self, "成功", "模版添加成功")
                self.load_templates()
            else:
                CustomMessageBox.warning(self, "失败", msg)
    
    def edit_template(self):
        """编辑席位模版"""
        row = self.templates_table.currentRow()
        if row < 0:
            CustomMessageBox.warning(self, "请选择", "请先选择要编辑的模版")
            return
        
        template_id = int(self.templates_table.item(row, 0).text())
        template = self.db.get_template(template_id)
        
        if template:
            dialog = TemplateEditDialog(self, template)
            if dialog.exec_() == QDialog.Accepted:
                data = dialog.get_data()
                old_id = template_id
                new_id = data.get('id')
                self.db.update_template(old_id, data.get('name'), data.get('description'), new_id=new_id, available=data.get('available'))
                self.load_templates()
    
    def delete_template(self):
        """删除席位模版"""
        row = self.templates_table.currentRow()
        if row < 0:
            CustomMessageBox.warning(self, "请选择", "请先选择要删除的模版")
            return
        
        template_id = int(self.templates_table.item(row, 0).text())
        reply = CustomMessageBox.question(self, "确认删除", f"确定要删除模版 {template_id} 吗？")
        if not reply:
            return
        
        success, msg = self.db.delete_template(template_id)
        CustomMessageBox.information(self, "成功", msg)
        self.load_templates()
    
    def on_template_available_clicked(self, row, col):
        """双击席位模版：可用状态列切换状态，其他列打开编辑对话框"""
        template_id = int(self.templates_table.item(row, 0).text())
        template = self.db.get_template(template_id)
        if not template:
            return
        
        if col == 3:
            # 可用状态列 - 切换状态
            new_available = 0 if template.get('available', 1) else 1
            self.db.update_template_available(template_id, new_available)
            self.load_templates()
        else:
            # 其他列 - 打开编辑对话框
            dialog = TemplateEditDialog(self, template)
            if dialog.exec_() == QDialog.Accepted:
                data = dialog.get_data()
                self.db.update_template(template_id, data.get('name'), data.get('description'), available=data.get('available'))
                self.load_templates()
    
    def preview_template(self):
        """预览席位模版详情（可编辑时段）"""
        row = self.templates_table.currentRow()
        if row < 0:
            CustomMessageBox.warning(self, "请选择", "请先选择要预览的模版")
            return
        
        template_id = int(self.templates_table.item(row, 0).text())
        template = self.db.get_template(template_id)
        
        if template:
            slots = self.db.get_template_time_slots(template_id)
            
            dialog = QDialog(self)
            dialog.setWindowTitle(f"模版预览 - {template['name']}")
            dialog.setMinimumSize(700, 480)
            dialog.setStyleSheet("""
                QDialog { background-color: #161b22; }
                QLabel { color: #c9d1d9; }
                QLineEdit { padding: 6px; border: 1px solid #30363d; border-radius: 4px; background: #21262d; color: #c9d1d9; font-size: 13px; }
                QPushButton { background-color: #58a6ff; color: white; padding: 8px 20px; border: none; border-radius: 4px; font-size: 13px; }
                QComboBox { padding: 6px; border: 1px solid #30363d; border-radius: 4px; background: #21262d; color: #c9d1d9; font-size: 13px; }
            """)
            
            layout = QVBoxLayout(dialog)
            
            # 标题和状态
            header_layout = QHBoxLayout()
            
            title = QLabel(f"模版: {template['name']} ({template_id})")
            title.setStyleSheet("font-size: 16px; font-weight: bold; color: #58a6ff;")
            header_layout.addWidget(title)
            
            header_layout.addStretch()
            
            available_text = "可用" if template.get('available', 1) else "禁用"
            available_color = "#3fb950" if template.get('available', 1) else "#f85149"
            status_label = QLabel(f"状态: {available_text}")
            status_label.setStyleSheet(f"font-size: 14px; font-weight: bold; color: {available_color};")
            header_layout.addWidget(status_label)
            
            layout.addLayout(header_layout)
            
# 时段配置标题
            config_title = QLabel("时段配置")
            config_title.setStyleSheet("font-size: 14px; font-weight: bold; color: #58a6ff; padding: 15px 0 10px 0;")
            layout.addWidget(config_title)

            table = QTableWidget()
            table.setColumnCount(5)
            table.setHorizontalHeaderLabels(["班次", "管制时段", "管制席", "助理时段", "助理席"])
            table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
            table.setStyleSheet("""
                QTableWidget { border: 1px solid #30363d; background-color: #0d1117; color: #c9d1d9; gridline-color: #30363d; }
                QTableWidget::item { padding: 8px; text-align: center; border: 1px solid #30363d; font-size: 14px; }
                QTableWidget::item:selected { background-color: #58a6ff; color: white; }
                QHeaderView::section { background: #21262d; font-weight: bold; padding: 10px; color: #58a6ff; border: 1px solid #30363d; font-size: 14px; }
            """)
            
            # 编辑状态下不默认填充时间段，让用户手动添加
            # slots为空时表格显示为0行
            if not slots:
                table.setRowCount(0)
            else:
                table.setRowCount(len(slots))
                for i, slot in enumerate(slots):
                    # 班次下拉框
                    shift_combo = QComboBox()
                    shift_combo.addItems(['早班', '下午班', '晚班'])
                    shift_combo.setCurrentText(slot.get('shift_name', '早班'))
                    table.setCellWidget(i, 0, shift_combo)
                    
                    # 管制时段输入框
                    ctrl_time_edit = QLineEdit(slot.get('ctrl_time', ''))
                    ctrl_time_edit.setAlignment(Qt.AlignCenter)
                    ctrl_time_edit.setPlaceholderText("0850-1040")
                    table.setCellWidget(i, 1, ctrl_time_edit)
                    
                    # 管制席下拉框
                    ctrl_combo = QComboBox()
                    ctrl_combo.addItems(['', 'A', 'B', 'C', 'D'])
                    ctrl_combo.setCurrentText(slot.get('ctrl_position', ''))
                    table.setCellWidget(i, 2, ctrl_combo)
                    
                    # 助理时段输入框
                    asst_time_edit = QLineEdit(slot.get('asst_time', ''))
                    asst_time_edit.setAlignment(Qt.AlignCenter)
                    asst_time_edit.setPlaceholderText("0850-1040")
                    table.setCellWidget(i, 3, asst_time_edit)
                    
                    # 助理席下拉框
                    asst_combo = QComboBox()
                    asst_combo.addItems(['', 'A', 'B', 'C', 'D'])
                    asst_combo.setCurrentText(slot.get('asst_position', ''))
                    table.setCellWidget(i, 4, asst_combo)
            
            table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            table.setAlternatingRowColors(True)
            table.verticalHeader().setDefaultSectionSize(45)
            layout.addWidget(table)
            
            btn_layout = QHBoxLayout()
            
            add_btn = QPushButton("添加行")
            add_btn.clicked.connect(lambda: self._add_time_slot_row(table))
            btn_layout.addWidget(add_btn)
            
            del_btn = QPushButton("删除行")
            del_btn.clicked.connect(lambda: self._del_time_slot_row(table))
            btn_layout.addWidget(del_btn)
            
            btn_layout.addStretch()
            
            save_btn = QPushButton("保存")
            save_btn.clicked.connect(lambda: self._save_template_slots_only(template_id, table, dialog))
            btn_layout.addWidget(save_btn)
            
            close_btn = QPushButton("关闭")
            close_btn.clicked.connect(dialog.close)
            btn_layout.addWidget(close_btn)
            
            layout.addLayout(btn_layout)
            
            dialog.exec_()

    def _save_template_slots_only(self, template_id, table, dialog):
        """只保存时段配置"""
        slots = []
        for i in range(table.rowCount()):
            # 获取班次
            shift_name = table.cellWidget(i, 0).currentText() if table.cellWidget(i, 0) else ''
            
            # 获取管制时段 (QLineEdit)
            ctrl_time = ''
            if table.cellWidget(i, 1):
                ctrl_time = table.cellWidget(i, 1).text()
            
            # 获取管制席
            ctrl_position = table.cellWidget(i, 2).currentText() if table.cellWidget(i, 2) else ''
            
            # 获取助理时段 (QLineEdit)
            asst_time = ''
            if table.cellWidget(i, 3):
                asst_time = table.cellWidget(i, 3).text()
            
            # 获取助理席
            asst_position = table.cellWidget(i, 4).currentText() if table.cellWidget(i, 4) else ''
            
            slots.append({
                'shift_name': shift_name,
                'ctrl_time': ctrl_time,
                'ctrl_position': ctrl_position,
                'asst_time': asst_time,
                'asst_position': asst_position
            })
        
        self.db.set_template_time_slots(template_id, slots)
        CustomMessageBox.information(self, "成功", "时段保存成功")
        dialog.close()
    
    def _add_time_slot_row(self, table):
        """添加时段行"""
        row = table.rowCount()
        table.insertRow(row)
        
        shift_combo = QComboBox()
        shift_combo.addItems(['早班', '下午班', '晚班'])
        table.setCellWidget(row, 0, shift_combo)
        
        # 管制时段输入框
        ctrl_time_edit = QLineEdit('')
        ctrl_time_edit.setAlignment(Qt.AlignCenter)
        ctrl_time_edit.setPlaceholderText("0850-1040")
        table.setCellWidget(row, 1, ctrl_time_edit)
        
        ctrl_combo = QComboBox()
        ctrl_combo.addItems(['', 'A', 'B', 'C', 'D'])
        table.setCellWidget(row, 2, ctrl_combo)
        
        # 助理时段输入框
        asst_time_edit = QLineEdit('')
        asst_time_edit.setAlignment(Qt.AlignCenter)
        asst_time_edit.setPlaceholderText("0850-1040")
        table.setCellWidget(row, 3, asst_time_edit)
        
        asst_combo = QComboBox()
        asst_combo.addItems(['', 'A', 'B', 'C', 'D'])
        table.setCellWidget(row, 4, asst_combo)
    
    def _del_time_slot_row(self, table):
        """删除时段行"""
        if table.rowCount() > 0:
            table.removeRow(table.rowCount() - 1)
    
    def _save_template_all(self, template_id, table, name_edit, status_combo, dialog):
        """保存模版所有信息"""
        # 保存基本信息
        name = name_edit.text().strip()
        available = 1 if status_combo.currentIndex() == 0 else 0
        
        # 获取旧description
        old_template = self.db.get_template(template_id)
        description = old_template.get('description', '') if old_template else ''
        
        self.db.update_template(template_id, name, description)
        self.db.update_template_available(template_id, available)
        
        # 保存时段配置
        slots = []
        for i in range(table.rowCount()):
            slot = {
                'shift_name': table.cellWidget(i, 0).currentText() if table.cellWidget(i, 0) else '',
                'ctrl_time': table.item(i, 1).text() if table.item(i, 1) else '',
                'ctrl_position': table.cellWidget(i, 2).currentText() if table.cellWidget(i, 2) else '',
                'asst_time': table.item(i, 3).text() if table.item(i, 3) else '',
                'asst_position': table.cellWidget(i, 4).currentText() if table.cellWidget(i, 4) else ''
            }
            slots.append(slot)
        
        self.db.set_template_time_slots(template_id, slots)
        CustomMessageBox.information(self, "成功", "模版保存成功")
        dialog.close()
        self.load_templates()
    
    def edit_template_time_slots(self, template_id, template):
        """打开编辑时段对话框（保留兼容，双击调用）"""
        # 选中对应行并打开编辑
        for row in range(self.templates_table.rowCount()):
            if self.templates_table.item(row, 0).text() == str(template_id):
                self.templates_table.selectRow(row)
                break
        self.preview_template()

    def load_rules(self):
        rules = self.db.get_all_rules()
        self.rules_table.setRowCount(len(rules))
        
        for i, rule in enumerate(rules):
            rule = dict(rule)
            self.rules_table.setItem(i, 0, QTableWidgetItem(str(rule['id'])))
            self.rules_table.item(i, 0).setTextAlignment(Qt.AlignCenter)
            self.rules_table.setItem(i, 1, QTableWidgetItem(rule['name']))
            self.rules_table.item(i, 1).setTextAlignment(Qt.AlignCenter)
            self.rules_table.setItem(i, 2, QTableWidgetItem(rule.get('description', '')))
            self.rules_table.item(i, 2).setTextAlignment(Qt.AlignCenter)
            score_item = QTableWidgetItem(f"{rule['score_modifier']:+.2f}")
            score_item.setTextAlignment(Qt.AlignCenter)
            score_item.setForeground(QColor(231, 76, 60) if rule['score_modifier'] < 0 else QColor(39, 174, 96))
            self.rules_table.setItem(i, 3, score_item)
            active_text = "✓ 启用" if rule['active'] else "✗ 禁用"
            active_item = QTableWidgetItem(active_text)
            active_item.setTextAlignment(Qt.AlignCenter)
            active_item.setForeground(QColor(39, 174, 96) if rule['active'] else QColor(149, 165, 166))
            self.rules_table.setItem(i, 4, active_item)
        
        active_count = sum(1 for r in rules if r['active'])
        self.rules_stats_label.setText(f"共 {len(rules)} 条规则，启用 {active_count} 条")
    
    def on_rules_double_clicked(self, index):
        row = index.row()
        rule_id = int(self.rules_table.item(row, 0).text())
        rule = self.db.get_rule_by_id(rule_id)
        dialog = RuleEditDialog(self, dict(rule))
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            success, msg = self.db.update_rule(rule_id, data['name'], data['description'], data['score_modifier'], data['active'])
            if success:
                self.load_rules()
            else:
                CustomMessageBox.warning(self, "失败", msg)
    
    def add_rule(self):
        dialog = RuleEditDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            success, msg = self.db.add_rule(data['name'], data['description'], data['score_modifier'])
            if success:
                CustomMessageBox.information(self, "成功", msg)
                self.load_rules()
            else:
                CustomMessageBox.warning(self, "失败", msg)
    
    def edit_rule(self):
        row = self.rules_table.currentRow()
        if row < 0:
            CustomMessageBox.warning(self, "请选择", "请先选择要编辑的规则")
            return
        rule_id = int(self.rules_table.item(row, 0).text())
        rule = self.db.get_rule_by_id(rule_id)
        dialog = RuleEditDialog(self, dict(rule))
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            success, msg = self.db.update_rule(rule_id, data['name'], data['description'], data['score_modifier'], data['active'])
            if success:
                CustomMessageBox.information(self, "成功", msg)
                self.load_rules()
            else:
                CustomMessageBox.warning(self, "失败", msg)
    
    def delete_rule(self):
        row = self.rules_table.currentRow()
        if row < 0:
            CustomMessageBox.warning(self, "请选择", "请先选择要删除的规则")
            return
        rule_id = int(self.rules_table.item(row, 0).text())
        rule_name = self.rules_table.item(row, 1).text()
        if CustomMessageBox.question(self, "确认删除", f"确定要删除规则「{rule_name}」吗？"):
            success, msg = self.db.delete_rule(rule_id)
            if success:
                CustomMessageBox.information(self, "成功", msg)
                self.load_rules()
            else:
                CustomMessageBox.warning(self, "失败", msg)
    
    def create_person_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("搜索:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("输入姓名搜索...")
        self.search_edit.textChanged.connect(self.on_search)
        search_layout.addWidget(self.search_edit)
        search_layout.addWidget(QLabel("  等级:"))
        self.filter_combo = QComboBox()
        self.filter_combo.addItems(["全部", "C1", "C2", "C3", "I", "S"])
        self.filter_combo.currentTextChanged.connect(self.load_data)
        search_layout.addWidget(self.filter_combo)
        layout.addLayout(search_layout)
        
        toolbar = QHBoxLayout()
        
        self.add_person_btn = QPushButton("添加人员")
        self.add_person_btn.setStyleSheet("""
            QPushButton {
                background-color: #21262d;
                color: #3fb950;
                border: 1px solid #3fb950;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #30363d;
                border-color: #3fb950;
            }
        """)
        self.add_person_btn.clicked.connect(self.add_person)
        toolbar.addWidget(self.add_person_btn)
        
        self.edit_person_btn = QPushButton("编辑信息")
        self.edit_person_btn.setStyleSheet("""
            QPushButton {
                background-color: #21262d;
                color: #58a6ff;
                border: 1px solid #58a6ff;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #30363d;
                border-color: #58a6ff;
            }
        """)
        self.edit_person_btn.clicked.connect(self.edit_person)
        toolbar.addWidget(self.edit_person_btn)
        
        self.detail_person_btn = QPushButton("查看详情")
        self.detail_person_btn.setStyleSheet("""
            QPushButton {
                background-color: #21262d;
                color: #a371f7;
                border: 1px solid #a371f7;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #30363d;
                border-color: #a371f7;
            }
        """)
        self.detail_person_btn.clicked.connect(self.view_person_detail)
        toolbar.addWidget(self.detail_person_btn)
        
        self.delete_person_btn = QPushButton("删除人员")
        self.delete_person_btn.setStyleSheet("""
            QPushButton {
                background-color: #21262d;
                color: #f85149;
                border: 1px solid #f85149;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #30363d;
                border-color: #f85149;
            }
        """)
        self.delete_person_btn.clicked.connect(self.delete_person)
        toolbar.addWidget(self.delete_person_btn)
        
        toolbar.addStretch()
        
        self.toggle_active_btn = QPushButton("切换状态")
        self.toggle_active_btn.setStyleSheet("""
            QPushButton {
                background-color: #21262d;
                color: #c9d1d9;
                border: 1px solid #c9d1d9;
                border-radius: 6px;
                padding: 8px 16px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #30363d;
                border-color: #58a6ff;
                color: #58a6ff;
            }
        """)
        self.toggle_active_btn.clicked.connect(self.toggle_active)
        toolbar.addWidget(self.toggle_active_btn)
        
        layout.addLayout(toolbar)
        
        self.person_table = QTableWidget()
        self.person_table.setColumnCount(5)
        self.person_table.setHorizontalHeaderLabels(["姓名", "等级", "分数", "参与排班", "生效规则"])
        self.person_table.horizontalHeader().setStretchLastSection(True)
        self.person_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.person_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.person_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.person_table.setAlternatingRowColors(True)
        self.person_table.doubleClicked.connect(self.on_person_double_clicked)
        self.person_table.cellClicked.connect(self.on_person_cell_clicked)
        self.person_table.setItemDelegateForColumn(0, CenterAlignDelegate(self.person_table))
        self.person_table.setColumnWidth(0, 120)
        self.person_table.setStyleSheet("""
            QTableWidget { border: 1px solid #dcdcdc; gridline-color: #e0e0e0; background: white; }
            QTableWidget::item { padding: 5px; text-align: center; }
        """)
        layout.addWidget(self.person_table)
        
        stats_layout = QHBoxLayout()
        stats_layout.addSpacing(10)
        
        self.stats_label = QLabel("总计: 0人")
        self.stats_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        stats_layout.addWidget(self.stats_label)
        
        stats_layout.addStretch()
        
        self.c1_count_label = QLabel("C1: 0人")
        self.c1_count_label.setStyleSheet("color: #e74c3c; font-weight: bold; font-size: 14px;")
        stats_layout.addWidget(self.c1_count_label)
        
        self.c2_count_label = QLabel("C2: 0人")
        self.c2_count_label.setStyleSheet("color: #f39c12; font-weight: bold; font-size: 14px;")
        stats_layout.addWidget(self.c2_count_label)
        
        self.c3_count_label = QLabel("C3: 0人")
        self.c3_count_label.setStyleSheet("color: #27ae60; font-weight: bold; font-size: 14px;")
        stats_layout.addWidget(self.c3_count_label)
        
        self.i_count_label = QLabel("I: 0人")
        self.i_count_label.setStyleSheet("color: #3498db; font-weight: bold; font-size: 14px;")
        stats_layout.addWidget(self.i_count_label)
        
        self.s_count_label = QLabel("S: 0人")
        self.s_count_label.setStyleSheet("color: #9b59b6; font-weight: bold; font-size: 14px;")
        stats_layout.addWidget(self.s_count_label)
        
        layout.addLayout(stats_layout)
        
        self.load_data()
        return widget
    
    def create_schedule_tab(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        seats_group = QGroupBox("席位设置")
        seats_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                border: 2px solid #58a6ff;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                background-color: #161b22;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 5px;
                color: #58a6ff;
            }
        """)
        seats_layout = QVBoxLayout(seats_group)
        
        seats_data = self.db.get_all_seats()
        for seat_data in seats_data:
            seat_widget = SeatWidget(seat_data, self.db)
            seat_widget.available_changed.connect(self.on_seat_available_changed)
            seat_widget.count_changed.connect(self.on_seat_count_changed)
            seat_widget.score_changed.connect(self.on_seat_score_changed)
            seat_widget.app_name_changed.connect(self.on_seat_app_name_changed)
            seats_layout.addWidget(seat_widget)
            self.seat_widgets.append(seat_widget)
        
        layout.addWidget(seats_group)
        
        self.seat_stats_label = QLabel("可用席位: 5个, 总容量: 15人")
        self.seat_stats_label.setStyleSheet("font-weight: bold; font-size: 14px; padding: 12px; background-color: #161b22; border-radius: 8px; color: #58a6ff;")
        layout.addWidget(self.seat_stats_label)

        result_group = QGroupBox("排班结果")
        result_group.setStyleSheet("""
            QGroupBox { font-weight: bold; font-size: 14px; border: 2px solid #238636; border-radius: 8px; margin-top: 10px; padding-top: 10px; background-color: #161b22; }
            QGroupBox::title { subcontrol-origin: margin; left: 15px; padding: 0 5px; color: #58a6ff; }
        """)
        result_layout = QVBoxLayout(result_group)
        
        all_seats = self.db.get_all_seats()
        for seat_data in all_seats:
            seat = Seat(seat_data['app_name'], bool(seat_data.get('available', 1)), seat_data.get('persons_count', 3), seat_data.get('required_score', 5), seat_data.get('template_id'))
            result_widget = SeatResultWidget(seat)
            result_layout.addWidget(result_widget)
            self.seat_result_widgets.append(result_widget)
        
        layout.addWidget(result_group)
        
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)
        
        schedule_btn = QPushButton("🚀 开始排班")
        schedule_btn.setMinimumHeight(50)
        schedule_btn.setMinimumWidth(180)
        schedule_btn.setStyleSheet("""
            QPushButton { background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #2ecc71, stop:1 #27ae60); color: white; font-size: 15px; font-weight: bold; border: none; border-radius: 8px; padding: 8px 16px; }
            QPushButton:hover { background-color: #27ae60; }
        """)
        schedule_btn.clicked.connect(self.run_schedule)
        btn_layout.addWidget(schedule_btn)
        
        # 排班结果/告警信息显示框
        self.schedule_result_display = QTextEdit()
        self.schedule_result_display.setReadOnly(True)
        self.schedule_result_display.setMaximumHeight(80)
        self.schedule_result_display.setPlaceholderText("排班结果和告警信息将显示在这里...")
        self.schedule_result_display.setStyleSheet("""
            QTextEdit {
                background-color: #161b22;
                color: #c9d1d9;
                border: 2px solid #30363d;
                border-radius: 8px;
                padding: 10px;
                font-size: 12px;
                font-family: "SF Mono", "Menlo", monospace;
            }
        """)
        btn_layout.addWidget(self.schedule_result_display, 1)  # stretch=1
        
        clear_btn = QPushButton("🗑️ 清除结果")
        clear_btn.setMinimumHeight(50)
        clear_btn.setMinimumWidth(140)
        clear_btn.setStyleSheet("""
            QPushButton { background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #95a5a6, stop:1 #7f8c8d); color: white; font-size: 14px; border: none; border-radius: 8px; }
        """)
        clear_btn.clicked.connect(self.clear_schedule_result)
        btn_layout.addWidget(clear_btn)
        
        layout.addLayout(btn_layout)
        
        return widget
    
    def create_preview_tab(self):
        widget = QWidget()
        main_layout = QVBoxLayout(widget)
        main_layout.setContentsMargins(5, 5, 5, 5)
        
        info_label = QLabel("排班预览：按样表格式显示席位、人员、执勤时段安排")
        info_label.setStyleSheet("color: #8b949e; font-size: 12px; padding: 10px; background-color: #161b22; border-radius: 5px;")
        main_layout.addWidget(info_label)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border: none;")
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        main_layout.addWidget(scroll)
        
        self.preview_container = QWidget()
        self.preview_layout = QVBoxLayout(self.preview_container)
        self.preview_layout.setSpacing(10)
        self.preview_layout.setContentsMargins(0, 0, 0, 0)
        scroll.setWidget(self.preview_container)

        export_btn = QPushButton("📊 导出Excel")
        export_btn.setMinimumHeight(45)
        export_btn.setMinimumWidth(150)
        export_btn.setStyleSheet("""
            QPushButton { background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3498db, stop:1 #2980b9); color: white; font-size: 14px; font-weight: bold; border: none; border-radius: 5px; padding: 8px 20px; }
        """)
        export_btn.clicked.connect(self.export_to_excel)
        main_layout.addWidget(export_btn)
        
        self.refresh_preview()
        return widget
    
    def create_duty_data_tab(self):
        widget = QWidget()
        main_layout = QVBoxLayout(widget)
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        title_label = QLabel("执勤数据")
        title_label.setStyleSheet("font-size: 18px; font-weight: bold; color: #c9d1d9; padding: 5px;")
        main_layout.addWidget(title_label)
        
        info_label = QLabel("双击表格可编辑席位执勤次数")
        info_label.setStyleSheet("color: #8b949e; font-size: 12px; padding: 10px; background-color: #161b22; border-radius: 5px;")
        main_layout.addWidget(info_label)
        
        self.duty_table = QTableWidget()
        # 获取物理席位数据，设置动态列数和表头
        duty_seats = self.db.get_all_seats()
        seat_names = [s['app_name'] for s in duty_seats]  # 显示物理席位APP编号
        col_count = 3 + len(seat_names)  # 姓名、等级、上次席位 + 物理席位数量
        self.duty_table.setColumnCount(col_count)
        headers = ["姓名", "等级", "上次席位"] + seat_names
        self.duty_table.setHorizontalHeaderLabels(headers)
        self.duty_table.horizontalHeader().setStretchLastSection(True)
        self.duty_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.duty_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.duty_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.duty_table.setAlternatingRowColors(True)
        self.duty_table.doubleClicked.connect(self.on_duty_data_double_clicked)
        self.duty_table.setStyleSheet("""
            QTableWidget { border: 1px solid #ddd; border-radius: 5px; }
            QTableWidget::item { padding: 5px; }
            QTableWidget::item:selected { background-color: #3498db; color: white; }
        """)
        main_layout.addWidget(self.duty_table)
        
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(10)
        
        update_btn = QPushButton("✓ 更新")
        update_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 8px 20px; border: none; border-radius: 5px;")
        update_btn.clicked.connect(self.on_duty_data_updated)
        btn_layout.addWidget(update_btn)
        
        delete_btn = QPushButton("🗑️ 删除")
        delete_btn.setStyleSheet("background-color: #e74c3c; color: white; padding: 8px 20px; border: none; border-radius: 5px;")
        delete_btn.clicked.connect(self.on_duty_data_deleted)
        btn_layout.addWidget(delete_btn)
        
        cancel_btn = QPushButton("✗ 取消")
        cancel_btn.setStyleSheet("background-color: #95a5a6; color: white; padding: 8px 20px; border: none; border-radius: 5px;")
        cancel_btn.clicked.connect(self.refresh_duty_data)
        btn_layout.addWidget(cancel_btn)
        
        btn_layout.addStretch()
        main_layout.addLayout(btn_layout)
        
        self.refresh_duty_data()
        return widget
    
    def refresh_duty_data(self):
        persons = self.db.get_all_persons_with_duty_stats()
        seats = self.db.get_all_seats()
        
        self.duty_table.setRowCount(len(persons))
        
        for i, person in enumerate(persons):
            self.duty_table.setItem(i, 0, QTableWidgetItem(person['name']))
            self.duty_table.item(i, 0).setTextAlignment(Qt.AlignCenter)
            self.duty_table.setItem(i, 1, QTableWidgetItem(person['level']))
            self.duty_table.item(i, 1).setTextAlignment(Qt.AlignCenter)
            last_seat = person.get('last_seat_name', '') or '无'
            self.duty_table.setItem(i, 2, QTableWidgetItem(last_seat))
            self.duty_table.item(i, 2).setTextAlignment(Qt.AlignCenter)
            
            duty_map = person.get('duty_stats', {})  # duty_stats 已经是字典 {'APP01': 5, 'APP02': 3}
            
            for j, seat in enumerate(seats):
                count = duty_map.get(seat['app_name'], 0)
                item = QTableWidgetItem(str(count))
                item.setTextAlignment(Qt.AlignCenter)
                item.setData(Qt.UserRole, seat['app_name'])  # 存储app_name
                self.duty_table.setItem(i, 3 + j, item)
        
        self.duty_table.setColumnWidth(0, 100)
        self.duty_table.setColumnWidth(1, 80)
        self.duty_table.setColumnWidth(2, 120)
        for j in range(len(seats)):
            self.duty_table.setColumnWidth(3 + j, 100)
    
    def on_duty_data_updated(self):
        CustomMessageBox.information(self, "提示", "数据已更新！")
    
    def on_duty_data_deleted(self):
        selected = self.duty_table.selectedItems()
        if not selected:
            CustomMessageBox.warning(self, "提示", "请先选择要删除的数据行！")
            return
        
        row = selected[0].row()
        person_id = self.get_person_id_from_duty_row(row)
        if not person_id:
            return
        
        reply = CustomMessageBox.question(self, "确认", "确定要清空该人员的执勤数据吗？")
        if reply:
            self.db.clear_person_duty_stats(person_id)
            self.refresh_duty_data()
    
    def on_duty_data_double_clicked(self, index):
        row = index.row()
        col = index.column()
        
        if col < 3:
            return
        
        person_id = self.get_person_id_from_duty_row(row)
        if not person_id:
            return
        
        person = self.db.get_person_by_id(person_id)
        if not person:
            return
        
        seat_col = col - 3
        seats = self.db.get_all_seats()
        if seat_col >= len(seats):
            return
        
        seat = seats[seat_col]
        
        # 安全检查：确保单元格存在
        item = self.duty_table.item(row, col)
        if not item:
            return
        current_count = item.text() if item.text() else "0"
        
        dialog = DutyEditDialog(person, seat, current_count, self)
        if dialog.exec_():
            new_count = dialog.get_count()
            new_last_seat = dialog.get_last_seat()
            
            if new_count != '':
                self.db.set_duty_count(person_id, seat['app_name'], int(new_count))
            if new_last_seat is not None:
                self.db.set_person_last_seat(person_id, new_last_seat)
            
            self.refresh_duty_data()
    
    def get_person_id_from_duty_row(self, row):
        name_item = self.duty_table.item(row, 0)
        if name_item:
            person = self.db.get_person_by_name(name_item.text())
            if person:
                return person['id']
        return None
    
    def refresh_preview(self):
        # 检查所有席位是否已填满人员
        unfilled_seats = []
        for widget in self.seat_result_widgets:
            if not widget.seat.available:
                continue
            if not widget.seat.app_name or widget.seat.app_name == "空":
                continue
            
            seat_id_str = widget.seat.app_name
            if seat_id_str in self.schedule_data.selections:
                selected_with_none = self.schedule_data.selections[seat_id_str]
                filled_count = len([n for n in selected_with_none if n])
                seat_count = widget.seat.persons_count
                if filled_count < seat_count:
                    unfilled_seats.append(f"{widget.seat.app_name}({filled_count}/{seat_count})")
        
        # 如果有席位未填满，清空预览区域并显示提示
        while self.preview_layout.count():
            item = self.preview_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        if unfilled_seats:
            # 显示未填满的席位
            tip_label = QLabel(f"⚠️ 以下席位人员未填满：\n{', '.join(unfilled_seats)}")
            tip_label.setStyleSheet("color: #e74c3c; font-size: 14px; padding: 20px;")
            tip_label.setAlignment(Qt.AlignCenter)
            self.preview_layout.addWidget(tip_label)
            return

        seat_map = {}
        seat_app_map = {}
        
        # 按物理席位名称排序，APP03放在最后
        def seat_sort_key(widget):
            app = widget.seat.app_name or ''
            num = int(app.replace('APP', '')) if app.replace('APP', '').isdigit() else 0
            if app == 'APP03':
                num = 999  # APP03放最后
            return (num, widget.seat.app_name)
        
        sorted_widgets = sorted(self.seat_result_widgets, key=seat_sort_key)
        
        # 从数据层获取选择，保留位置信息
        # schedule_data.selections 格式: {seat_id: [name or None, ...]}
        for widget in sorted_widgets:
            # 跳过空app_name的席位
            if not widget.seat.available:
                continue
            if not widget.seat.app_name or widget.seat.app_name == "空":
                continue
            
            # 从数据层获取选择
            seat_id_str = widget.seat.app_name
            if seat_id_str in self.schedule_data.selections:
                selected_with_none = self.schedule_data.selections[seat_id_str]
                # 过滤None但保留位置信息
                selected_persons = [n for n in selected_with_none if n]
                # 同时记录是否显示4人席位（通过检查原始数据）
                seat_map[widget.seat.app_name] = selected_persons
                # 记录原始数据用于判断席位人数
                if not hasattr(self, '_preview_seat_counts'):
                    self._preview_seat_counts = {}
                self._preview_seat_counts[widget.seat.app_name] = len([n for n in selected_with_none if n])
            else:
                seat_map[widget.seat.app_name] = []
                if not hasattr(self, '_preview_seat_counts'):
                    self._preview_seat_counts = {}
                self._preview_seat_counts[widget.seat.app_name] = 0
            
            seat_app_map[widget.seat.app_name] = widget.seat.app_name
        
        # 获取每个席位的时段配置（从数据库读取）
        seat_shift_times = {}  # {app_name: [slot_dict...]}
        for widget in sorted_widgets:
            if not widget.seat.available:
                continue
            if not widget.seat.app_name or widget.seat.app_name == "空":
                continue
            
            seat_id_str = widget.seat.app_name
            # 获取席位关联的template_id
            seat_info = self.schedule_data.seats.get(seat_id_str, {})
            template_id = seat_info.get('template_id')
            if template_id:
                slots = self.db.get_template_time_slots(template_id)
                seat_shift_times[seat_id_str] = slots
        
        def get_person(persons, pos):
            if not persons:
                return ''
            if pos == 'A':
                return persons[0] if len(persons) > 0 else ''
            elif pos == 'B':
                return persons[1] if len(persons) > 1 else ''
            elif pos == 'C':
                return persons[2] if len(persons) > 2 else ''
            return ''
        
        # 按物理席位名称排序，APP03放最后
        def app_sort_key(item):
            app = item[1] or ''
            num = int(app.replace('APP', '')) if app.replace('APP', '').isdigit() else 0
            if app == 'APP03':
                num = 999
            return (num, item[0])
        
        available_seats = sorted(
            [(name, seat_app_map.get(name, name)) for name in seat_map.keys()],
key=app_sort_key
        )
        
        def create_seat_block(seat_name, app_name, persons, is_4person, shift_slots_from_db):
            seat_frame = QFrame()
            seat_frame.setStyleSheet("""
                QFrame { background-color: #161b22; border: 1px solid #30363d; border-radius: 4px; }
            """)
            seat_layout = QVBoxLayout(seat_frame)
            seat_layout.setContentsMargins(0, 0, 0, 0)
            seat_layout.setSpacing(0)
            
            # 实际选择的人员数量
            actual_count = len([p for p in persons if p])
            
            # 第一行：物理席位名称
            name_row = QFrame()
            name_row.setStyleSheet("background-color: #30363d;")
            name_layout = QHBoxLayout(name_row)
            name_layout.setContentsMargins(0, 0, 0, 0)
            name_layout.setSpacing(0)

            name_label = QLabel(app_name)
            name_label.setFont(QFont("Arial", 11, QFont.Bold))
            name_label.setStyleSheet("color: #ffffff; padding: 4px;")
            name_label.setAlignment(Qt.AlignCenter)
            name_layout.addWidget(name_label)
            seat_layout.addWidget(name_row)

            # 数据行表头
            header_row = QFrame()
            header_row.setStyleSheet("background-color: #21262d;")
            header_layout = QHBoxLayout(header_row)
            header_layout.setContentsMargins(0, 0, 0, 0)
            header_layout.setSpacing(0)

            for col_text in ['班次', '管制时段', '管制席', '助理时段', '助理席']:
                label = QLabel(col_text)
                label.setFont(QFont("Arial", 9, QFont.Bold))
                label.setStyleSheet("color: #c9d1d9; padding: 2px; border: 1px solid #30363d;")
                label.setAlignment(Qt.AlignCenter)
                header_layout.addWidget(label)
            seat_layout.addWidget(header_row)

            # 按row_order排序的时段数据
            sorted_slots = sorted(shift_slots_from_db, key=lambda x: x.get('row_order', 0)) if shift_slots_from_db else []
            
            # 按班次分组，每班次只显示第一个时段
            seen_shifts = set()
            for slot in sorted_slots:
                shift_name = slot.get('shift_name', '')
                ctrl_time = slot.get('ctrl_time', '')
                ctrl_pos = slot.get('ctrl_position', '')
                asst_time = slot.get('asst_time', '')
                asst_pos = slot.get('asst_position', '')
                
                # 判断是否是该班次的第一个时段
                is_first_of_shift = shift_name not in seen_shifts
                seen_shifts.add(shift_name)
                
                # 时段行
                data_row = QFrame()
                data_layout = QHBoxLayout(data_row)
                data_layout.setContentsMargins(0, 0, 0, 0)
                data_layout.setSpacing(0)
                
                # 班次列
                shift_label = QLabel(shift_name if is_first_of_shift else '')
                shift_label.setFont(QFont("Arial", 10, QFont.Bold) if is_first_of_shift else QFont("Arial", 10))
                shift_label.setStyleSheet("color: #c9d1d9; padding: 2px; border: 1px solid #30363d;")
                shift_label.setAlignment(Qt.AlignCenter)
                data_layout.addWidget(shift_label)
                
                # 少于3人时，每个班次第一行按顺序填人
                if actual_count < 3:
                    # 第一行填人员
                    ctrl_person = persons[0] if len(persons) > 0 else ''
                    asst_person = persons[1] if len(persons) > 1 else ''
                else:
                    # 3人及以上按位置填
                    ctrl_person = get_person(persons, ctrl_pos) if ctrl_pos else ''
                    asst_person = get_person(persons, asst_pos) if asst_pos else ''
                
                # 管制时段
                ctrl_time_label = QLabel(ctrl_time)
                ctrl_time_label.setStyleSheet("color: #c9d1d9; padding: 2px; border: 1px solid #30363d;")
                ctrl_time_label.setAlignment(Qt.AlignCenter)
                data_layout.addWidget(ctrl_time_label)

                # 管制席人员
                ctrl_person_label = QLabel(ctrl_person)
                ctrl_person_label.setStyleSheet("color: #58a6ff; padding: 2px; border: 1px solid #30363d;")
                ctrl_person_label.setAlignment(Qt.AlignCenter)
                data_layout.addWidget(ctrl_person_label)
                
                # 助理时段
                asst_time_label = QLabel(asst_time)
                asst_time_label.setStyleSheet("color: #c9d1d9; padding: 2px; border: 1px solid #30363d;")
                asst_time_label.setAlignment(Qt.AlignCenter)
                data_layout.addWidget(asst_time_label)

                # 助理席人员
                asst_person_label = QLabel(asst_person)
                asst_person_label.setStyleSheet("color: #58a6ff; padding: 2px; border: 1px solid #30363d;")
                asst_person_label.setAlignment(Qt.AlignCenter)
                data_layout.addWidget(asst_person_label)

                seat_layout.addWidget(data_row)
            
            return seat_frame
        
        for idx, (seat_name, app_name) in enumerate(available_seats):
            persons = seat_map.get(seat_name, [])
            # 使用实际的席位人数来判断
            seat_count = getattr(self, '_preview_seat_counts', {}).get(seat_name, 0)
            # 如果没有选择，获取席位配置的人数
            if seat_count == 0:
                for widget in self.seat_result_widgets:
                    if widget.seat.app_name == seat_name:
                        seat_count = widget.seat.persons_count
                        self._preview_seat_counts[widget.seat.app_name] = len([n for n in (self.schedule_data.selections.get(seat_id_str) or []) if n])
                        break
            is_4person = seat_count == 4
            # 获取该席位的时段配置（从数据库读取）
            shift_slots_from_db = seat_shift_times.get(app_name, [])
            seat_frame = create_seat_block(seat_name, app_name, persons, is_4person, shift_slots_from_db)
            seat_frame.setMinimumWidth(400)
            seat_frame.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Minimum)
            
            if idx % 2 == 0:
                row_frame = QFrame()
                row_frame.setStyleSheet("QFrame { background-color: transparent; }")
                row_layout = QHBoxLayout(row_frame)
                row_layout.setContentsMargins(0, 0, 0, 0)
                row_layout.setSpacing(10)
                self.preview_layout.addWidget(row_frame)
            else:
                row_frame = self.preview_layout.itemAt(self.preview_layout.count() - 1).widget()
            
            row_frame.layout().addWidget(seat_frame)
    
    def clear_schedule_result(self):
        self.current_schedule = None
        # 清除UI
        for widget in self.seat_result_widgets:
            widget.clear_selection()
        # 清除数据层
        self.schedule_data.selections = {}
        self._refresh_dropdowns()
        # 清除结果显示框
        self.schedule_result_display.clear()
        self.schedule_result_display.setStyleSheet("""
            QTextEdit {
                background-color: #161b22;
                color: #c9d1d9;
                border: 2px solid #30363d;
                border-radius: 8px;
                padding: 10px;
                font-size: 12px;
                font-family: "SF Mono", "Menlo", monospace;
            }
        """)
        self.status_bar.showMessage("已清除所有选择")
    
    def update_seat_result_widgets(self):
        seats = self.db.get_all_seats()
        
        for i, widget in enumerate(self.seat_result_widgets):
            if i < len(seats):
                new_seat = Seat(seats[i]['app_name'], bool(seats[i].get('available', 1)), seats[i].get('persons_count', 3), seats[i].get('required_score', 5), seats[i].get('template_id'))
                
                if new_seat.persons_count != widget.seat.persons_count:
                    widget.seat = new_seat
                    widget.recreate_combos(new_seat.persons_count)
                    active_persons = self.db.get_active_persons()
                    widget.update_persons_list(active_persons)
                else:
                    widget.seat = new_seat
                
                widget.set_available(new_seat.available)
                widget.title_label.setText(f"{widget.seat.app_name}/{len(active_persons)}/{widget.seat.persons_count}人")
    
    def refresh_ui(self):
        """从数据层刷新UI（被动刷新）"""
        for widget in self.seat_result_widgets:
            seat_id_str = widget.seat.app_name
            if seat_id_str in self.schedule_data.seats:
                available = self.schedule_data.seats[seat_id_str].get('available', True)
                widget.set_available(available)
                widget.update_count()
                # 检查规则并更新边框样式
                widget._check_and_update_border_style_noemit()
    
    def on_seat_available_changed(self, seat_id, available):
        # seat_id现在是字符串（app_name）
        seat_id_str = seat_id
        
        self.db.update_seat_available(seat_id_str, 1 if available else 0)
        
        # 更新数据层
        if seat_id_str in self.schedule_data.seats:
            self.schedule_data.seats[seat_id_str]['available'] = available
        
        # 禁用时清空选择
        if not available:
            selections = dict(self.schedule_data.selections)
            if seat_id_str in selections:
                del selections[seat_id_str]
                self.schedule_data.selections = selections
        else:
            # 启用时确保数据存在
            if seat_id_str not in self.schedule_data.selections:
                self.schedule_data.selections[seat_id_str] = [None] * self.schedule_data.seats[seat_id_str].get('count', 3)
        
        # 直接更新排班结果widget的禁用样式，不调用refresh_ui
        for widget in self.seat_result_widgets:
            if widget.seat.app_name == seat_id_str:
                widget.set_available(available)
                # 检查规则并更新边框样式
                widget._check_and_update_border_style_noemit()
                break
        
        # 刷新左侧Tab中的席位设置（刷新可用状态显示）
        for widget in self.seat_widgets:
            if widget.seat_id == seat_id:
                widget.seat_data['available'] = 1 if available else 0
                widget.available_switch.blockSignals(True)
                widget.available_switch.setChecked(available)
                widget.available_switch.setText("启用" if available else "禁用")
                widget.available_switch.blockSignals(False)
        
        self.status_bar.showMessage(f"席位 {seat_id_str} 已{'启用' if available else '禁用'}")
        
        self.update_seat_stats()
        self.update_stats()
        self.refresh_preview()
    
    def on_seat_count_changed(self, seat_id, count):
        # seat_id现在是字符串（app_name）
        seat_id_str = seat_id  # 直接使用字符串
        self.db.update_seat_persons_count(seat_id_str, count)
        
        # 更新数据层
        if seat_id_str in self.schedule_data.seats:
            self.schedule_data.seats[seat_id_str]['count'] = count
        
        for widget in self.seat_result_widgets:
            if widget.seat.app_name == seat_id_str:
                widget.seat.persons_count = count  # 更新席位人数
                widget.recreate_combos(count)
                active_persons = self.db.get_active_persons()
                widget.update_persons_list(active_persons)
                widget._check_and_update_border_style()  # 刷新边框样式
                break
        
        # 检查规则并更新显示框
        warnings = check_all_rules(
            self.schedule_data.selections,
            self.schedule_data.seats,
            self.schedule_data.persons,
            self.template_slots
        )
        if warnings:
            self.schedule_result_display.setText(format_warnings(warnings))
        else:
            self.schedule_result_display.setText("")
        
        self.update_seat_stats()
        self.status_bar.showMessage(f"席位 {seat_id} 人数已设为 {count}")
        
        # 刷新左侧Tab中的席位设置（刷新人数显示）
        for widget in self.seat_widgets:
            if widget.seat_id == seat_id:
                widget.count_spin.blockSignals(True)
                widget.count_spin.setValue(count)
                widget.count_spin.blockSignals(False)
                break
        
        # 刷新排班预览
        self.refresh_preview()
    
    def on_seat_score_changed(self, seat_id, required_score):
        # seat_id现在是字符串（app_name）
        seat_id_str = seat_id  # 直接使用字符串
        self.db.update_seat_required_score(seat_id_str, required_score)
        
        # 更新数据层
        if seat_id_str in self.schedule_data.seats:
            self.schedule_data.seats[seat_id_str]['required'] = required_score
        
        # 更新 widget.seat 对象
        for widget in self.seat_result_widgets:
            if widget.seat.app_name == seat_id_str:
                widget.seat.required_score = required_score
                widget.update_count()  # 刷新分数显示
                widget._check_and_update_border_style()  # 刷新边框样式
                break
        
        # 检查规则并更新显示框
        warnings = check_all_rules(
            self.schedule_data.selections,
            self.schedule_data.seats,
            self.schedule_data.persons,
            self.template_slots
        )
        if warnings:
            self.schedule_result_display.setText(format_warnings(warnings))
        else:
            self.schedule_result_display.setText("")
        
        self.update_seat_stats()
        self.status_bar.showMessage(f"席位 {seat_id} 分数要求已设为 {required_score}")
        
        # 刷新左侧Tab中的席位设置（刷新分数显示）
        for widget in self.seat_widgets:
            if widget.seat_id == seat_id:
                widget.score_combo.blockSignals(True)
                widget.score_combo.setCurrentText(str(required_score))
                widget.score_combo.blockSignals(False)
                break
        
        # 刷新排班预览
        self.refresh_preview()
    
    def on_seat_app_name_changed(self, seat_id, value):
        # value 是 "ID:名称" 格式，需要提取ID并保存到template_id字段
        template_id = None
        if value and value != "空":
            try:
                template_id = int(value.split(":")[0])
            except:
                pass
        
        # 更新数据库
        self.db.update_seat_template(seat_id, template_id)
        
        # 更新数据层
        if seat_id in self.schedule_data.seats:
            self.schedule_data.seats[seat_id]['template_id'] = template_id
        
        # 刷新排班预览
        self.refresh_preview()
        self.status_bar.showMessage(f"席位 {seat_id} 已关联模版 {value}")
    
    def update_seat_stats(self):
        seats = self.db.get_all_seats()
        available_count = sum(1 for s in seats if s['available'])
        total_capacity = sum(s['persons_count'] for s in seats if s['available'])
        self.seat_stats_label.setText(f"可用席位: {available_count}个, 总容量: {total_capacity}人")
    
    def load_data(self):
        filter_level = self.filter_combo.currentText()
        
        if filter_level == "全部":
            rows = self.db.get_all_persons()
        else:
            all_rows = self.db.get_all_persons()
            rows = [r for r in all_rows if r['level'] == filter_level]
        
        self.person_table.setRowCount(len(rows))
        
        for i, row in enumerate(rows):
            row = dict(row)
            name_item = QTableWidgetItem(row['name'])
            name_item.setTextAlignment(Qt.AlignCenter)
            self.person_table.setItem(i, 0, name_item)
            
            level_item = QTableWidgetItem(row['level'])
            level_item.setTextAlignment(Qt.AlignCenter)
            if row['level'] == "C1":
                level_item.setBackground(QColor(231, 76, 60))
                level_item.setForeground(QColor(255, 255, 255))
            elif row['level'] == "C2":
                level_item.setBackground(QColor(243, 156, 18))
            elif row['level'] == "C3":
                level_item.setBackground(QColor(39, 174, 96))
                level_item.setForeground(QColor(255, 255, 255))
            elif row['level'] == "I":
                level_item.setBackground(QColor(52, 152, 219))
                level_item.setForeground(QColor(255, 255, 255))
            elif row['level'] == "S":
                level_item.setBackground(QColor(155, 89, 182))
                level_item.setForeground(QColor(255, 255, 255))
            self.person_table.setItem(i, 1, level_item)
            
            score_item = QTableWidgetItem(str(row['score']))
            score_item.setTextAlignment(Qt.AlignCenter)
            self.person_table.setItem(i, 2, score_item)
            
            active_text = "✓ 是" if row['active'] else "✗ 否"
            active_item = QTableWidgetItem(active_text)
            active_item.setTextAlignment(Qt.AlignCenter)
            if not row['active']:
                active_item.setForeground(QColor(149, 165, 166))
            self.person_table.setItem(i, 3, active_item)
            
            person_rules = self.db.get_person_rules(row['id'])
            total_modifier = 0.0
            enabled_rules = []
            for r in person_rules:
                r_dict = dict(r)
                if r_dict.get('enabled') == 1 and r_dict.get('active') == 1:
                    enabled_rules.append(r_dict['name'])
                    total_modifier += r_dict.get('score_modifier', 0)
            
            effective_score = row['score'] + total_modifier
            rules_text = f"{effective_score:.2f}"
            if enabled_rules:
                rules_text += f" ({', '.join(enabled_rules)})"
            else:
                rules_text += " (无规则)"
            
            rules_item = QTableWidgetItem(rules_text)
            rules_item.setTextAlignment(Qt.AlignCenter)
            if total_modifier > 0:
                rules_item.setForeground(QColor(39, 174, 96))
            elif total_modifier < 0:
                rules_item.setForeground(QColor(231, 76, 60))
            else:
                rules_item.setForeground(QColor(149, 165, 166))
            self.person_table.setItem(i, 4, rules_item)
        
        self.person_table.setColumnWidth(0, 120)
        self.person_table.setColumnWidth(1, 100)
        self.person_table.setColumnWidth(2, 60)
        self.person_table.setColumnWidth(3, 80)
        self.person_table.setColumnWidth(4, 200)
        
        self.update_stats()
        
        if hasattr(self, 'seat_stats_label'):
            self.update_seat_stats()
    
    def update_stats(self):
        stats = self.db.get_statistics()
        self.stats_label.setText(f"总计: {stats['total']}人 (参与: {stats['active']}人)")
        self.c1_count_label.setText(f"C1: {stats['c1']}人")
        self.c2_count_label.setText(f"C2: {stats['c2']}人")
        self.c3_count_label.setText(f"C3: {stats['c3']}人")
        self.i_count_label.setText(f"I: {stats['i']}人")
        self.s_count_label.setText(f"S: {stats['s']}人")
    
    def on_person_double_clicked(self, index):
        person_id = self.get_selected_person_id()
        if person_id:
            dialog = PersonRuleSelectDialog(self, person_id, self.db)
            dialog.exec_()
            self.load_data()
    
    def on_person_cell_clicked(self, row, col):
        if col == 3:
            person_id = self.get_selected_person_id()
            if person_id:
                person = self.db.get_person_by_id(person_id)
                person_dict = dict(person)
                new_active = 0 if person_dict['active'] else 1
                self.db.update_person(person_id, person_dict['name'], person_dict['level'], new_active)
                self.load_data()
    
    def on_search(self, text):
        if text:
            rows = self.db.search_persons(text)
            self.person_table.setRowCount(len(rows))
            
            for i, row in enumerate(rows):
                name_item = QTableWidgetItem(row['name'])
                name_item.setTextAlignment(Qt.AlignCenter)
                self.person_table.setItem(i, 0, name_item)
                
                level_item = QTableWidgetItem(row['level'])
                level_item.setTextAlignment(Qt.AlignCenter)
                self.person_table.setItem(i, 1, level_item)
                
                score_item = QTableWidgetItem(str(row['score']))
                score_item.setTextAlignment(Qt.AlignCenter)
                self.person_table.setItem(i, 2, score_item)
                
                active_text = "✓ 是" if row['active'] else "✗ 否"
                active_item = QTableWidgetItem(active_text)
                active_item.setTextAlignment(Qt.AlignCenter)
                self.person_table.setItem(i, 3, active_item)
                
                person_rules = self.db.get_person_rules(row['id'])
                total_modifier = 0.0
                enabled_rules = []
                for r in person_rules:
                    r_dict = dict(r)
                    if r_dict.get('enabled') == 1 and r_dict.get('active') == 1:
                        enabled_rules.append(r_dict['name'])
                        total_modifier += r_dict.get('score_modifier', 0)
                
                effective_score = row['score'] + total_modifier
                rules_text = f"{effective_score:.2f}"
                if enabled_rules:
                    rules_text += f" ({', '.join(enabled_rules)})"
                
                rules_item = QTableWidgetItem(rules_text)
                rules_item.setTextAlignment(Qt.AlignCenter)
                self.person_table.setItem(i, 4, rules_item)
        else:
            self.load_data()
    
    def get_selected_person_id(self):
        row = self.person_table.currentRow()
        if row < 0:
            return None
        name = self.person_table.item(row, 0).text()
        person = self.db.get_person_by_name(name)
        return person['id'] if person else None
    
    def add_person(self):
        dialog = PersonEditDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            success, msg = self.db.add_person(data['name'], data['level'], data['active'])
            if success:
                CustomMessageBox.information(self, "成功", msg)
                self.load_data()
                self.status_bar.showMessage(f"已添加: {data['name']}")
            else:
                CustomMessageBox.warning(self, "失败", msg)
    
    def edit_person(self):
        person_id = self.get_selected_person_id()
        if person_id is None:
            CustomMessageBox.warning(self, "请选择", "请先选择要编辑的人员")
            return
        
        person = self.db.get_person_by_id(person_id)
        dialog = PersonEditDialog(self, dict(person))
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            # 使用用户输入的分数，而非默认分数
            success, msg = self.db.update_person(person_id, data['name'], data['level'], data['active'], data['score'])
            if success:
                CustomMessageBox.information(self, "成功", msg)
                self.load_data()
                self.status_bar.showMessage(f"已更新: {data['name']}")
            else:
                CustomMessageBox.warning(self, "失败", msg)
    
    def view_person_detail(self):
        person_id = self.get_selected_person_id()
        if person_id is None:
            CustomMessageBox.warning(self, "请选择", "请先选择要查看的人员")
            return
        
        person = self.db.get_person_by_id(person_id)
        dialog = PersonDetailDialog(self, dict(person))
        dialog.exec_()
    
    def delete_person(self):
        person_id = self.get_selected_person_id()
        if person_id is None:
            CustomMessageBox.warning(self, "请选择", "请先选择要删除的人员")
            return
        
        person = self.db.get_person_by_id(person_id)
        if CustomMessageBox.question(self, "确认删除", f"确定要删除 {person['name']} ({person['level']}) 吗?\n此操作不可恢复!"):
            success, msg = self.db.delete_person(person_id)
            if success:
                CustomMessageBox.information(self, "成功", msg)
                self.load_data()
            else:
                CustomMessageBox.warning(self, "失败", msg)
    
    def toggle_active(self):
        person_id = self.get_selected_person_id()
        if person_id is None:
            CustomMessageBox.warning(self, "请选择", "请先选择人员")
            return
        
        person = self.db.get_person_by_id(person_id)
        new_status = 0 if person['active'] else 1
        success, msg = self.db.update_person(person_id, person['name'], person['level'], new_status, person['score'])
        
        if success:
            status_text = "参与排班" if new_status else "不参与排班"
            CustomMessageBox.information(self, "成功", f"{person['name']} 已设置为: {status_text}")
            self.load_data()
        else:
            CustomMessageBox.warning(self, "失败", msg)
    
    def _run_schedule_original(self):
        """开始排班核心逻辑 - 使用统一数据源"""
        try:
            available_seats = self.db.get_available_seats()
            
            if not available_seats:
                self.schedule_result_display.setText("席位未启用，请至少启用一个席位")
                self.schedule_result_display.setStyleSheet("""
                    QTextEdit {
                        background-color: #161b22;
                        color: #e74c3c;
                        border: 2px solid #30363d;
                        border-radius: 8px;
                        padding: 10px;
                        font-size: 12px;
                        font-family: "SF Mono", "Menlo", monospace;
                    }
                """)
                return
            
            active_persons = self.db.get_active_persons()
            
            if not active_persons:
                self.schedule_result_display.setText("没有参与排班的人员")
                self.schedule_result_display.setStyleSheet("""
                    QTextEdit {
                        background-color: #161b22;
                        color: #e74c3c;
                        border: 2px solid #30363d;
                        border-radius: 8px;
                        padding: 10px;
                        font-size: 12px;
                        font-family: "SF Mono", "Menlo", monospace;
                    }
                """)
                return
            
            # 使用统一数据源 schedule_data.selections（格式：[A, B, C, D]，None表示未选）
            user_selections = {}  # {seat_id: [name or None, ...]} 保留位置
            all_selected_names = []
            
            for seat_id, names in self.schedule_data.selections.items():
                # names 是 [A, B, C, D] 格式，保留None以维持位置
                if any(n for n in names):  # 只要有一个非空
                    user_selections[seat_id] = list(names)  # 复制，保留位置
                    all_selected_names.extend([n for n in names if n])
            
            # 检查重复
            if len(all_selected_names) != len(set(all_selected_names)):
                self.schedule_result_display.setText("同一个人不能被多次选择！")
                return
            
            # 检查C1/C2规则
            person_map = {p['name']: dict(p) for p in active_persons}
            for seat_id, selected_names in user_selections.items():
                # seat_id就是app_name
                seat_name = seat_id  # 直接使用app_name
                
                c1_count = sum(1 for name in selected_names if person_map.get(name, {}).get('level') == 'C1')
                if c1_count > 1:
                    self.schedule_result_display.setText(f"{seat_name}: C1人员不能超过1个！")
                    return
                
                has_c1 = any(person_map.get(name, {}).get('level') == 'C1' for name in selected_names)
                has_c2 = any(person_map.get(name, {}).get('level') == 'C2' for name in selected_names)
                if has_c1 and has_c2:
                    self.schedule_result_display.setText(f"{seat_name}: C1和C2不能同席！")
                    return
            
            # 剩余人员
            remaining_persons = [p for p in active_persons if p['name'] not in all_selected_names]
            
            # 检查席位分要求（手动选择也要检测）
            for seat_id, selected_names in user_selections.items():
                # seat_id就是app_name，直接从seats字典获取
                seat_info = next((s for s in available_seats if s['app_name'] == seat_id), None)
                if not seat_info:
                    continue
                seat_name = seat_info['app_name']  # 使用app_name
                required_score = seat_info.get('required_score', 0)
                
                # 获取该席位已选人员的分数（包括规则修改）
                selected_people = []
                for name in selected_names:
                    if name is None:
                        continue
                    p = person_map.get(name)
                    if p:
                        score_mod = self.db.get_person_score_modifier(p['id'])
                        effective_score = p['score'] + score_mod
                        selected_people.append({'name': name, 'level': p['level'], 'effective_score': effective_score})
                
                # 检查席位分要求（>=3人时检查，4人席位分别检查ABC/BCD组合）
                if len(selected_people) >= 3 and required_score > 0:
                    # 4人席位：分别检查ABC和BCD组合
                    if seat_info.get('persons_count') == 4:
                        # ABC组合（位置0,1,2）
                        if len(selected_people) >= 3:
                            abc_score = sum(p['effective_score'] for p in selected_people[:3])
                            if abc_score < required_score:
                                self.schedule_result_display.setText(
                                    f"{seat_name}: ABC组合分数不足({abc_score}/{required_score}分)"
                                )
                                return
                        # BCD组合（位置1,2,3）
                        if len(selected_people) >= 4:
                            bcd_score = sum(p['effective_score'] for p in selected_people[1:4])
                            if bcd_score < required_score:
                                self.schedule_result_display.setText(
                                    f"{seat_name}: BCD组合分数不足({bcd_score}/{required_score}分)"
                                )
                                return
                    else:
                        # 非4人席位：检查总分
                        total_score = sum(p['effective_score'] for p in selected_people)
                        if total_score < required_score:
                            self.schedule_result_display.setText(
                                f"{seat_name}: 分数不足({total_score}/{required_score}分)"
                            )
                            return
            
            # 检查C1/C2列限制（手动选择也要检测）
            # 构造 check_all_rules 需要的数据格式
            seats_dict = {}
            for s in available_seats:
                seats_dict[s['app_name']] = {
                    'name': s['app_name'],
                    'count': s['persons_count'],
                    'available': s['available']
                }
            
            persons_dict = {}
            for p in active_persons:
                score_mod = self.db.get_person_score_modifier(p['id'])
                persons_dict[p['name']] = {
                    'name': p['name'],
                    'level': p['level'],
                    'score': p['score'],
                    'effective_score': p['score'] + score_mod
                }
            
            warnings = check_all_rules(user_selections, seats_dict, persons_dict, self.template_slots)
            if warnings:
                from models.scheduler import format_warnings
                warning_text = format_warnings(warnings)
                # 在结果显示框中显示
                self.schedule_result_display.setText(warning_text)
                self.schedule_result_display.setStyleSheet("""
                    QTextEdit {
                        background-color: #161b22;
                        color: #e74c3c;
                        border: 2px solid #30363d;
                        border-radius: 8px;
                        padding: 10px;
                        font-size: 12px;
                        font-family: "SF Mono", "Menlo", monospace;
                    }
                """)
                return
            
            # 如果人员刚好够，直接完成
            if not remaining_persons:
                self._update_ui_from_selections(user_selections)
                # 在结果显示框中显示
                self.schedule_result_display.setText("✅ 所有人员已安排完成！")
                self.schedule_result_display.setStyleSheet("""
                    QTextEdit {
                        background-color: #161b22;
                        color: #2ecc71;
                        border: 2px solid #2ecc71;
                        border-radius: 8px;
                        padding: 10px;
                        font-size: 12px;
                        font-family: "SF Mono", "Menlo", monospace;
                    }
                """)
                self.status_bar.showMessage("排班完成")
                self.refresh_preview()
                return
            
            # 构建人员对象（包括锁定人员和剩余人员）
            # 锁定的人员也要传给 scheduler
            all_person_objs = []
            
            # 先添加锁定的人员
            for seat_id, names in user_selections.items():
                for pos_idx, name in enumerate(names):
                    if name is None:
                        continue
                    p = person_map.get(name)
                    if p:
                        score_mod = self.db.get_person_score_modifier(p['id'])
                        total_count = self.db.get_person_total_count(p['id'])
                        seat_history_list = self.db.get_person_seat_history(p['id'])
                        seat_history = {r['app_name']: r['count'] for r in seat_history_list}  # 使用app_name
                        last_seat_app_name = p.get('last_seat_app_name')
                        all_person_objs.append(Person(p['id'], p['name'], p['level'], p['score'], True, score_mod, seat_history, total_count, True, position=pos_idx, last_seat_app_name=last_seat_app_name))
            
            # 再添加剩余人员
            for p in remaining_persons:
                p_dict = dict(p)
                score_mod = self.db.get_person_score_modifier(p_dict['id'])
                total_count = self.db.get_person_total_count(p_dict['id'])
                seat_history_list = self.db.get_person_seat_history(p_dict['id'])
                seat_history = {r['app_name']: r['count'] for r in seat_history_list}  # 使用app_name
                last_seat_app_name = p_dict.get('last_seat_app_name')
                all_person_objs.append(Person(p_dict['id'], p_dict['name'], p_dict['level'], p_dict['score'], True, score_mod, seat_history, total_count, False, last_seat_app_name=last_seat_app_name))
            
            seats = []
            for s in available_seats:
                seat = Seat(s['app_name'], bool(s.get('available', 1)), s.get('persons_count', 3), s.get('required_score', 5), s.get('template_id'))
                # 如果席位有模版，加载时段配置供scheduler使用
                if seat.template_id:
                    time_slots = self.db.get_template_time_slots(seat.template_id)
                    seat._time_slots = time_slots if time_slots else []
                seats.append(seat)
            
            # 预填充（包含位置索引）- 从 user_selections 构建
            # user_selections的key现在是app_name
            pre_filled = {}
            for seat_id, names in user_selections.items():
                # seat_id就是app_name（如'APP01'）
                seat_obj = next((s for s in seats if s.app_name == seat_id), None)
                if not seat_obj:
                    continue
                pre_filled[seat_id] = []
                for pos_idx, name in enumerate(names):
                    if name is None:
                        continue
                    # 从 all_person_objs 中找到对应的人员（锁定人员）
                    locked_person = next((p for p in all_person_objs if p.name == name and p.locked), None)
                    if locked_person:
                        pre_filled[seat_id].append(locked_person)
            
            self.status_bar.showMessage("正在排班...")
            QApplication.processEvents()
            
            # 传入所有人员（锁定+剩余）
            scheduler = ShiftScheduler(all_person_objs, seats)
            seats_dict, message = scheduler.generate_schedule_with_prefill(pre_filled)
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.schedule_result_display.setText(f"排班异常: {str(e)}")
            self.schedule_result_display.setStyleSheet("""
                QTextEdit {
                    background-color: #161b22;
                    color: #e74c3c;
                    border: 2px solid #30363d;
                    border-radius: 8px;
                    padding: 10px;
                    font-size: 12px;
                    font-family: "SF Mono", "Menlo", monospace;
                }
            """)
            self.status_bar.showMessage("排班出错")
            return
        
        # 计算已安排人数
        total_selected = len(all_selected_names)
        
        if seats_dict is None:
            # 排班失败，在结果显示框中显示
            self.schedule_result_display.setText(f"❌ 排班失败\n\n{message}")
            self.schedule_result_display.setStyleSheet("""
                QTextEdit {
                    background-color: #161b22;
                    color: #e74c3c;
                    border: 2px solid #30363d;
                    border-radius: 8px;
                    padding: 10px;
                    font-size: 12px;
                    font-family: "SF Mono", "Menlo", monospace;
                }
            """)
            self.status_bar.showMessage("排班失败")
        else:
            # 合并排班结果到 user_selections（保留手动选择的位置）
            # 使用排班返回的 Person 对象的 position 属性
            
            user_selections = {}  # 重新构建，格式: {app_name: [name1, name2, ...]}
            
            for seat_id, persons in seats_dict.items():
                # seat_id就是app_name
                seat_obj = next((s for s in seats if s.app_name == seat_id), None)
                max_count = seat_obj.persons_count if seat_obj else 4
                
                # 按位置索引构建列表
                seat_list = [None] * max_count
                for p in persons:
                    if p.position is not None and p.position < max_count:
                        seat_list[p.position] = p.name
                    else:
                        # 没有位置信息的尝试找空位
                        for i in range(max_count):
                            if seat_list[i] is None:
                                seat_list[i] = p.name
                                break
                
                user_selections[seat_id] = seat_list
            # 合并手动选择和自动排班结果

            # 更新数据层（格式：[A, B, C, D]）
            selections_for_ui = {}
            for seat_id, names in user_selections.items():
                # seat_id就是app_name字符串（如'APP01'）
                arr = [None, None, None, None]
                for i, name in enumerate(names):
                    if i < 4:
                        arr[i] = name
                selections_for_ui[seat_id] = arr
            
            self.schedule_data.selections = selections_for_ui
            
            # 先刷新下拉框（包含已选人员），在设置选择之前
            self._refresh_dropdowns()
            
            # 更新UI
            self._update_ui_from_selections(user_selections)
            
            total_placed = sum(len([n for n in v if n]) for v in user_selections.values())
            unplaced = len(active_persons) - total_placed
            
            # 在结果显示框中显示（不弹窗）
            if unplaced > 0 or message != "排班成功":
                result_text = f"✅ 已安排: {total_placed}人\n❌ 未安排: {unplaced}人\n\n{message}"
            else:
                result_text = "✅ 所有人员已安排完成！"
            
            self.schedule_result_display.setText(result_text)
            
            # 根据是否有告警调整样式
            if unplaced > 0 or "未安排" in message or "空位置" in message:
                self.schedule_result_display.setStyleSheet("""
                    QTextEdit {
                        background-color: #161b22;
                        color: #e74c3c;
                        border: 2px solid #30363d;
                        border-radius: 8px;
                        padding: 10px;
                        font-size: 12px;
                        font-family: "SF Mono", "Menlo", monospace;
                    }
                """)
            else:
                self.schedule_result_display.setStyleSheet("""
                    QTextEdit {
                        background-color: #161b22;
                        color: #2ecc71;
                        border: 2px solid #2ecc71;
                        border-radius: 8px;
                        padding: 10px;
                        font-size: 12px;
                        font-family: "SF Mono", "Menlo", monospace;
                    }
                """)
            
            self.status_bar.showMessage(f"排班完成 - 已安排{total_placed}人")
            self.refresh_preview()
    
    def _update_ui_from_selections(self, selections):
        """根据 selections 更新UI（排班完成后的同步）"""
        for widget in self.seat_result_widgets:
            seat_id = widget.seat.app_name
            if seat_id in selections:
                names = selections[seat_id]
                # 使用 blockSignals 阻止信号循环
                for combo in widget.person_combos:
                    combo.blockSignals(True)
                for i, name in enumerate(names):
                    if i < len(widget.person_combos):
                        idx = widget.person_combos[i].findData(name)
                        if idx >= 0:
                            widget.person_combos[i].setCurrentIndex(idx)
                for combo in widget.person_combos:
                    combo.blockSignals(False)
                widget.update_count()
        
        # 检查并更新所有下拉框的边框样式（规则违反时变红）
        for widget in self.seat_result_widgets:
            widget._check_and_update_border_style()
    
    def export_to_excel(self):
        try:
            from openpyxl import Workbook
            from PyQt5.QtWidgets import QFileDialog
            # 获取上次导出目录（使用数据库持久化）
            last_dir = self.db.get_setting('last_export_dir', os.path.expanduser('~'))
            if not os.path.exists(last_dir):
                last_dir = os.path.expanduser('~')

            file_path, _ = QFileDialog.getSaveFileName(self, "导出排班表", last_dir, "Excel文件 (*.xlsx)")
            if not file_path:
                return
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'

            # 保存导出目录到数据库
            self.db.set_setting('last_export_dir', os.path.dirname(file_path))

            # 获取席位人员数据
            seat_map = {}
            seat_persons_count = {}

            for seat_id, names in self.schedule_data.selections.items():
                # seat_id就是app_name (如'APP01')
                valid_names = [n for n in names if n]
                if not valid_names:
                    continue

                # 获取席位信息
                seat_info = self.schedule_data.seats.get(seat_id, {})
                app_name = seat_info.get('app', seat_id)

                # 跳过空app_name的席位
                if not app_name or app_name == "空":
                    continue

                seat_map[app_name] = valid_names
                seat_persons_count[app_name] = len(valid_names)

            # 填充排班表：2天周期
            # 早班+晚班 = 第1天，下午班 = 第2天
            # 4人席位：第1天ABC，第2天BCD（A休息）

            # 先统计执勤次数（不保存）
            # 先统计执勤次数
            duty_summary = []  # [(人员名, 席位名, 执勤次数), ...]

            for app_name, persons in seat_map.items():
                persons_count = seat_persons_count.get(app_name, 0)

                if not persons:
                    continue

                # 获取该席位的template_id和时段配置
                seat_info = self.schedule_data.seats.get(app_name, {})
                template_id = seat_info.get('template_id')

                # 获取时段配置，按shift_name分组
                shift_slots = {}
                if template_id:
                    slots = self.db.get_template_time_slots(template_id)
                    for slot in slots:
                        shift_name = slot.get('shift_name', '')
                        if shift_name:
                            if shift_name not in shift_slots:
                                shift_slots[shift_name] = []
                            shift_slots[shift_name].append(slot)

                # 统计每个人参与的班次
                for p in persons:
                    if not p:
                        continue
                    person = self.db.get_person_by_name(p)
                    if not person:
                        continue

                    # 判断参与的班次
                    has_morning_or_evening = False
                    has_afternoon = False

                    for shift_name, slots_list in shift_slots.items():
                        # 检查该班次是否有人在这个位置执勤
                        for slot in slots_list:
                            ctrl_pos = slot.get('ctrl_position', '')
                            asst_pos = slot.get('asst_position', '')

                            # 根据位置判断人员属于哪个时段
                            person_pos_idx = None
                            if p in persons:
                                try:
                                    person_pos_idx = persons.index(p)
                                except:
                                    continue

                            if person_pos_idx is not None:
                                pos_map = {0: 'A', 1: 'B', 2: 'C', 3: 'D'}
                                person_pos = pos_map.get(person_pos_idx, '')

                                if ctrl_pos == person_pos or asst_pos == person_pos:
                                    if shift_name in ['早班', '晚班']:
                                        has_morning_or_evening = True
                                    elif shift_name == '下午班':
                                        has_afternoon = True

                    duty_count = 0
                    if has_morning_or_evening:
                        duty_count += 1
                    if has_afternoon:
                        duty_count += 1

                    if duty_count > 0:
                        duty_summary.append((p, app_name, duty_count))

            # 显示统计确认对话框
            if duty_summary:
                summary_text = "即将更新以下人员的执勤次数：\n\n"
                for name, seat, count in duty_summary:
                    summary_text += f"• {name}: {seat} +{count}次\n"

                reply = CustomMessageBox.question(self, "确认执勤统计", summary_text + "\n是否确认保存？")
                if not reply:
                    return
            
            # 导出Excel文件
            from openpyxl import Workbook
            from openpyxl.styles import Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "排班结果"
            
            # 按app_name排序输出，每个席位分块
            for app_name in sorted(seat_map.keys()):
                persons = seat_map.get(app_name, [])
                actual_count = len([p for p in persons if p])
                
                # 获取时段配置
                seat_info = self.schedule_data.seats.get(app_name, {})
                template_id = seat_info.get('template_id')
                
                slots_by_shift = {}
                if template_id:
                    slots = self.db.get_template_time_slots(template_id)
                    # 按shift_name分组
                    for slot in slots:
                        shift = slot.get('shift_name', '')
                        if shift not in slots_by_shift:
                            slots_by_shift[shift] = []
                        slots_by_shift[shift].append(slot)
                
                # 获取助手函数
                def get_person(persons, pos):
                    if not persons:
                        return ''
                    if pos == 'A':
                        return persons[0] if len(persons) > 0 else ''
                    elif pos == 'B':
                        return persons[1] if len(persons) > 1 else ''
                    elif pos == 'C':
                        return persons[2] if len(persons) > 2 else ''
                    return ''
                
                # 获取当前行号（用于设置居中）
                current_row = ws.max_row + 1
                
                # 第一行：席位名称（居中）
                seat_row = [app_name, '', '', '', '', '']
                ws.append(seat_row)
                
                # 设置第一行居中
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
                
                current_row += 1
                
                # 第二行：表头
                header_row = ['', '班次', '管制时段', '管制席', '助理时段', '助理席']
                ws.append(header_row)
                
                # 后续行：每个班次只显示第一次
                last_shift = ''
                for shift_name in ['早班', '晚班', '下午班']:
                    slots_list = slots_by_shift.get(shift_name, [])
                    if not slots_list:
                        continue
                    
                    # 显示该班次所有时段
                    for slot in slots_list:
                        ctrl_time = slot.get('ctrl_time', '')
                        ctrl_pos = slot.get('ctrl_position', '')
                        asst_time = slot.get('asst_time', '')
                        asst_pos = slot.get('asst_position', '')

                        # 少于3人时，每人按顺序填
                        if actual_count < 3:
                            ctrl_person = persons[0] if len(persons) > 0 else ''
                            asst_person = persons[1] if len(persons) > 1 else ''
                        else:
                            ctrl_person = get_person(persons, ctrl_pos) if ctrl_pos else ''
                            asst_person = get_person(persons, asst_pos) if asst_pos else ''

                        # 连续重复的班次留空
                        shift_display = '' if shift_name == last_shift else shift_name
                        last_shift = shift_name

                        row_data = [
                            '',  # 席位列留空
                            shift_display,
                            ctrl_time,
                            ctrl_person,
                            asst_time,
                            asst_person,
                        ]
                        ws.append(row_data)
            
            today = datetime.now().strftime("%Y-%m-%d")
            
            # 构建 schedule_history 记录（同时保存到 schedule_history 和 person_duty_stats）
            # 构建 history_records（使用确认过的 duty_summary，确保数据一致）
            # duty_summary: [(person_name, app_name, duty_count), ...]
            person_duty_map = {}  # {(name, app_name): duty_count}
            for name, app_name, duty_count in duty_summary:
                person_duty_map[(name, app_name)] = duty_count

            history_records = []  # [(person_id, seat_id, app_name, time_slot, duration_minutes), ...]
            for app_name, persons in seat_map.items():
                if not persons:
                    continue

                for p in persons:
                    if not p:
                        continue
                    person = self.db.get_person_by_name(p)
                    if not person:
                        continue
                    # seat_id 在 main_UI 中等于 app_name
                    # duration 使用确认对话框中计算的实际执勤次数
                    duration = person_duty_map.get((p, app_name), 1)
                    history_records.append((person['id'], app_name, app_name, f'{app_name}-导出', duration))

            self.db.save_schedule_with_seat_ids(today, history_records)

            wb.save(file_path)

            # 刷新统计数据
            self.update_stats()
            self.refresh_duty_data()

            CustomMessageBox.information(self, "导出成功", f"排班表已导出到:\n{file_path}")
            self.status_bar.showMessage(f"已导出: {file_path}")
        
        except ImportError:
            CustomMessageBox.warning(self, "导出失败", "请先安装 openpyxl 库：\npip install openpyxl")
        except Exception as e:
            CustomMessageBox.warning(self, "导出失败", str(e))
    
    def closeEvent(self, event):
        self.db.close()
        event.accept()


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setStyleSheet("""
        /* === 全局主题 === */
        * {
            background-color: #0d1117;
            color: #c9d1d9;
            font-family: "SF Mono", "Menlo", "Monaco", monospace;
            font-size: 12px;
        }
        
        /* === 标题和文字 === */
        QMainWindow {
            background-color: #0d1117;
        }
        QLabel {
            color: #58a6ff;
            font-weight: bold;
        }
        
        /* === 对话框 === */
        QDialog {
            background-color: #161b22;
            color: #c9d1d9;
        }
        QDialog QLabel {
            color: #8b949e;
        }
        
        /* === 输入框 === */
        QLineEdit, QComboBox, QSpinBox {
            background-color: #21262d;
            color: #c9d1d9;
            border: 1px solid #30363d;
            border-radius: 6px;
            padding: 8px 12px;
            selection-background-color: #1f6feb;
        }
        QLineEdit:focus, QComboBox:focus {
            border-color: #388bfd;
        }
        QTextEdit {
            background-color: #21262d;
            color: #c9d1d9;
            border: 1px solid #30363d;
            border-radius: 6px;
        }
        
        /* === 下拉框展开列表 === */
        QComboBox QAbstractItemView {
            background-color: #161b22;
            color: #c9d1d9;
            border: 1px solid #30363d;
            selection-background-color: #1f6feb;
            selection-color: #ffffff;
        }
        
        /* === 按钮 === */
        QPushButton {
            background-color: #238636;
            color: #ffffff;
            border: none;
            border-radius: 6px;
            padding: 10px 20px;
            font-weight: bold;
        }
        QPushButton:hover {
            background-color: #2ea043;
        }
        QPushButton:pressed {
            background-color: #196c2e;
        }
        
        /* === 表格 === */
        QTableWidget {
            background-color: #0d1117;
            color: #c9d1d9;
            gridline-color: #30363d;
            border: 1px solid #30363d;
            border-radius: 8px;
        }
        QTableWidget::item {
            background-color: #0d1117;
            color: #c9d1d9;
            padding: 8px;
            border: none;
        }
        QTableWidget::item:alternate {
            background-color: #161b22;
        }
        QTableWidget::item:selected {
            background-color: #1f6feb;
            color: #ffffff;
        }
        
        /* === 表格表头 === */
        QHeaderView::section {
            background-color: #21262d;
            color: #58a6ff;
            padding: 10px;
            border: none;
            border-bottom: 2px solid #388bfd;
            font-weight: bold;
        }
        QTableWidget QHeaderView::section:horizontal {
            background-color: #21262d;
            color: #58a6ff;
            padding: 12px;
            border: none;
            border-bottom: 2px solid #388bfd;
            font-weight: bold;
        }
        QTableWidget QHeaderView::section:vertical {
            background-color: #21262d;
            color: #58a6fb;
            padding: 10px;
            border: none;
            border-right: 2px solid #388bfd;
            font-weight: bold;
        }
        
        /* === 分组框 === */
        QGroupBox {
            background-color: #161b22;
            color: #58a6ff;
            border: 1px solid #30363d;
            border-radius: 10px;
            margin-top: 15px;
            padding-top: 10px;
            font-weight: bold;
            font-size: 14px;
        }
        QGroupBox::title {
            color: #58a6ff;
            subcontrol-origin: margin;
            left: 15px;
            padding: 0 10px;
        }
        
        /* === 消息框 === */
        QMessageBox {
            background-color: #161b22;
        }
        QMessageBox QLabel {
            color: #c9d1d9;
            font-size: 14px;
        }
        
        /* === 选项框 === */
        QCheckBox {
            color: #c9d1d9;
            spacing: 8px;
        }
        QCheckBox::indicator {
            width: 18px;
            height: 18px;
            border: 2px solid #30363d;
            border-radius: 4px;
            background-color: #21262d;
        }
        QCheckBox::indicator:checked {
            background-color: #238636;
            border-color: #238636;
        }
        
        /* === 滚动条 === */
        QScrollBar:vertical {
            background-color: #21262d;
            width: 12px;
            border-radius: 6px;
        }
        QScrollBar::handle:vertical {
            background-color: #484f58;
            border-radius: 6px;
            min-height: 20px;
        }
        QScrollBar::handle:vertical:hover {
            background-color: #6e7681;
        }
        QScrollBar:horizontal {
            background-color: #21262d;
            height: 12px;
            border-radius: 6px;
        }
        QScrollBar::handle:horizontal {
            background-color: #484f58;
            border-radius: 6px;
            min-width: 20px;
        }
    """)
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()