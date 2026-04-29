#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
排班管理软件 - 西南空管局版本
功能：管理人员等级(C1/C2/C3)，根据约束条件自动排班
数据存储：SQLite数据库
约束：C1与C1不可同席，C2与C1不可同席
"""

import sys
import random
import sqlite3
import os
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QTableWidget, QTableWidgetItem, QPushButton, QLabel,
    QLineEdit, QComboBox, QCheckBox, QGroupBox, QFormLayout, QMessageBox,
    QSpinBox, QDoubleSpinBox, QTextEdit, QStatusBar, QDialog, QDialogButtonBox,
    QFrame, QGridLayout, QStyledItemDelegate, QScrollArea, QHeaderView,
    QSizePolicy
)
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QFont, QColor

# 获取脚本所在目录，用于数据库文件
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DB_FILE = os.path.join(SCRIPT_DIR, "scheduler.db")


class Database:
    """数据库管理类"""
    
    def __init__(self):
        self.conn = sqlite3.connect(DB_FILE)
        self.conn.row_factory = sqlite3.Row
        self.create_tables()
        self.init_default_seats()
        self.init_default_rules()
        self.init_default_time_slots()
        self.init_test_data_if_empty()
    
    def create_tables(self):
        """创建数据表"""
        cursor = self.conn.cursor()

        # 人员表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS persons (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                level TEXT NOT NULL,
                score INTEGER DEFAULT 0,
                active INTEGER DEFAULT 1,
                locked INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 为已存在的数据库添加locked列（如果不存在）
        cursor.execute("PRAGMA table_info(persons)")
        columns = [col[1] for col in cursor.fetchall()]
        if 'locked' not in columns:
            cursor.execute("ALTER TABLE persons ADD COLUMN locked INTEGER DEFAULT 0")

        # 席位表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS seats (
                id INTEGER PRIMARY KEY,
                name TEXT NOT NULL,
                app_name TEXT,
                available INTEGER DEFAULT 1,
                persons_count INTEGER DEFAULT 3,
                required_score INTEGER DEFAULT 5
            )
        ''')

        # 为已存在的数据库添加app_name列（如果不存在）
        cursor.execute("PRAGMA table_info(seats)")
        columns = [col[1] for col in cursor.fetchall()]
        if 'app_name' not in columns:
            cursor.execute("ALTER TABLE seats ADD COLUMN app_name TEXT")

        # 规则表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS rules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                description TEXT,
                score_modifier REAL DEFAULT 0,
                active INTEGER DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # 人员规则关联表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS person_rules (
                person_id INTEGER,
                rule_id INTEGER,
                enabled INTEGER DEFAULT 1,
                PRIMARY KEY (person_id, rule_id),
                FOREIGN KEY (person_id) REFERENCES persons(id) ON DELETE CASCADE,
                FOREIGN KEY (rule_id) REFERENCES rules(id) ON DELETE CASCADE
            )
        ''')

        # 排班历史记录表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS schedule_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                schedule_date DATE NOT NULL,
                person_id INTEGER NOT NULL,
                seat_id INTEGER NOT NULL,
                time_slot TEXT NOT NULL,
                duration_minutes INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (person_id) REFERENCES persons(id) ON DELETE CASCADE,
                FOREIGN KEY (seat_id) REFERENCES seats(id) ON DELETE CASCADE
            )
        ''')

        # 时间段配置表
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS time_slots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                duration_minutes INTEGER NOT NULL,
                shift_type TEXT NOT NULL,
                seat_position TEXT NOT NULL
            )
        ''')

        self.conn.commit()
    
    def init_default_seats(self):
        """初始化默认5个席位"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM seats")
        if cursor.fetchone()[0] == 0:
            default_seats = [
                (1, "席位一", "APP01", 1, 3, 5),
                (2, "席位二", "APP02", 1, 3, 5),
                (3, "席位三", "APP03", 1, 3, 5),
                (4, "席位四", "APP04", 1, 3, 5),
                (5, "席位五", "APP05", 1, 3, 5),
            ]
            cursor.executemany(
                "INSERT OR IGNORE INTO seats (id, name, app_name, available, persons_count, required_score) VALUES (?, ?, ?, ?, ?, ?)",
                default_seats
            )
            self.conn.commit()
    
    def init_test_data_if_empty(self):
        """初始化测试数据（如果数据库为空）"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM persons")
        if cursor.fetchone()[0] == 0:
            # 等级对应的分数
            level_scores = {"C1": 1, "C2": 2, "C3": 3, "I": 4, "S": 4}
            
            # 添加测试人员 - C1/C2/C3/I/S 各等级
            test_persons = [
                ("张伟", "C1", 1), ("李强", "C1", 1),
                ("王芳", "C2", 1), ("刘洋", "C2", 1), ("陈明", "C2", 1),
                ("赵敏", "C3", 1), ("黄磊", "C3", 1), ("周涛", "C3", 1), ("吴静", "C3", 1),
                ("郑凯", "I", 1), ("孙悦", "I", 1), ("钱磊", "I", 1), ("赵强", "I", 1), ("王涛", "I", 1),
                ("冯琳", "S", 1), ("陈晨", "S", 1), ("林雪", "S", 1),
            ]
            # 添加分数
            test_data = [(name, level, level_scores.get(level, 0), active) for name, level, active in test_persons]
            cursor.executemany(
                "INSERT INTO persons (name, level, score, active) VALUES (?, ?, ?, ?)",
                test_data
            )
            self.conn.commit()
    
    def get_all_persons(self):
        """获取所有人员"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM persons ORDER BY level, name")
        return cursor.fetchall()
    
    def get_active_persons(self):
        """获取参与排班的人员"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM persons WHERE active = 1 ORDER BY level, name")
        return cursor.fetchall()
    
    def get_available_seats(self):
        """获取可用的席位"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM seats WHERE available = 1 ORDER BY id")
        rows = cursor.fetchall()
        # 转换为字典
        return [dict(row) for row in rows]
    
    def get_all_seats(self):
        """获取所有席位"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM seats ORDER BY id")
        rows = cursor.fetchall()
        # 转换为字典，因为sqlite3.Row不支持get方法
        return [dict(row) for row in rows]
    
    def update_seat_available(self, seat_id, available):
        """更新席位可用状态"""
        cursor = self.conn.cursor()
        cursor.execute("UPDATE seats SET available = ? WHERE id = ?", (available, seat_id))
        self.conn.commit()
    
    def update_seat_persons_count(self, seat_id, count):
        """更新席位人数"""
        cursor = self.conn.cursor()
        cursor.execute("UPDATE seats SET persons_count = ? WHERE id = ?", (count, seat_id))
        self.conn.commit()
    
    def update_seat_required_score(self, seat_id, required_score):
        """更新席位要求分数"""
        cursor = self.conn.cursor()
        cursor.execute("UPDATE seats SET required_score = ? WHERE id = ?", (required_score, seat_id))
        self.conn.commit()
    
    def update_seat_app_name(self, seat_id, app_name):
        """更新席位APP名称"""
        cursor = self.conn.cursor()
        cursor.execute("UPDATE seats SET app_name = ? WHERE id = ?", (app_name, seat_id))
        self.conn.commit()
    
    def add_person(self, name, level, active=1):
        """添加人员"""
        try:
            cursor = self.conn.cursor()
            cursor.execute('''
                INSERT INTO persons (name, level, active)
                VALUES (?, ?, ?)
            ''', (name, level, active))
            self.conn.commit()
            return True, "添加成功"
        except sqlite3.IntegrityError:
            return False, "人员已存在"
        except Exception as e:
            return False, str(e)
    
    def update_person(self, person_id, name, level, active, locked=None):
        """更新人员信息"""
        try:
            cursor = self.conn.cursor()
            if locked is not None:
                cursor.execute('''
                    UPDATE persons
                    SET name = ?, level = ?, active = ?, locked = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (name, level, active, locked, person_id))
            else:
                cursor.execute('''
                    UPDATE persons
                    SET name = ?, level = ?, active = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (name, level, active, person_id))
            self.conn.commit()
            return True, "更新成功"
        except sqlite3.IntegrityError:
            return False, "人员已存在"
        except Exception as e:
            return False, str(e)

    def update_person_locked(self, person_id, locked):
        """更新人员锁定状态"""
        cursor = self.conn.cursor()
        cursor.execute("UPDATE persons SET locked = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?", (locked, person_id))
        self.conn.commit()
        try:
            cursor = self.conn.cursor()
            cursor.execute("DELETE FROM persons WHERE id=?", (person_id,))
            self.conn.commit()
            return True, "删除成功"
        except Exception as e:
            return False, str(e)
    
    def get_person_by_id(self, person_id):
        """根据ID获取人员"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM persons WHERE id=?", (person_id,))
        return cursor.fetchone()

    def get_person_by_name(self, name):
        """根据姓名获取人员"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM persons WHERE name=?", (name,))
        return cursor.fetchone()

    def update_person_active(self, person_id, active):
        """更新人员参与排班状态"""
        cursor = self.conn.cursor()
        cursor.execute("UPDATE persons SET active = ? WHERE id = ?", (active, person_id))
        self.conn.commit()

    def search_persons(self, keyword):
        """搜索人员"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT * FROM persons 
            WHERE name LIKE ?
            ORDER BY level, name
        ''', (f'%{keyword}%',))
        return cursor.fetchall()
    
    def get_statistics(self):
        """获取统计信息"""
        cursor = self.conn.cursor()
        
        cursor.execute("SELECT COUNT(*) as total FROM persons")
        total = cursor.fetchone()['total']
        
        cursor.execute("SELECT COUNT(*) as active FROM persons WHERE active = 1")
        active = cursor.fetchone()['active']
        
        cursor.execute("SELECT level, COUNT(*) as count FROM persons WHERE active = 1 GROUP BY level")
        level_stats = {row['level']: row['count'] for row in cursor.fetchall()}
        
        cursor.execute("SELECT COUNT(*) as available FROM seats WHERE available = 1")
        available_seats = cursor.fetchone()['available']
        
        return {
            'total': total,
            'active': active,
            'c1': level_stats.get('C1', 0),
            'c2': level_stats.get('C2', 0),
            'c3': level_stats.get('C3', 0),
            'i': level_stats.get('I', 0),
            's': level_stats.get('S', 0),
            'available_seats': available_seats
        }

    # ==================== 规则管理 ====================

    def get_all_rules(self):
        """获取所有规则"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM rules ORDER BY id")
        return cursor.fetchall()

    def get_active_rules(self):
        """获取启用的规则"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM rules WHERE active = 1 ORDER BY id")
        return cursor.fetchall()

    def add_rule(self, name, description, score_modifier):
        """添加规则"""
        try:
            cursor = self.conn.cursor()
            cursor.execute('''
                INSERT INTO rules (name, description, score_modifier, active)
                VALUES (?, ?, ?, 1)
            ''', (name, description, score_modifier))
            self.conn.commit()
            return True, "添加成功"
        except sqlite3.IntegrityError:
            return False, "规则已存在"
        except Exception as e:
            return False, str(e)

    def update_rule(self, rule_id, name, description, score_modifier, active):
        """更新规则"""
        try:
            cursor = self.conn.cursor()
            cursor.execute('''
                UPDATE rules
                SET name = ?, description = ?, score_modifier = ?, active = ?
                WHERE id = ?
            ''', (name, description, score_modifier, active, rule_id))
            self.conn.commit()
            return True, "更新成功"
        except sqlite3.IntegrityError:
            return False, "规则已存在"
        except Exception as e:
            return False, str(e)

    def delete_rule(self, rule_id):
        """删除规则"""
        try:
            cursor = self.conn.cursor()
            cursor.execute("DELETE FROM rules WHERE id = ?", (rule_id,))
            self.conn.commit()
            return True, "删除成功"
        except Exception as e:
            return False, str(e)

    def get_rule_by_id(self, rule_id):
        """获取规则"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT * FROM rules WHERE id = ?", (rule_id,))
        return cursor.fetchone()

    def get_person_rules(self, person_id):
        """获取人员的规则"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT r.*, pr.enabled
            FROM rules r
            LEFT JOIN person_rules pr ON r.id = pr.rule_id AND pr.person_id = ?
            ORDER BY r.id
        ''', (person_id,))
        return cursor.fetchall()

    def set_person_rule(self, person_id, rule_id, enabled):
        """设置人员的规则"""
        try:
            cursor = self.conn.cursor()
            cursor.execute('''
                INSERT OR REPLACE INTO person_rules (person_id, rule_id, enabled)
                VALUES (?, ?, ?)
            ''', (person_id, rule_id, enabled))
            self.conn.commit()
            return True
        except Exception as e:
            return False

    def get_person_score_modifier(self, person_id):
        """获取人员的规则分数修改值"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT COALESCE(SUM(r.score_modifier), 0) as total
            FROM person_rules pr
            JOIN rules r ON pr.rule_id = r.id
            WHERE pr.person_id = ? AND pr.enabled = 1 AND r.active = 1
        ''', (person_id,))
        return cursor.fetchone()['total']

    def init_default_rules(self):
        """初始化默认规则（如果为空）"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM rules")
        if cursor.fetchone()[0] == 0:
            default_rules = [
                ("生病", "因生病扣除分数", -0.1),
                ("请假", "因请假扣除分数", -0.2),
                ("培训", "参加培训扣除分数", -0.15),
                ("出差", "因出差扣除分数", -0.1),
                ("值班超时", "值班超时奖励", 0.1),
            ]
            cursor.executemany(
                "INSERT INTO rules (name, description, score_modifier) VALUES (?, ?, ?)",
                default_rules
            )
            self.conn.commit()

    def init_default_time_slots(self):
        """初始化默认时间段（如果为空）"""
        cursor = self.conn.cursor()
        cursor.execute("SELECT COUNT(*) FROM time_slots")
        if cursor.fetchone()[0] == 0:
            # 从Excel样表解析的时间段
            default_slots = [
                # 早班 - APP04
                ("早班-时段1", "0850", "1040", 110, "早班", "APP04-管制席"),
                ("早班-时段2", "1030", "1200", 90, "早班", "APP04-管制席"),
                ("早班-时段3", "1150", "1240", 50, "早班", "APP04-管制席"),
                # 早班 - APP04 助理席
                ("早班-时段1", "0840", "1000", 80, "早班", "APP04-助理席"),
                ("早班-时段2", "0950", "1120", 90, "早班", "APP04-助理席"),
                ("早班-时段3", "1110", "1230", 80, "早班", "APP04-助理席"),
                # 早班 - APP05
                ("早班-时段1", "0840", "1030", 110, "早班", "APP05-管制席"),
                ("早班-时段2", "1020", "1150", 90, "早班", "APP05-管制席"),
                ("早班-时段3", "1140", "1230", 50, "早班", "APP05-管制席"),
                # 早班 - APP05 助理席
                ("早班-时段1", "0850", "0950", 60, "早班", "APP05-助理席"),
                ("早班-时段2", "0940", "1110", 90, "早班", "APP05-助理席"),
                ("早班-时段3", "1100", "1240", 100, "早班", "APP05-助理席"),
                # 晚班 - APP04
                ("晚班-时段1", "1800", "1940", 100, "晚班", "APP04-管制席"),
                ("晚班-时段2", "1930", "2100", 90, "晚班", "APP04-管制席"),
                ("晚班-时段3", "2050", "2220", 90, "晚班", "APP04-管制席"),
                # 晚班 - APP04 助理席
                ("晚班-时段1", "1750", "1900", 70, "晚班", "APP04-助理席"),
                ("晚班-时段2", "1850", "2020", 90, "晚班", "APP04-助理席"),
                ("晚班-时段3", "2010", "2140", 90, "晚班", "APP04-助理席"),
                ("晚班-时段4", "2130", "2210", 40, "晚班", "APP04-助理席"),
                # 晚班 - APP05
                ("晚班-时段1", "1750", "1930", 100, "晚班", "APP05-管制席"),
                ("晚班-时段2", "1920", "2050", 90, "晚班", "APP05-管制席"),
                ("晚班-时段3", "2040", "2210", 90, "晚班", "APP05-管制席"),
                # 晚班 - APP05 助理席
                ("晚班-时段1", "1800", "1850", 50, "晚班", "APP05-助理席"),
                ("晚班-时段2", "1840", "2010", 90, "晚班", "APP05-助理席"),
                ("晚班-时段3", "2000", "2130", 90, "晚班", "APP05-助理席"),
                ("晚班-时段4", "2120", "2220", 60, "晚班", "APP05-助理席"),
                # 下午班 - APP04
                ("下午班-时段1", "1230", "1400", 90, "下午班", "APP04-管制席"),
                ("下午班-时段2", "1350", "1550", 120, "下午班", "APP04-管制席"),
                ("下午班-时段3", "1540", "1730", 110, "下午班", "APP04-管制席"),
                ("下午班-时段4", "1720", "1810", 50, "下午班", "APP04-管制席"),
                # 下午班 - APP04 助理席
                ("下午班-时段1", "1220", "1320", 60, "下午班", "APP04-助理席"),
                ("下午班-时段2", "1310", "1500", 110, "下午班", "APP04-助理席"),
                ("下午班-时段3", "1450", "1650", 120, "下午班", "APP04-助理席"),
                ("下午班-时段4", "1640", "1800", 80, "下午班", "APP04-助理席"),
                # 下午班 - APP05
                ("下午班-时段1", "1220", "1350", 90, "下午班", "APP05-管制席"),
                ("下午班-时段2", "1340", "1540", 120, "下午班", "APP05-管制席"),
                ("下午班-时段3", "1530", "1720", 110, "下午班", "APP05-管制席"),
                ("下午班-时段4", "1710", "1800", 50, "下午班", "APP05-管制席"),
                # 下午班 - APP05 助理席
                ("下午班-时段1", "1230", "1310", 40, "下午班", "APP05-助理席"),
                ("下午班-时段2", "1300", "1450", 110, "下午班", "APP05-助理席"),
                ("下午班-时段3", "1440", "1640", 120, "下午班", "APP05-助理席"),
                ("下午班-时段4", "1630", "1810", 100, "下午班", "APP05-助理席"),
            ]
            cursor.executemany(
                "INSERT INTO time_slots (name, start_time, end_time, duration_minutes, shift_type, seat_position) VALUES (?, ?, ?, ?, ?, ?)",
                default_slots
            )
            self.conn.commit()

    def get_person_seat_history(self, person_id, limit=10):
        """获取人员在各席位的最近排班历史（用于均衡）"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT s.id as seat_id, s.name as seat_name,
                   COUNT(*) as count, SUM(duration_minutes) as total_minutes
            FROM schedule_history h
            JOIN seats s ON h.seat_id = s.id
            WHERE h.person_id = ?
            GROUP BY s.id
            ORDER BY count DESC
        ''', (person_id,))
        return cursor.fetchall()

    def get_person_total_hours(self, person_id, days=30):
        """获取人员最近N天的总上班时长（分钟）"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT COALESCE(SUM(duration_minutes), 0) as total
            FROM schedule_history 
            WHERE person_id = ? AND schedule_date >= date('now', '-' || ? || ' days')
        ''', (person_id, days))
        return cursor.fetchone()['total']

    def get_seat_assignment_count(self, seat_id, days=30):
        """获取席位最近N天的安排次数"""
        cursor = self.conn.cursor()
        cursor.execute('''
            SELECT COUNT(*) as cnt FROM schedule_history 
            WHERE seat_id = ? AND schedule_date >= date('now', '-' || ? || ' days')
        ''', (seat_id, days))
        return cursor.fetchone()['cnt']

    def save_schedule(self, schedule_date, assignments):
        """保存排班结果到历史记录
        assignments: [(person_id, seat_id, time_slot, duration_minutes), ...]
        """
        cursor = self.conn.cursor()
        # 先删除当天的记录
        cursor.execute("DELETE FROM schedule_history WHERE schedule_date = ?", (schedule_date,))
        # 插入新记录
        cursor.executemany('''
            INSERT INTO schedule_history (schedule_date, person_id, seat_id, time_slot, duration_minutes)
            VALUES (?, ?, ?, ?, ?)
        ''', assignments)
        self.conn.commit()

    def get_time_slots(self, shift_type=None):
        """获取时间段配置"""
        cursor = self.conn.cursor()
        if shift_type:
            cursor.execute("SELECT * FROM time_slots WHERE shift_type = ? ORDER BY seat_position, start_time", (shift_type,))
        else:
            cursor.execute("SELECT * FROM time_slots ORDER BY shift_type, seat_position, start_time")
        return cursor.fetchall()

    def close(self):
        """关闭数据库连接"""
        self.conn.close()


class Person:
    """人员类(用于排班)"""
    def __init__(self, id, name, level, score=0, active=True, score_modifier=0, seat_history=None, total_hours=0, locked=False):
        self.id = id
        self.name = name
        self.level = level
        self.score = score
        self.active = active
        self.score_modifier = score_modifier  # 规则分数修改
        self.seat_history = seat_history or {}  # {seat_id: count}
        self.total_hours = total_hours  # 最近30天总时长(分钟)
        self.locked = locked  # 是否锁定（手动指定的排班不动）

    @property
    def effective_score(self):
        """实际参与排班的分数（基础分 + 规则修改）"""
        return self.score + self.score_modifier

    def is_c_level(self):
        """是否为C1或C2"""
        return self.level in ("C1", "C2")


class Seat:
    """席位类"""
    def __init__(self, id, name, app_name=None, available=True, persons_count=3, required_score=5):
        self.id = id
        self.name = name
        self.app_name = app_name or name.replace('席位', 'APP0')
        self.available = available
        self.persons_count = persons_count
        self.required_score = required_score


class ShiftScheduler:
    """排班调度器"""
    
    def __init__(self, persons, seats):
        self.persons = persons
        self.seats = seats
    
    def can_place_together(self, p1, p2):
        """检查两人是否可以同席"""
        # C1 单独使用席位（不能与C1或C2同席）
        if p1.level == "C1" and (p2.level == "C1" or p2.level == "C2"):
            return False
        if p2.level == "C1" and (p1.level == "C1" or p1.level == "C2"):
            return False
        
        # C2 可与C2或C3同席
        if p1.level == "C2" and p2.level not in ("C2", "C3", "I", "S"):
            return False
        if p2.level == "C2" and p1.level not in ("C2", "C3", "I", "S"):
            return False
        
        # I 和 S 等级可以与任意等级同席
        return True

    def calc_balance_score(self, person, seat_id, seats_dict):
        """计算人员到席位的均衡分数（越低越优先选择）
        考虑：1. 该席位历史安排次数越少越好
              2. 人员总工时越多越不优先
              3. 同一人同一席位历史次数越少越好
        """
        score = 0

        # 同一席位历史安排次数越多，得分越高（不优先）
        history_seat_count = person.seat_history.get(seat_id, 0)
        score += history_seat_count * 10

        # 人员总工时（越多越不优先）
        score += person.total_hours / 60 * 2  # 每小时加2分

        return score

    def select_best_seat(self, person, seats_dict):
        """为人员选择最优席位"""
        best_seat = None
        best_score = float('inf')

        for seat in self.seats:
            if not seat.available:
                continue
            if len(seats_dict.get(seat.id, [])) >= seat.persons_count:
                continue

            # 检查是否能加入
            if not self.can_add_to_seat(seats_dict.get(seat.id, []), person, seat.required_score):
                continue

            # 计算均衡分数
            balance_score = self.calc_balance_score(person, seat.id, seats_dict)
            if balance_score < best_score:
                best_score = balance_score
                best_seat = seat

        return best_seat

    def can_add_to_seat(self, seat_people, new_person, required_score=5):
        """检查是否可以添加到现有席位"""
        # 先检查等级兼容性
        for p in seat_people:
            if not self.can_place_together(p, new_person):
                return False

        # 检查C1/C2人数不超过席位的50%
        # 如果席位3人，则C1+C2最多1人（向下取整）
        # 注意：空席时允许第1个C级加入，50%限制只对后续加入生效
        if new_person.is_c_level() and len(seat_people) > 0:
            seat_capacity = len(seat_people) + 1  # 当前人数+新人
            max_c = seat_capacity // 2
            current_c_count = sum(1 for p in seat_people if p.is_c_level())
            if current_c_count >= max_c:
                return False

        # 检查分数是否满足要求（两两分数相加需>required_score）
        # 席位有3人时，每对组合都需满足
        # 只有1人时，分数门槛减半（两人刚开始配合，要求稍宽松）
        if len(seat_people) >= 2:
            all_people = seat_people + [new_person]
            from itertools import combinations
            for p1, p2 in combinations(all_people, 2):
                if p1.effective_score + p2.effective_score <= required_score:
                    return False
        elif len(seat_people) == 1:
            # 只有1人时，分数需要>required_score的一半（向上取整）
            import math
            threshold = math.ceil(required_score / 2)
            if seat_people[0].effective_score + new_person.effective_score <= threshold:
                return False

        return True
    
    def check_seat_score(self, seat_people, required_score):
        """检查席位内所有人员是否满足分数要求"""
        if len(seat_people) < 2:
            return True

        from itertools import combinations
        for p1, p2 in combinations(seat_people, 2):
            if p1.effective_score + p2.effective_score <= required_score:
                return False
        return True
    
    def generate_schedule(self):
        """生成排班表 - 使用均衡算法安排人员"""
        available = [p for p in self.persons if p.active]
        total_capacity = sum(s.persons_count for s in self.seats if s.available)

        best_seats = None
        best_balance_score = float('inf')
        best_assigned = 0

        for attempt in range(2000):
            seats = {s.id: [] for s in self.seats if s.available}
            assigned = set()

            # 打乱顺序但保持均衡考虑
            persons_copy = list(available)
            random.shuffle(persons_copy)

            # 按均衡分数排序，优先安排分数高（需要优先安排）的人员
            # 分数高意味着：历史工时少、同一席位安排少
            persons_copy.sort(key=lambda p: (
                p.total_hours,  # 工时少的先安排
                -sum(p.seat_history.get(s.id, 0) for s in self.seats)  # 历史席位次数少的先安排
            ))

            # 尝试安排每个人
            for person in persons_copy:
                # 使用均衡选择
                best_seat = self.select_best_seat(person, seats)
                if best_seat:
                    seats[best_seat.id].append(person)
                    assigned.add(person.name)

            # 验证约束
            valid = True
            balance_score = 0
            for seat in self.seats:
                if not seat.available:
                    continue
                seat_people = seats.get(seat.id, [])
                # 检查等级兼容性
                for i, p1 in enumerate(seat_people):
                    for p2 in seat_people[i+1:]:
                        if not self.can_place_together(p1, p2):
                            valid = False
                            break
                    if not valid:
                        break
                if not valid:
                    break
                # 检查分数要求
                if not self.check_seat_score(seat_people, seat.required_score):
                    valid = False
                    break
                # 计算均衡分数
                for p in seat_people:
                    balance_score += p.total_hours

            if valid:
                current_assigned = len(assigned)
                # 优先选择安排人数多的，其次选择均衡分数低的
                if current_assigned > best_assigned or \
                   (current_assigned == best_assigned and balance_score < best_balance_score):
                    best_assigned = current_assigned
                    best_seats = seats
                    best_balance_score = balance_score

                if current_assigned >= total_capacity:
                    break

        if best_seats:
            # 提取所有已安排人员的名字
            assigned_names = {person.name for persons in best_seats.values() for person in persons}
            unplaced = [p for p in available if p.name not in assigned_names]

            placed_count = sum(len(v) for v in best_seats.values())
            if len(unplaced) == 0:
                return best_seats, "排班成功"
            else:
                # 显示未安排的人员
                unplaced_names = [p.name for p in unplaced[:3]]
                msg = f"已安排{placed_count}人/"
                if len(unplaced) > 3:
                    msg += f"还有{len(unplaced)}人待安排: {','.join(unplaced_names)}..."
                else:
                    msg += f"未安排: {','.join(unplaced_names)}"
                return best_seats, msg

        return None, "无法生成有效的排班方案"
    
    def generate_schedule_with_prefill(self, pre_filled):
        """生成排班表 - 考虑预先安排的人员，锁定人员优先
        pre_filled: {seat_id: [Person objects]}
        锁定的人员(locked=True)不会被移动，只填充剩余位置
        """
        all_persons = list(self.persons)
        
        # 分离锁定和非锁定人员
        locked_persons = [p for p in all_persons if p.active and p.locked]
        unlocked_persons = [p for p in all_persons if p.active and not p.locked]
        
        # 构建锁定人员的席位映射
        locked_seat_map = {}  # {person_name: seat_id}
        for seat_id, persons in pre_filled.items():
            for p in persons:
                if p.locked:
                    locked_seat_map[p.name] = seat_id
        
        # 获取总容量
        total_capacity = sum(s.persons_count for s in self.seats if s.available)
        
        # 锁定人员占用的总席位容量
        locked_seat_ids = set(locked_seat_map.values())
        locked_count = len(locked_persons)
        
        best_seats = None
        best_assigned = 0
        best_unplaced = []
        
        for attempt in range(2000):
            random.shuffle(unlocked_persons)
            
            # 构建初始席位：先放锁定人员
            seats = {s.id: [] for s in self.seats if s.available}
            assigned = set()
            
            # 放入锁定人员
            for p in locked_persons:
                seat_id = locked_seat_map.get(p.name)
                if seat_id and seat_id in seats:
                    seats[seat_id].append(p)
                    assigned.add(p.name)
            
            # 按均衡分数排序解锁人员（工时少的先安排）
            unlocked_persons.sort(key=lambda p: (
                p.total_hours,
                -sum(p.seat_history.get(s.id, 0) for s in self.seats)
            ))
            
            # 尝试安排剩余每个人
            for person in unlocked_persons:
                # 按均衡分数选择最优席位
                best_seat = self._select_best_seat_for_person(person, seats)
                if best_seat:
                    seats[best_seat.id].append(person)
                    assigned.add(person.name)
            
            # 验证约束
            valid = True
            for seat in self.seats:
                if not seat.available:
                    continue
                seat_people = seats.get(seat.id, [])
                # 检查等级兼容性
                for i, p1 in enumerate(seat_people):
                    for p2 in seat_people[i+1:]:
                        if not self.can_place_together(p1, p2):
                            valid = False
                            break
                    if not valid:
                        break
                if not valid:
                    break
                # 检查分数要求
                if not self.check_seat_score(seat_people, seat.required_score):
                    valid = False
                    break
            
            if valid:
                current_assigned = len(assigned)
                if current_assigned > best_assigned:
                    best_assigned = current_assigned
                    best_seats = seats
                    # 记录未安排的人员
                    best_unplaced = [p for p in all_persons if p.active and p.name not in assigned]
                
                if current_assigned >= total_capacity:
                    break
        
        if best_seats:
            placed_count = sum(len(v) for v in best_seats.values())
            if len(best_unplaced) == 0:
                return best_seats, "排班成功"
            else:
                unplaced_names = [p.name for p in best_unplaced[:3]]
                msg = f"已安排{placed_count}人/"
                if len(best_unplaced) > 3:
                    msg += f"还有{len(best_unplaced)}人待安排: {','.join(unplaced_names)}..."
                else:
                    msg += f"未安排: {','.join(unplaced_names)}"
                return best_seats, msg
        
        return None, "无法生成有效的排班方案"
    
    def _select_best_seat_for_person(self, person, seats):
        """为解锁人员选择最优席位（考虑均衡）"""
        best_seat = None
        best_score = float('inf')
        
        for seat in self.seats:
            if not seat.available:
                continue
            if len(seats.get(seat.id, [])) >= seat.persons_count:
                continue
            
            current_people = seats.get(seat.id, [])
            if not self.can_add_to_seat(current_people, person, seat.required_score):
                continue
            
            # 计算均衡分数
            balance_score = self.calc_balance_score(person, seat.id, seats)
            if balance_score < best_score:
                best_score = balance_score
                best_seat = seat
        
        return best_seat


class PersonEditDialog(QDialog):
    """人员编辑/添加对话框"""
    def __init__(self, parent=None, person_data=None):
        super().__init__(parent)
        self.person_data = person_data
        self.setup_ui()
        
    def setup_ui(self):
        self.setWindowTitle("编辑人员信息" if self.person_data else "添加人员")
        self.setMinimumWidth(350)
        self.setStyleSheet("""
            QLineEdit, QComboBox {
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 3px;
            }
        """)
        
        layout = QFormLayout(self)
        layout.setSpacing(12)
        
        self.name_edit = QLineEdit()
        if self.person_data:
            self.name_edit.setText(self.person_data['name'])
        layout.addRow("姓名 *", self.name_edit)
        
        self.level_combo = QComboBox()
        self.level_combo.addItems(["C1", "C2", "C3", "I", "S"])
        if self.person_data:
            self.level_combo.setCurrentText(self.person_data['level'])
        self.level_combo.currentTextChanged.connect(self.on_level_changed)
        layout.addRow("等级 *", self.level_combo)
        
        # 分数显示（根据等级自动设置）
        self.score_label = QLabel()
        self.score_label.setStyleSheet("font-weight: bold; color: #1565c0;")
        layout.addRow("分数", self.score_label)
        
        self.active_checkbox = QCheckBox("参与排班")
        self.active_checkbox.setChecked(True)
        if self.person_data:
            self.active_checkbox.setChecked(bool(self.person_data['active']))
        layout.addRow("", self.active_checkbox)
        
        # 初始化分数显示
        self.on_level_changed(self.level_combo.currentText())
        
        buttons = QDialogButtonBox(
            QDialogButtonBox.Ok | QDialogButtonBox.Cancel
        )
        buttons.accepted.connect(self.validate_input)
        buttons.rejected.connect(self.reject)
        layout.addRow("", buttons)
    
    def on_level_changed(self, level):
        """等级变化时自动更新分数"""
        level_scores = {"C1": 1, "C2": 2, "C3": 3, "I": 4, "S": 4}
        score = level_scores.get(level, 0)
        self.score_label.setText(f"{score} 分")
    
    def validate_input(self):
        if not self.name_edit.text().strip():
            CustomMessageBox.warning(self, "输入错误", "请输入姓名")
            return
        self.accept()
    
    def get_data(self):
        level_scores = {"C1": 1, "C2": 2, "C3": 3, "I": 4, "S": 4}
        return {
            'name': self.name_edit.text().strip(),
            'level': self.level_combo.currentText(),
            'score': level_scores.get(self.level_combo.currentText(), 0),
            'active': 1 if self.active_checkbox.isChecked() else 0
        }


class PersonDetailDialog(QDialog):
    """人员详情对话框"""
    def __init__(self, parent=None, person_data=None):
        super().__init__(parent)
        self.person_data = person_data
        self.setup_ui()
    
    def setup_ui(self):
        self.setWindowTitle("人员详细信息")
        self.setMinimumSize(350, 300)
        
        layout = QVBoxLayout(self)
        
        info_text = QTextEdit()
        info_text.setReadOnly(True)
        info_text.setStyleSheet("background-color: #f5f5f5;")
        
        p = self.person_data
        status = "✓ 参与排班" if p['active'] else "✗ 不参与排班"
        level_color = "#e74c3c" if p['level'] == "C1" else ("#f39c12" if p['level'] == "C2" else "#27ae60")
        
        html = f"""
        <h2 style="color: {level_color};">{p['name']}</h2>
        <p><b>等级：</b><span style="color: {level_color};">{p['level']}</span></p>
        <p><b>分数：</b><span style="color: #1565c0; font-weight: bold;">{p['score']}分</span></p>
        <p><b>状态：</b>{status}</p>
        <hr>
        <p><small>添加时间：{p.get('created_at', '未知')}</small></p>
        <p><small>更新时间：{p.get('updated_at', '未知')}</small></p>
        """
        info_text.setHtml(html)
        layout.addWidget(info_text)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        close_btn = QPushButton("关闭")
        close_btn.setStyleSheet("background-color: #3498db; color: white; padding: 6px 20px; border: none; border-radius: 4px;")
        close_btn.clicked.connect(self.accept)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)


class CustomMessageBox(QDialog):
    """自定义消息框 - 白底黑字"""
    def __init__(self, parent=None, title="", message="", icon_type="info"):
        super().__init__(parent)
        self.setWindowTitle(title)
        self.setMinimumWidth(350)
        self.icon_type = icon_type
        self.setup_ui(message)

    def setup_ui(self, message):
        self.setStyleSheet("""
            QDialog {
                background-color: #ffffff;
            }
        """)

        layout = QVBoxLayout(self)
        layout.setSpacing(15)

        # 图标和消息
        h_layout = QHBoxLayout()

        # 图标
        icon_label = QLabel()
        if self.icon_type == "warning":
            icon_label.setText("⚠️")
        elif self.icon_type == "question":
            icon_label.setText("❓")
        else:
            icon_label.setText("✅")
        icon_label.setStyleSheet("font-size: 32px;")
        icon_label.setFixedWidth(50)
        h_layout.addWidget(icon_label)

        # 消息文字
        msg_label = QLabel(message)
        msg_label.setStyleSheet("color: black; font-size: 14px;")
        msg_label.setWordWrap(True)
        h_layout.addWidget(msg_label)

        layout.addLayout(h_layout)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        if self.icon_type == "question":
            self.yes_btn = QPushButton("是")
            self.yes_btn.setStyleSheet("""
                QPushButton {
                    background-color: #27ae60;
                    color: white;
                    padding: 8px 25px;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                }
                QPushButton:hover { background-color: #1e8449; }
            """)
            self.yes_btn.clicked.connect(self.accept)
            self.no_btn = QPushButton("否")
            self.no_btn.setStyleSheet("""
                QPushButton {
                    background-color: #7f8c8d;
                    color: white;
                    padding: 8px 25px;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                }
                QPushButton:hover { background-color: #707b7c; }
            """)
            self.no_btn.clicked.connect(self.reject)
            btn_layout.addWidget(self.yes_btn)
            btn_layout.addWidget(self.no_btn)
        else:
            self.ok_btn = QPushButton("确定")
            self.ok_btn.setStyleSheet("""
                QPushButton {
                    background-color: #3498db;
                    color: white;
                    padding: 8px 25px;
                    border: none;
                    border-radius: 4px;
                    font-weight: bold;
                }
                QPushButton:hover { background-color: #2980b9; }
            """)
            self.ok_btn.clicked.connect(self.accept)
            btn_layout.addWidget(self.ok_btn)

        layout.addLayout(btn_layout)

    def exec_(self):
        return QDialog.exec_(self)

    @staticmethod
    def information(parent, title, message):
        dialog = CustomMessageBox(parent, title, message, "info")
        dialog.exec_()

    @staticmethod
    def warning(parent, title, message):
        dialog = CustomMessageBox(parent, title, message, "warning")
        dialog.exec_()

    @staticmethod
    def question(parent, title, message):
        dialog = CustomMessageBox(parent, title, message, "question")
        result = dialog.exec_()
        return result == QDialog.Accepted


class RuleEditDialog(QDialog):
    """规则编辑对话框"""
    def __init__(self, parent=None, rule_data=None):
        super().__init__(parent)
        self.rule_data = rule_data
        self.setup_ui()

    def setup_ui(self):
        self.setWindowTitle("编辑规则" if self.rule_data else "添加规则")
        self.setMinimumWidth(400)
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
            }
            QLabel {
                color: #333333;
                font-size: 13px;
            }
            QLineEdit, QTextEdit {
                padding: 8px;
                border: 2px solid #3498db;
                border-radius: 5px;
                background-color: white;
                color: #333333;
                font-size: 13px;
            }
        """)

        layout = QFormLayout(self)
        layout.setSpacing(15)

        self.name_edit = QLineEdit()
        if self.rule_data:
            self.name_edit.setText(self.rule_data['name'])
        layout.addRow("规则名称 *", self.name_edit)

        self.desc_edit = QLineEdit()
        if self.rule_data:
            self.desc_edit.setText(self.rule_data.get('description', ''))
        layout.addRow("描述", self.desc_edit)

        # 分数修改
        score_layout = QHBoxLayout()
        self.score_spin = QDoubleSpinBox()
        self.score_spin.setRange(-10, 10)
        self.score_spin.setSingleStep(0.05)
        self.score_spin.setDecimals(2)
        self.score_spin.setValue(self.rule_data['score_modifier'] if self.rule_data else 0)
        self.score_spin.setStyleSheet("""
            QDoubleSpinBox {
                padding: 8px;
                border: 2px solid #3498db;
                border-radius: 5px;
                background-color: white;
                color: #333333;
            }
            QDoubleSpinBox::up-button, QDoubleSpinBox::down-button {
                background-color: #3498db;
                color: white;
            }
        """)
        score_layout.addWidget(self.score_spin)
        score_label = QLabel("分（负数扣分，正数加分）")
        score_label.setStyleSheet("color: #333333;")
        score_layout.addWidget(score_label)
        score_layout.addStretch()
        layout.addRow("分数修改", score_layout)

        # 启用状态
        self.active_check = QCheckBox("启用此规则")
        self.active_check.setChecked(self.rule_data['active'] if self.rule_data else True)
        self.active_check.setStyleSheet("""
            QCheckBox {
                color: #333333;
                font-size: 13px;
            }
            QCheckBox::indicator {
                width: 20px;
                height: 20px;
                border: 2px solid #27ae60;
                border-radius: 3px;
                background-color: white;
            }
            QCheckBox::indicator:checked {
                background-color: #27ae60;
            }
        """)
        layout.addRow("", self.active_check)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        ok_btn = QPushButton("确定")
        ok_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                padding: 8px 25px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #1e8449; }
        """)
        ok_btn.clicked.connect(self.validate_input)
        btn_layout.addWidget(ok_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #7f8c8d;
                color: white;
                padding: 8px 25px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #707b7c; }
        """)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addRow("", btn_layout)

    def validate_input(self):
        if not self.name_edit.text().strip():
            CustomMessageBox.warning(self, "输入错误", "请输入规则名称")
            return
        self.accept()

    def get_data(self):
        return {
            'name': self.name_edit.text().strip(),
            'description': self.desc_edit.text().strip(),
            'score_modifier': self.score_spin.value(),
            'active': 1 if self.active_check.isChecked() else 0
        }


class PersonRuleSelectDialog(QDialog):
    """人员规则选择对话框"""
    def __init__(self, parent=None, person_id=None, db=None):
        super().__init__(parent)
        self.person_id = person_id
        self.db = db
        self.setup_ui()
        self.load_rules()

    def setup_ui(self):
        self.setWindowTitle("选择生效规则")
        self.setMinimumWidth(450)
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
            }
            QLabel {
                color: #333333;
                font-size: 14px;
            }
            QCheckBox {
                color: #333333;
                font-size: 13px;
                padding: 5px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }
        """)

        layout = QVBoxLayout(self)

        # 提示
        layout.addWidget(QLabel("勾选要生效的规则（分数将根据规则计算）"))

        # 规则列表
        self.rule_widget = QWidget()
        self.rule_widget.setStyleSheet("background-color: white; border-radius: 5px;")
        self.rule_layout = QVBoxLayout(self.rule_widget)
        self.rule_layout.setSpacing(5)
        self.rule_layout.setContentsMargins(5, 5, 5, 5)
        self.checkboxes = {}

        scroll = QScrollArea()
        scroll.setWidget(self.rule_widget)
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("""
            QScrollArea {
                background-color: white;
                border: 1px solid #dcdcdc;
                border-radius: 5px;
            }
        """)
        scroll.setMinimumHeight(200)
        layout.addWidget(scroll)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        ok_btn = QPushButton("确定")
        ok_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                padding: 8px 25px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #1e8449; }
        """)
        ok_btn.clicked.connect(self.save_and_close)
        btn_layout.addWidget(ok_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #7f8c8d;
                color: white;
                padding: 8px 25px;
                border: none;
                border-radius: 4px;
                font-weight: bold;
            }
            QPushButton:hover { background-color: #707b7c; }
        """)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)

    def load_rules(self):
        """加载规则列表（只显示启用的规则）"""
        all_rules = self.db.get_active_rules()  # 只获取启用的规则
        person_rules = self.db.get_person_rules(self.person_id)

        # 创建规则ID到启用状态的映射
        rule_status = {}
        for r in person_rules:
            r_dict = dict(r)
            rule_status[r_dict['id']] = r_dict.get('enabled', 0) == 1

        for rule in all_rules:
            rule_dict = dict(rule)
            desc = rule_dict.get('description', '') or ''
            cb = QCheckBox(f"{rule_dict['name']} ({rule_dict['score_modifier']:+.2f}分) - {desc}")
            cb.setChecked(rule_status.get(rule_dict['id'], False))
            cb.setStyleSheet("""
                QCheckBox {
                    color: #333333;
                    font-size: 13px;
                    padding: 8px;
                }
                QCheckBox::indicator {
                    width: 18px;
                    height: 18px;
                    border: 2px solid #27ae60;
                    border-radius: 3px;
                    background-color: white;
                }
                QCheckBox::indicator:checked {
                    background-color: #27ae60;
                }
            """)
            self.checkboxes[rule_dict['id']] = cb
            self.rule_layout.addWidget(cb)

    def save_and_close(self):
        """保存规则选择"""
        for rule_id, cb in self.checkboxes.items():
            self.db.set_person_rule(self.person_id, rule_id, 1 if cb.isChecked() else 0)
        self.accept()


class CenterAlignDelegate(QStyledItemDelegate):
    """姓名列居中对齐委托 - 固定4汉字宽度"""
    def initStyleOption(self, option, index):
        super().initStyleOption(option, index)
        option.displayAlignment = Qt.AlignCenter


class SeatWidget(QWidget):
    """单个席位设置控件"""
    available_changed = pyqtSignal(int, bool)
    count_changed = pyqtSignal(int, int)
    score_changed = pyqtSignal(int, int)
    app_name_changed = pyqtSignal(int, str)
    
    def __init__(self, seat_data):
        super().__init__()
        self.seat_id = seat_data['id']
        self.seat_data = seat_data
        self.setup_ui(seat_data)
    
    def setup_ui(self, seat_data):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(5, 5, 5, 5)
        
        # 席位名称
        name_label = QLabel(seat_data['name'])
        name_label.setStyleSheet("font-weight: bold; font-size: 14px; min-width: 60px;")
        layout.addWidget(name_label)
        
        # APP名称选择
        app_label = QLabel("  APP:")
        layout.addWidget(app_label)
        
        self.app_combo = QComboBox()
        self.app_combo.addItems(["APP01", "APP02", "APP03", "APP04", "APP05"])
        # 获取当前app_name
        current_app = seat_data.get('app_name') or seat_data['name'].replace('席位', 'APP0')
        self.app_combo.setCurrentText(current_app)
        self.app_combo.setFixedWidth(80)
        self.app_combo.currentTextChanged.connect(lambda v: self.app_name_changed.emit(self.seat_id, v))
        layout.addWidget(self.app_combo)
        
        # 可用状态开关
        available_label = QLabel("  可用:")
        layout.addWidget(available_label)
        
        self.available_switch = QCheckBox()
        self.available_switch.setText("启用" if seat_data['available'] else "禁用")
        self.available_switch.setChecked(bool(seat_data['available']))
        self.available_switch.setStyleSheet("""
            QCheckBox {
                font-weight: bold;
            }
        """)
        self.available_switch.stateChanged.connect(lambda state: self.on_available_changed(state))
        layout.addWidget(self.available_switch)
        
        # 人数设置
        count_label = QLabel("  人数:")
        layout.addWidget(count_label)
        
        self.count_spin = QSpinBox()
        self.count_spin.setRange(1, 3)  # 席位人数1-3人
        self.count_spin.setValue(seat_data['persons_count'])
        self.count_spin.setFixedWidth(60)
        self.count_spin.valueChanged.connect(lambda v: self.count_changed.emit(self.seat_id, v))
        layout.addWidget(self.count_spin)
        
        # 分数要求设置
        score_label = QLabel("  分数要求:")
        layout.addWidget(score_label)
        
        self.score_combo = QComboBox()
        self.score_combo.addItems(["4", "5", "6", "7"])
        self.score_combo.setCurrentText(str(seat_data['required_score']))
        self.score_combo.setEditable(True)
        self.score_combo.setInsertPolicy(QComboBox.NoInsert)
        self.score_combo.setFixedWidth(60)
        self.score_combo.currentTextChanged.connect(lambda v: self.score_changed.emit(self.seat_id, int(v)) if v else None)
        layout.addWidget(self.score_combo)
        
        # 状态显示
        self.status_label = QLabel("🟢 已启用" if seat_data['available'] else "🔴 已禁用")
        self.status_label.setStyleSheet("margin-left: 10px;")
        layout.addWidget(self.status_label)
        
        layout.addStretch()
        
        self.update_style(seat_data['available'])
    
    def on_available_changed(self, state):
        available = state == Qt.Checked
        self.available_switch.setText("启用" if available else "禁用")
        self.status_label.setText("🟢 已启用" if available else "🔴 已禁用")
        self.available_changed.emit(self.seat_id, available)
        self.update_style(available)
    
    def update_style(self, available):
        if available:
            self.setStyleSheet("""
                SeatWidget {
                    background-color: #e8f5e9;
                    border: 2px solid #27ae60;
                    border-radius: 8px;
                }
            """)
        else:
            self.setStyleSheet("""
                SeatWidget {
                    background-color: #ffebee;
                    border: 2px solid #e74c3c;
                    border-radius: 8px;
                }
            """)


class SeatResultWidget(QWidget):
    """排班结果显示控件 - 右侧为下拉框选择人员"""
    seat_changed = pyqtSignal()  # 人员选择变化信号
    
    def __init__(self, seat, parent=None):
        super().__init__(parent)
        self.seat = seat
        self.person_combos = []  # 存储下拉框
        self.setup_ui()
    
    def setup_ui(self):
        self.setMinimumHeight(55)
        self.setStyleSheet("""
            SeatResultWidget {
                background-color: #ffffff;
                border: 2px solid #e0e0e0;
                border-radius: 8px;
                margin: 4px 0;
            }
        """)
        
        layout = QHBoxLayout(self)
        layout.setContentsMargins(15, 10, 15, 10)
        layout.setSpacing(15)
        
        # 左侧：席位名称 + 人数
        left_layout = QVBoxLayout()
        left_layout.setSpacing(4)
        left_layout.setContentsMargins(0, 0, 0, 0)
        
        self.title_label = QLabel(f"【{self.seat.name}→{self.seat.app_name}】")
        self.title_label.setStyleSheet("""
            font-weight: bold; 
            font-size: 16px; 
            color: #1565c0;
            background-color: transparent;
        """)
        left_layout.addWidget(self.title_label)
        
        self.count_label = QLabel(f"(0/{self.seat.persons_count}人)")
        self.count_label.setStyleSheet("color: #424242; font-size: 14px; font-weight: bold;")
        left_layout.addWidget(self.count_label)
        
        layout.addLayout(left_layout)
        
        # 分隔线
        separator = QFrame()
        separator.setFrameShape(QFrame.VLine)
        separator.setFixedWidth(2)
        separator.setStyleSheet("color: #bdbdbd;")
        layout.addWidget(separator)
        
        # 右侧：人员下拉框列表 - 只有可用席位才显示
        self.persons_container = QWidget()
        self.persons_layout = QHBoxLayout(self.persons_container)
        self.persons_layout.setSpacing(6)
        self.persons_layout.setContentsMargins(0, 0, 0, 0)
        
        # 根据席位人数创建对应数量的下拉框
        for i in range(self.seat.persons_count):
            combo = QComboBox()
            combo.setFixedWidth(100)
            combo.setStyleSheet("""
                QComboBox {
                    padding: 5px 8px;
                    border: 2px solid #27ae60;
                    border-radius: 5px;
                    background-color: #2c3e50;
                    color: white;
                    font-size: 14px;
                    font-weight: bold;
                }
                QComboBox:hover {
                    background-color: #34495e;
                }
                QComboBox::drop-down {
                    border: none;
                    width: 20px;
                }
                QComboBox::down-arrow {
                    image: none;
                    border-left: 5px solid transparent;
                    border-right: 5px solid transparent;
                    border-top: 6px solid white;
                    margin-right: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: #2c3e50;
                    color: white;
                    selection-background-color: #3498db;
                    selection-color: white;
                }
            """)
            combo.currentTextChanged.connect(self.on_selection_changed)
            self.person_combos.append(combo)
            self.persons_layout.addWidget(combo)
        
        self.persons_layout.addStretch()
        
        # 根据可用状态显示/隐藏
        if self.seat.available:
            layout.addWidget(self.persons_container)
        else:
            self.persons_container.hide()
    
    def on_selection_changed(self):
        """选择变化时更新计数"""
        self.update_count()
        self.seat_changed.emit()
    
    def update_count(self):
        """更新已选人数显示"""
        selected = sum(1 for combo in self.person_combos if combo.currentText() and combo.currentText() != "未选择")
        self.count_label.setText(f"({selected}/{self.seat.persons_count}人)")
    
    def update_persons_list(self, persons_list, excluded_names=None):
        """更新下拉框中的人员列表
        excluded_names: 需要排除的人员名称列表（其他席位已选择的）
        """
        if excluded_names is None:
            excluded_names = []
        
        # 保存当前选择
        current_selections = [combo.currentData() for combo in self.person_combos]
        
        # 阻止所有信号，避免循环触发
        for combo in self.person_combos:
            combo.blockSignals(True)
        
        for combo in self.person_combos:
            combo.clear()
            combo.addItem("未选择", None)
            for p in persons_list:
                # 支持字典、sqlite3.Row和对象三种格式
                try:
                    name = p['name']  # 字典或sqlite3.Row
                except:
                    name = p.name     # 对象
                try:
                    level = p['level']
                except:
                    level = p.level
                
                # 排除已选择的人员（但保留当前下拉框已选的人员）
                if name in excluded_names and name not in current_selections:
                    continue
                    
                combo.addItem(f"{name}({level})", name)
        
        # 恢复选择（检查是否仍然有效）
        for i, selection in enumerate(current_selections):
            if selection:
                index = self.person_combos[i].findData(selection)
                if index >= 0:
                    self.person_combos[i].setCurrentIndex(index)
        
        # 解除信号阻止
        for combo in self.person_combos:
            combo.blockSignals(False)
        
        self.update_count()
    
    def on_selection_changed(self):
        """选择变化时更新计数并通知刷新"""
        self.update_count()
        self.seat_changed.emit()  # 发出信号通知刷新
    
    def refresh_all_dropdowns(self, persons_list, all_selected, current_combo_index):
        """刷新所有下拉框，排除已选择的人员
        persons_list: 所有可用人员列表
        all_selected: 所有已选择的人员名称
        current_combo_index: 当前正在操作的combo索引（不清除其选择）
        """
        for i, combo in enumerate(self.person_combos):
            current_selection = combo.currentData() if i == current_combo_index else None
            
            combo.blockSignals(True)
            combo.clear()
            combo.addItem("未选择", None)
            
            for p in persons_list:
                try:
                    name = p['name']
                except:
                    name = p.name
                try:
                    level = p['level']
                except:
                    level = p.level
                
                # 当前下拉框已选的人员保留显示，其他已选择的排除
                excluded = [n for n in all_selected if n != current_selection]
                if name in excluded:
                    continue
                    
                combo.addItem(f"{name}({level})", name)
            
            combo.blockSignals(False)
            
            # 恢复当前选择
            if current_selection:
                idx = combo.findData(current_selection)
                if idx >= 0:
                    combo.setCurrentIndex(idx)
    
    def get_selected_persons(self):
        """获取已选择的人员名称列表"""
        selected = []
        for combo in self.person_combos:
            name = combo.currentData()
            if name:
                selected.append(name)
        return selected
    
    def set_selected_persons(self, names):
        """设置已选择的人员"""
        for i, name in enumerate(names):
            if i < len(self.person_combos):
                index = self.person_combos[i].findData(name)
                if index >= 0:
                    self.person_combos[i].setCurrentIndex(index)
        self.update_count()
    
    def clear_selection(self):
        """清空选择"""
        for combo in self.person_combos:
            combo.blockSignals(True)
            combo.setCurrentIndex(0)
            combo.blockSignals(False)
        self.update_count()
    
    def set_available(self, available):
        """设置席位是否可用，控制下拉框显示/隐藏"""
        self.seat.available = available  # 更新席位对象
        if available:
            self.persons_container.show()
        else:
            self.clear_selection()
            self.persons_container.hide()
    
    def recreate_combos(self, new_count):
        """重建下拉框 - 当席位人数变化时调用"""
        # 保存当前选择
        current_selections = self.get_selected_persons()
        
        # 清除旧的下拉框
        while self.persons_layout.count():
            item = self.persons_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        self.person_combos = []
        self.seat.persons_count = new_count
        self.count_label.setText(f"(0/{new_count}人)")
        
        # 创建新数量的下拉框 - 使用与setup_ui一致的样式
        for i in range(new_count):
            combo = QComboBox()
            combo.setFixedWidth(100)
            combo.setStyleSheet("""
                QComboBox {
                    padding: 5px 8px;
                    border: 2px solid #27ae60;
                    border-radius: 5px;
                    background-color: #2c3e50;
                    color: white;
                    font-size: 14px;
                    font-weight: bold;
                }
                QComboBox:hover {
                    background-color: #34495e;
                }
                QComboBox::drop-down {
                    border: none;
                    width: 20px;
                }
                QComboBox::down-arrow {
                    image: none;
                    border-left: 5px solid transparent;
                    border-right: 5px solid transparent;
                    border-top: 6px solid white;
                    margin-right: 5px;
                }
                QComboBox QAbstractItemView {
                    background-color: #2c3e50;
                    color: white;
                    selection-background-color: #3498db;
                    selection-color: white;
                }
            """)
            combo.currentTextChanged.connect(self.on_selection_changed)
            self.person_combos.append(combo)
            self.persons_layout.addWidget(combo)
        
        self.persons_layout.addStretch()
        
        # 重新加载人员列表
        self.update_persons_list_from_db()
        
        # 尝试恢复之前的选择（最多恢复min(len(原选择), 新下拉框数量)个）
        for i, name in enumerate(current_selections[:new_count]):
            idx = self.person_combos[i].findData(name)
            if idx >= 0:
                self.person_combos[i].setCurrentIndex(idx)
    
    def update_persons_list_from_db(self):
        """从数据库刷新人员列表"""
        from scheduler import Person  # 避免循环导入
        # 重新获取人员列表（这里需要访问MainWindow的db）
        # 暂时先不更新，实际使用时会通过MainWindow调用update_persons_list


class MainWindow(QMainWindow):
    """主窗口"""
    def __init__(self):
        super().__init__()
        self.db = Database()
        self.seat_widgets = []
        self.seat_result_widgets = []
        self.current_schedule = None  # 保存当前排班结果
        self.setup_ui()
        self.load_data()
        self.init_schedule_combos()  # 初始化排班页面下拉框
    
    def init_schedule_combos(self):
        """初始化排班页面的人员下拉框"""
        active_persons = self.db.get_active_persons()
        
        # 先初始化所有下拉框
        for widget in self.seat_result_widgets:
            widget.update_persons_list(active_persons)
        
        # 连接信号 - 当选择变化时，刷新所有下拉框排除已选人员
        for widget in self.seat_result_widgets:
            widget.seat_changed.connect(lambda w=widget: self.on_seat_selection_changed(w))
    
    def on_seat_selection_changed(self, changed_widget):
        """当某个席位选择变化时，刷新所有下拉框排除已选择的人员"""
        active_persons = self.db.get_active_persons()
        
        # 收集所有已选择的人员（从所有席位）
        all_selected = []
        for widget in self.seat_result_widgets:
            if widget.seat.available:
                all_selected.extend(widget.get_selected_persons())
        
        # 移除重复
        all_selected = list(set(all_selected))
        
        # 刷新所有席位的下拉框
        for widget in self.seat_result_widgets:
            if widget.seat.available:
                widget.update_persons_list(active_persons, all_selected)
    
    def setup_ui(self):
        self.setWindowTitle("排班管理系统 - 西南空管局")
        self.setMinimumSize(950, 750)
        
        # 设置全局样式
        self.setStyleSheet("""
            QMainWindow {
                background-color: #f5f7fa;
            }
            QLabel {
                color: #2c3e50;
            }
        """)
        
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QVBoxLayout(central_widget)
        
        tabs = QTabWidget()
        tabs.addTab(self.create_rules_tab(), "⚙️ 规则设置")
        tabs.addTab(self.create_person_tab(), "👥 人员管理")
        tabs.addTab(self.create_schedule_tab(), "📅 排班设置")
        tabs.addTab(self.create_preview_tab(), "📊 排班预览")

        main_layout.addWidget(tabs)
        
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("就绪")

    def create_rules_tab(self):
        """规则设置页面"""
        widget = QWidget()
        layout = QVBoxLayout(widget)

        # 说明
        info_label = QLabel("规则设置：定义各种分数调整规则（如生病扣分、培训扣分等），然后在人员管理中为每个人员选择生效的规则")
        info_label.setStyleSheet("color: #7f8c8d; font-size: 13px; padding: 10px; background-color: #f5f5f5; border-radius: 5px;")
        layout.addWidget(info_label)

        # 工具栏
        toolbar = QHBoxLayout()

        add_btn = QPushButton("➕ 添加规则")
        add_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 8px;")
        add_btn.clicked.connect(self.add_rule)
        toolbar.addWidget(add_btn)

        edit_btn = QPushButton("✏️ 编辑规则")
        edit_btn.setStyleSheet("background-color: #3498db; color: white; padding: 8px;")
        edit_btn.clicked.connect(self.edit_rule)
        toolbar.addWidget(edit_btn)

        delete_btn = QPushButton("🗑️ 删除规则")
        delete_btn.setStyleSheet("background-color: #e74c3c; color: white; padding: 8px;")
        delete_btn.clicked.connect(self.delete_rule)
        toolbar.addWidget(delete_btn)

        toolbar.addStretch()
        layout.addLayout(toolbar)

        # 规则表格
        self.rules_table = QTableWidget()
        self.rules_table.setColumnCount(5)
        self.rules_table.setHorizontalHeaderLabels(["ID", "规则名称", "描述", "分数修改", "启用状态"])
        self.rules_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.rules_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.rules_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.rules_table.setAlternatingRowColors(True)
        self.rules_table.doubleClicked.connect(self.on_rules_double_clicked)
        self.rules_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #dcdcdc;
                background-color: white;
                color: #333333;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 5px;
                text-align: center;
                color: #333333;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                color: #333333;
                font-weight: bold;
                padding: 6px;
                border: 1px solid #e0e0e0;
            }
            QTableWidget::item:selected {
                background-color: #d5f5e3;
                color: #1e8449;
            }
        """)
        layout.addWidget(self.rules_table)

        # 统计
        self.rules_stats_label = QLabel()
        self.rules_stats_label.setStyleSheet("font-weight: bold; font-size: 14px; padding: 10px; color: #333333;")
        layout.addWidget(self.rules_stats_label)

        self.load_rules()
        return widget

    def load_rules(self):
        """加载规则列表"""
        rules = self.db.get_all_rules()
        self.rules_table.setRowCount(len(rules))

        for i, rule in enumerate(rules):
            rule = dict(rule)  # sqlite3.Row转dict
            self.rules_table.setItem(i, 0, QTableWidgetItem(str(rule['id'])))
            self.rules_table.item(i, 0).setTextAlignment(Qt.AlignCenter)

            name_item = QTableWidgetItem(rule['name'])
            name_item.setTextAlignment(Qt.AlignCenter)
            self.rules_table.setItem(i, 1, name_item)

            desc_item = QTableWidgetItem(rule.get('description', ''))
            desc_item.setTextAlignment(Qt.AlignCenter)
            self.rules_table.setItem(i, 2, desc_item)

            score_item = QTableWidgetItem(f"{rule['score_modifier']:+.2f}")
            score_item.setTextAlignment(Qt.AlignCenter)
            if rule['score_modifier'] < 0:
                score_item.setForeground(QColor(231, 76, 60))
            else:
                score_item.setForeground(QColor(39, 174, 96))
            self.rules_table.setItem(i, 3, score_item)

            active_text = "✓ 启用" if rule['active'] else "✗ 禁用"
            active_item = QTableWidgetItem(active_text)
            active_item.setTextAlignment(Qt.AlignCenter)
            if rule['active']:
                active_item.setForeground(QColor(39, 174, 96))
            else:
                active_item.setForeground(QColor(149, 165, 166))
            self.rules_table.setItem(i, 4, active_item)

        self.rules_table.setColumnWidth(0, 50)
        self.rules_table.setColumnWidth(1, 120)
        self.rules_table.setColumnWidth(2, 200)
        self.rules_table.setColumnWidth(3, 100)
        self.rules_table.setColumnWidth(4, 100)

        # 统计
        active_count = sum(1 for r in rules if r['active'])
        self.rules_stats_label.setText(f"共 {len(rules)} 条规则，启用 {active_count} 条")

    def on_rules_double_clicked(self, index):
        """双击规则表格 - 打开编辑对话框"""
        row = index.row()
        rule_id = int(self.rules_table.item(row, 0).text())
        rule = self.db.get_rule_by_id(rule_id)
        dialog = RuleEditDialog(self, dict(rule))
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            success, msg = self.db.update_rule(
                rule_id, data['name'], data['description'],
                data['score_modifier'], 1 if data['active'] else 0
            )
            if success:
                self.load_rules()
            else:
                CustomMessageBox.warning(self, "失败", msg)

    def add_rule(self):
        """添加规则"""
        dialog = RuleEditDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            success, msg = self.db.add_rule(data['name'], data['description'], data['score_modifier'])
            if success:
                CustomMessageBox.information(self, "成功", msg)
                self.load_rules()
            else:
                CustomMessageBox.warning(self, "失败", msg)

    def edit_rule(self):
        """编辑规则"""
        row = self.rules_table.currentRow()
        if row < 0:
            CustomMessageBox.warning(self, "请选择", "请先选择要编辑的规则")
            return

        rule_id = int(self.rules_table.item(row, 0).text())
        rule = self.db.get_rule_by_id(rule_id)
        dialog = RuleEditDialog(self, dict(rule))
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            success, msg = self.db.update_rule(rule_id, data['name'], data['description'], data['score_modifier'], data['active'])
            if success:
                CustomMessageBox.information(self, "成功", msg)
                self.load_rules()
            else:
                CustomMessageBox.warning(self, "失败", msg)

    def delete_rule(self):
        """删除规则"""
        row = self.rules_table.currentRow()
        if row < 0:
            CustomMessageBox.warning(self, "请选择", "请先选择要删除的规则")
            return

        rule_id = int(self.rules_table.item(row, 0).text())
        rule_name = self.rules_table.item(row, 1).text()

        if CustomMessageBox.question(self, "确认删除", f"确定要删除规则「{rule_name}」吗？"):
            success, msg = self.db.delete_rule(rule_id)
            if success:
                CustomMessageBox.information(self, "成功", msg)
                self.load_rules()
            else:
                CustomMessageBox.warning(self, "失败", msg)

    def create_person_tab(self):
        """人员管理页面"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # 搜索栏
        search_layout = QHBoxLayout()
        search_layout.addWidget(QLabel("搜索:"))
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("输入姓名搜索...")
        self.search_edit.textChanged.connect(self.on_search)
        search_layout.addWidget(self.search_edit)
        
        search_layout.addWidget(QLabel("  等级:"))
        self.filter_combo = QComboBox()
        self.filter_combo.addItems(["全部", "C1", "C2", "C3", "I", "S"])
        self.filter_combo.currentTextChanged.connect(self.load_data)
        search_layout.addWidget(self.filter_combo)
        
        layout.addLayout(search_layout)
        
        # 工具栏
        toolbar = QHBoxLayout()
        
        self.add_person_btn = QPushButton("➕ 添加人员")
        self.add_person_btn.setStyleSheet("background-color: #27ae60; color: white; padding: 8px;")
        self.add_person_btn.clicked.connect(self.add_person)
        toolbar.addWidget(self.add_person_btn)
        
        self.edit_person_btn = QPushButton("✏️ 编辑信息")
        self.edit_person_btn.setStyleSheet("background-color: #3498db; color: white; padding: 8px;")
        self.edit_person_btn.clicked.connect(self.edit_person)
        toolbar.addWidget(self.edit_person_btn)
        
        self.detail_person_btn = QPushButton("👁️ 查看详情")
        self.detail_person_btn.setStyleSheet("background-color: #9b59b6; color: white; padding: 8px;")
        self.detail_person_btn.clicked.connect(self.view_person_detail)
        toolbar.addWidget(self.detail_person_btn)
        
        self.delete_person_btn = QPushButton("🗑️ 删除人员")
        self.delete_person_btn.setStyleSheet("background-color: #e74c3c; color: white; padding: 8px;")
        self.delete_person_btn.clicked.connect(self.delete_person)
        toolbar.addWidget(self.delete_person_btn)
        
        toolbar.addStretch()
        
        self.toggle_active_btn = QPushButton("🔄 切换状态")
        self.toggle_active_btn.clicked.connect(self.toggle_active)
        toolbar.addWidget(self.toggle_active_btn)

        layout.addLayout(toolbar)
        
        # 表格 - 5列（姓名、等级、分数、参与排班、生效规则）
        self.person_table = QTableWidget()
        self.person_table.setColumnCount(5)
        self.person_table.setHorizontalHeaderLabels([
            "姓名", "等级", "分数", "参与排班", "生效规则"
        ])
        self.person_table.horizontalHeader().setStretchLastSection(True)
        self.person_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.person_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.person_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.person_table.setAlternatingRowColors(True)
        self.person_table.doubleClicked.connect(self.on_person_double_clicked)
        self.person_table.cellClicked.connect(self.on_person_cell_clicked)

        # 姓名列使用自定义委托（固定4汉字宽度）
        self.person_table.setItemDelegateForColumn(0, CenterAlignDelegate(self.person_table))
        self.person_table.setColumnWidth(0, 120)  # 固定宽度

        # 表格样式优化
        self.person_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #dcdcdc;
                gridline-color: #e0e0e0;
                background-color: white;
                color: #333333;
                font-size: 13px;
            }
            QTableWidget::item {
                padding: 5px;
                text-align: center;
                color: #333333;
            }
            QTableWidget::item:selected {
                background-color: #d5f5e3;
                color: #1e8449;
            }
            QHeaderView::section {
                background-color: #f5f5f5;
                color: #333333;
                font-weight: bold;
                font-size: 13px;
                padding: 6px;
                border: none;
                border-right: 1px solid #e0e0e0;
                border-bottom: 1px solid #e0e0e0;
                text-align: center;
            }
            QTableWidget::item:alternate {
                background-color: #fafafa;
            }
        """)

        layout.addWidget(self.person_table)
        
        # 统计
        stats_layout = QHBoxLayout()
        stats_layout.addSpacing(10)
        
        self.stats_label = QLabel("总计: 0人")
        self.stats_label.setStyleSheet("font-weight: bold; font-size: 14px;")
        stats_layout.addWidget(self.stats_label)
        
        stats_layout.addStretch()
        
        self.c1_count_label = QLabel("C1: 0人")
        self.c1_count_label.setStyleSheet("color: #e74c3c; font-weight: bold; font-size: 14px;")
        stats_layout.addWidget(self.c1_count_label)
        
        self.c2_count_label = QLabel("C2: 0人")
        self.c2_count_label.setStyleSheet("color: #f39c12; font-weight: bold; font-size: 14px;")
        stats_layout.addWidget(self.c2_count_label)
        
        self.c3_count_label = QLabel("C3: 0人")
        self.c3_count_label.setStyleSheet("color: #27ae60; font-weight: bold; font-size: 14px;")
        stats_layout.addWidget(self.c3_count_label)
        
        # I级统计
        self.i_count_label = QLabel("I: 0人")
        self.i_count_label.setStyleSheet("color: #3498db; font-weight: bold; font-size: 14px;")
        stats_layout.addWidget(self.i_count_label)
        
        # S级统计
        self.s_count_label = QLabel("S: 0人")
        self.s_count_label.setStyleSheet("color: #9b59b6; font-weight: bold; font-size: 14px;")
        stats_layout.addWidget(self.s_count_label)
        
        layout.addLayout(stats_layout)
        
        self.load_data()
        
        return widget
    
    def create_schedule_tab(self):
        """排班设置页面"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        
        # 上半部分：席位设置
        seats_group = QGroupBox("席位设置")
        seats_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                border: 2px solid #3498db;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                background-color: #f8fbff;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 5px;
                color: #2980b9;
            }
        """)
        seats_layout = QVBoxLayout(seats_group)
        
        # 加载席位设置控件
        seats_data = self.db.get_all_seats()
        for seat_data in seats_data:
            seat_widget = SeatWidget(seat_data)
            seat_widget.available_changed.connect(self.on_seat_available_changed)
            seat_widget.count_changed.connect(self.on_seat_count_changed)
            seat_widget.score_changed.connect(self.on_seat_score_changed)
            seat_widget.app_name_changed.connect(self.on_seat_app_name_changed)
            seats_layout.addWidget(seat_widget)
            self.seat_widgets.append(seat_widget)
        
        layout.addWidget(seats_group)
        
        # 席位统计
        self.seat_stats_label = QLabel("可用席位: 5个, 总容量: 15人")
        self.seat_stats_label.setStyleSheet("""
            font-weight: bold;
            font-size: 14px;
            padding: 12px;
            background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #e8f4fd, stop:1 #d4e9f7);
            border-radius: 8px;
            color: #2c5aa0;
        """)
        layout.addWidget(self.seat_stats_label)
        
        # 排班结果区域
        result_group = QGroupBox("排班结果")
        result_group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 14px;
                border: 2px solid #27ae60;
                border-radius: 8px;
                margin-top: 10px;
                padding-top: 10px;
                background-color: #f8fff8;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 15px;
                padding: 0 5px;
                color: #1e8449;
            }
        """)
        result_layout = QVBoxLayout(result_group)
        
        # 5个席位的结果显示
        all_seats = self.db.get_all_seats()
        for seat_data in all_seats:
            seat = Seat(seat_data['id'], seat_data['name'], seat_data.get('app_name'),
                       bool(seat_data['available']), seat_data['persons_count'], seat_data['required_score'])
            result_widget = SeatResultWidget(seat)
            result_layout.addWidget(result_widget)
            self.seat_result_widgets.append(result_widget)
        
        layout.addWidget(result_group)
        
        # 按钮区域
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(15)
        
        # 排班按钮
        schedule_btn = QPushButton("🚀 开始排班")
        schedule_btn.setMinimumHeight(50)
        schedule_btn.setMinimumWidth(180)
        schedule_btn.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #2ecc71, stop:1 #27ae60);
                color: white;
                font-size: 15px;
                font-weight: bold;
                border: none;
                border-radius: 8px;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #27ae60, stop:1 #1e8449);
            }
            QPushButton:pressed {
                background-color: #1e8449;
            }
        """)
        schedule_btn.clicked.connect(self.run_schedule)
        btn_layout.addWidget(schedule_btn)
        btn_layout.addStretch()
        
        # 清除结果按钮
        clear_btn = QPushButton("🗑️ 清除结果")
        clear_btn.setMinimumHeight(50)
        clear_btn.setMinimumWidth(140)
        clear_btn.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #95a5a6, stop:1 #7f8c8d);
                color: white;
                font-size: 14px;
                font-weight: bold;
                border: none;
                border-radius: 8px;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #7f8c8d, stop:1 #707b7c);
            }
            QPushButton:pressed {
                background-color: #707b7c;
            }
        """)
        clear_btn.clicked.connect(self.clear_schedule_result)
        btn_layout.addWidget(clear_btn)

        # 导出Excel按钮
        export_btn = QPushButton("📊 导出Excel")
        export_btn.setMinimumHeight(50)
        export_btn.setMinimumWidth(140)
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #3498db, stop:1 #2980b9);
                color: white;
                font-size: 14px;
                font-weight: bold;
                border: none;
                border-radius: 8px;
                padding: 8px 16px;
            }
            QPushButton:hover {
                background-color: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop:0 #2980b9, stop:1 #1f6dad);
            }
            QPushButton:pressed {
                background-color: #1f6dad;
            }
        """)
        export_btn.clicked.connect(self.export_to_excel)
        btn_layout.addWidget(export_btn)

        layout.addLayout(btn_layout)
        
        return widget

    def create_preview_tab(self):
        """排班预览页面 - 席位网格布局，整体滚动"""
        widget = QWidget()
        main_layout = QVBoxLayout(widget)
        main_layout.setContentsMargins(5, 5, 5, 5)

        # 说明
        info_label = QLabel("排班预览：按样表格式显示席位、人员、执勤时段安排")
        info_label.setStyleSheet("color: #7f8c8d; font-size: 13px; padding: 10px; background-color: #f5f5f5; border-radius: 5px;")
        main_layout.addWidget(info_label)

        # 滚动区域容纳所有席位块
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("border: none;")
        main_layout.addWidget(scroll)

        # 容器widget - 使用VBoxLayout实现2列网格
        self.preview_container = QWidget()
        self.preview_layout = QVBoxLayout(self.preview_container)
        self.preview_layout.setSpacing(10)
        self.preview_layout.setContentsMargins(0, 0, 0, 0)
        scroll.setWidget(self.preview_container)

        # 刷新按钮
        refresh_btn = QPushButton("🔄 刷新预览")
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                font-size: 13px;
                padding: 8px 20px;
                border: none;
                border-radius: 5px;
            }
            QPushButton:hover {
                background-color: #2980b9;
            }
        """)
        refresh_btn.clicked.connect(self.refresh_preview)
        main_layout.addWidget(refresh_btn)

        # 初始刷新
        self.refresh_preview()
        return widget

    def refresh_preview(self):
        """刷新排班预览 - 参考排班1.xlsx格式，单列完整显示"""
        # 清除旧的席位块
        while self.preview_layout.count():
            item = self.preview_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # 收集席位人员
        seat_map = {}  # seat_name -> [persons]
        seat_app_map = {}  # seat_name -> app_name
        for widget in self.seat_result_widgets:
            if widget.seat.available:
                seat_map[widget.seat.name] = widget.get_selected_persons()
                seat_app_map[widget.seat.name] = widget.seat.app_name

        # 每个席位的时间定义 - 完全按照排班.xlsx样表
        seat_times = {
            'APP01': {
                '早班': [
                    ('0840-1030', 'A', '0850-0950', 'B'),
                    ('1020-1150', 'B', '0940-1110', 'C'),
                    ('1140-1230', 'C', '1100-1240', 'A'),
                ],
                '晚班': [
                    ('1750-1930', 'C', '1800-1850', 'A'),
                    ('1920-2050', 'A', '1840-2010', 'B'),
                    ('2040-2210', 'B', '2000-2130', 'C'),
                    (None, None, '2120-2220', 'A'),
                ],
                '下午班': [
                    ('1220-1350', 'B', '1230-1310', 'C'),
                    ('1340-1540', 'C', '1300-1450', 'A'),
                    ('1530-1720', 'A', '1440-1640', 'B'),
                    ('1710-1800', 'B', '1630-1810', 'C'),
                ],
            },
            'APP02': {
                '早班': [
                    ('0850-1040', 'A', '0840-1000', 'B'),
                    ('1030-1200', 'B', '0950-1120', 'C'),
                    ('1150-1240', 'C', '1110-1230', 'A'),
                ],
                '晚班': [
                    ('1800-1940', 'C', '1750-1900', 'A'),
                    ('1930-2100', 'A', '1850-2020', 'B'),
                    ('2050-2220', 'B', '2010-2140', 'C'),
                    (None, None, '2130-2210', 'A'),
                ],
                '下午班': [
                    ('1230-1400', 'B', '1220-1320', 'C'),
                    ('1350-1550', 'C', '1310-1500', 'A'),
                    ('1540-1730', 'A', '1450-1650', 'B'),
                    ('1720-1810', 'B', '1640-1800', 'C'),
                ],
            },
            'APP03': {
                '早班': [
                    ('0850-1040', 'A', '0840-1000', 'B'),
                    ('1030-1200', 'B', '0950-1120', 'C'),
                    ('1150-1240', 'C', '1110-1230', 'A'),
                ],
                '晚班': [
                    ('1800-1940', 'C', '1750-1900', 'A'),
                    ('1930-2100', 'A', '1850-2020', 'B'),
                    ('2050-2220', 'B', '2010-2140', 'C'),
                    (None, None, '2130-2210', 'A'),
                ],
                '下午班': [
                    ('1230-1400', 'B', '1220-1320', 'C'),
                    ('1350-1550', 'C', '1310-1500', 'A'),
                    ('1540-1730', 'A', '1450-1650', 'B'),
                    ('1720-1810', 'B', '1640-1800', 'C'),
                ],
            },
            'APP04': {
                '早班': [
                    ('0850-1040', 'A', '0840-1000', 'B'),
                    ('1030-1200', 'B', '0950-1120', 'C'),
                    ('1150-1240', 'C', '1110-1230', 'A'),
                ],
                '晚班': [
                    ('1800-1940', 'C', '1750-1900', 'A'),
                    ('1930-2100', 'A', '1850-2020', 'B'),
                    ('2050-2220', 'B', '2010-2140', 'C'),
                    (None, None, '2130-2210', 'A'),
                ],
                '下午班': [
                    ('1230-1400', 'B', '1220-1320', 'C'),
                    ('1350-1550', 'C', '1310-1500', 'A'),
                    ('1540-1730', 'A', '1450-1650', 'B'),
                    ('1720-1810', 'B', '1640-1800', 'C'),
                ],
            },
            'APP05': {
                '早班': [
                    ('0840-1030', 'A', '0850-0950', 'B'),
                    ('1020-1150', 'B', '0940-1110', 'C'),
                    ('1140-1230', 'C', '1100-1240', 'A'),
                ],
                '晚班': [
                    ('1750-1930', 'C', '1800-1850', 'A'),
                    ('1920-2050', 'A', '1840-2010', 'B'),
                    ('2040-2210', 'B', '2000-2130', 'C'),
                    (None, None, '2120-2220', 'A'),
                ],
                '下午班': [
                    ('1220-1350', 'B', '1230-1310', 'C'),
                    ('1340-1540', 'C', '1300-1450', 'A'),
                    ('1530-1720', 'A', '1440-1640', 'B'),
                    ('1710-1800', 'B', '1630-1810', 'C'),
                ],
            },
        }

        def get_person(persons, pos):
            if not persons:
                return ''
            if pos == 'A':
                return persons[0] if len(persons) > 0 else ''
            elif pos == 'B':
                return persons[1] if len(persons) > 1 else ''
            elif pos == 'C':
                return persons[2] if len(persons) > 2 else ''
            return ''

        # 按APP名称排序
        available_seats = sorted(
            [(name, seat_app_map.get(name, name)) for name in seat_map.keys()],
            key=lambda x: x[1]
        )

        shifts_order = ['早班', '晚班', '下午班']

        def create_seat_block(seat_name, app_name, persons):
            """创建单个席位区块 - 参考排班1.xlsx格式"""
            # 席位区块Frame - 设置固定高度确保完全展开
            seat_frame = QFrame()
            seat_frame.setStyleSheet("""
                QFrame {
                    background-color: white;
                    border: 1px solid #3498db;
                    border-radius: 4px;
                }
            """)
            seat_layout = QVBoxLayout(seat_frame)
            seat_layout.setContentsMargins(0, 0, 0, 0)
            seat_layout.setSpacing(0)

            # 创建席位表格 - 与排班1.xlsx一致: 5列
            seat_table = QTableWidget()
            seat_table.setColumnCount(5)  # A(空/班次), B(执勤时段), C(管制席), D(执勤时段), E(助理席)
            seat_table.setHorizontalHeaderLabels(['班次', '执勤时段', '管制席', '执勤时段', '助理席'])
            seat_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
            seat_table.setEditTriggers(QTableWidget.NoEditTriggers)
            seat_table.verticalHeader().setVisible(False)
            seat_table.setShowGrid(True)
            seat_table.setStyleSheet("""
                QTableWidget {
                    border: none;
                    background-color: white;
                    color: #333333;
                    font-size: 12px;
                    gridline-color: #cccccc;
                }
                QTableWidget::item {
                    padding: 6px 4px;
                    text-align: center;
                    color: #333333;
                    border: 1px solid #dddddd;
                }
                QHeaderView::section {
                    background-color: #e8e8e8;
                    color: #333333;
                    font-weight: bold;
                    padding: 6px 4px;
                    border: 1px solid #cccccc;
                    font-size: 12px;
                }
            """)

            # 计算行数: 席位名称(1) + 早班(1标题+3数据), 晚班(1标题+4数据), 下午班(1标题+4数据) = 16行
            row_count = 16
            seat_table.setRowCount(row_count)

            # 设置固定行高
            row_height = 26
            for r in range(row_count):
                seat_table.setRowHeight(r, row_height)

            # 直接设置表格总高度 (表头30px + 数据行)
            header_height = 30
            total_table_height = header_height + row_count * row_height
            seat_table.setFixedHeight(total_table_height)

            # 设置列宽
            seat_table.horizontalHeader().setSectionResizeMode(QHeaderView.Fixed)
            seat_table.setColumnWidth(0, 60)   # 班次
            seat_table.setColumnWidth(1, 100)  # 执勤时段
            seat_table.setColumnWidth(2, 80)   # 管制席
            seat_table.setColumnWidth(3, 100)  # 执勤时段
            seat_table.setColumnWidth(4, 80)   # 助理席

            row = 0

            # 席位名称标题行 - 合并所有列，使用APP名称
            seat_name_item = QTableWidgetItem(app_name)
            seat_name_item.setFont(QFont("Arial", 12, QFont.Bold))
            seat_name_item.setBackground(QColor(200, 220, 240))
            seat_name_item.setTextAlignment(Qt.AlignCenter)
            seat_table.setItem(row, 0, seat_name_item)
            seat_table.setSpan(row, 0, 1, 5)
            row += 1

            for shift_name in shifts_order:
                # 班次标题行 - A列显示班次名，B-E合并
                shift_item = QTableWidgetItem(shift_name)
                shift_item.setFont(QFont("Arial", 11, QFont.Bold))
                shift_item.setBackground(QColor(220, 235, 250))
                shift_item.setTextAlignment(Qt.AlignCenter)
                seat_table.setItem(row, 0, shift_item)
                seat_table.setSpan(row, 0, 1, 5)
                row += 1

                # 时段数据行
                times = seat_times.get(app_name, {}).get(shift_name, [])
                for slot_idx in range(4):  # 固定4个时段
                    if slot_idx < len(times):
                        t1, pos1, t2, pos2 = times[slot_idx]
                        seat_table.setItem(row, 0, QTableWidgetItem(''))
                        seat_table.setItem(row, 1, QTableWidgetItem(t1 if t1 else ''))
                        seat_table.setItem(row, 2, QTableWidgetItem(get_person(persons, pos1)))
                        seat_table.setItem(row, 3, QTableWidgetItem(t2 if t2 else ''))
                        seat_table.setItem(row, 4, QTableWidgetItem(get_person(persons, pos2)))
                    else:
                        for i in range(5):
                            seat_table.setItem(row, i, QTableWidgetItem(''))
                    row += 1

            seat_layout.addWidget(seat_table)

            # 计算实际需要的框架高度 = 表格高度 + 额外空间确保完全显示
            seat_frame.setFixedHeight(total_table_height + 20)

            return seat_frame

        # 2列网格布局，每行2个席位
        cols = 2
        for idx, (seat_name, app_name) in enumerate(available_seats):
            row_idx = idx // cols
            col_idx = idx % cols
            persons = seat_map.get(seat_name, [])
            seat_frame = create_seat_block(seat_name, app_name, persons)

            # 创建行框架
            if col_idx == 0:
                row_frame = QFrame()
                row_frame.setStyleSheet("QFrame { background-color: transparent; }")
                row_layout = QHBoxLayout(row_frame)
                row_layout.setContentsMargins(0, 0, 0, 0)
                row_layout.setSpacing(10)
                self.preview_layout.addWidget(row_frame)
            else:
                row_frame = self.preview_layout.itemAt(self.preview_layout.count() - 1).widget()

            row_frame.layout().addWidget(seat_frame)

    def clear_schedule_result(self):
        """清除排班结果 - 清空所有选择"""
        self.current_schedule = None
        for widget in self.seat_result_widgets:
            widget.clear_selection()
        self.status_bar.showMessage("已清除所有选择")
    
    def update_seat_result_widgets(self):
        """更新席位结果控件的席位信息"""
        seats = self.db.get_all_seats()
        
        # 先更新席位信息
        for i, widget in enumerate(self.seat_result_widgets):
            if i < len(seats):
                new_seat = Seat(seats[i]['id'], seats[i]['name'], seats[i].get('app_name'),
                                  bool(seats[i]['available']), seats[i]['persons_count'], seats[i]['required_score'])
                
                # 如果席位人数变化了，重建下拉框
                if new_seat.persons_count != widget.seat.persons_count:
                    widget.seat = new_seat
                    widget.recreate_combos(new_seat.persons_count)
                    active_persons = self.db.get_active_persons()
                    widget.update_persons_list(active_persons)
                else:
                    widget.seat = new_seat
                
                # 更新可用状态
                widget.set_available(new_seat.available)
                
                widget.title_label.setText(f"【{widget.seat.name}→{widget.seat.app_name}】")
                widget.count_label.setText(f"(0/{seats[i]['persons_count']}人)")
    
    def on_seat_available_changed(self, seat_id, available):
        """席位可用状态改变"""
        self.db.update_seat_available(seat_id, 1 if available else 0)
        
        # 清除该席位的结果并显示/隐藏下拉框，同时更新widget.seat
        for widget in self.seat_result_widgets:
            if widget.seat.id == seat_id:
                widget.seat.available = available  # 直接更新widget.seat
                widget.clear_selection()
                widget.set_available(available)
                # 启用时重新加载人员列表
                if available:
                    active_persons = self.db.get_active_persons()
                    widget.update_persons_list(active_persons)
                break
        
        self.update_seat_stats()
        self.update_stats()  # 更新人员统计
        self.refresh_preview()  # 刷新预览
        self.status_bar.showMessage(f"席位 {seat_id} 已{'启用' if available else '禁用'}")
    
    def on_seat_count_changed(self, seat_id, count):
        """席位人数改变"""
        self.db.update_seat_persons_count(seat_id, count)
        
        # 重建该席位的下拉框
        for widget in self.seat_result_widgets:
            if widget.seat.id == seat_id:
                widget.recreate_combos(count)
                # 更新人员列表
                active_persons = self.db.get_active_persons()
                widget.update_persons_list(active_persons)
        
        self.update_seat_stats()
        self.status_bar.showMessage(f"席位 {seat_id} 人数已设为 {count}")
    
    def on_seat_score_changed(self, seat_id, required_score):
        """席位分数要求改变"""
        self.db.update_seat_required_score(seat_id, required_score)

        # 更新SeatResultWidget中的seat对象分数
        for widget in self.seat_result_widgets:
            if widget.seat.id == seat_id:
                widget.seat.required_score = required_score
                break

        self.update_seat_stats()
        self.status_bar.showMessage(f"席位 {seat_id} 分数要求已设为 {required_score}")
    
    def on_seat_app_name_changed(self, seat_id, app_name):
        """席位APP名称改变"""
        self.db.update_seat_app_name(seat_id, app_name)

        # 更新SeatResultWidget中的seat对象APP名称
        for widget in self.seat_result_widgets:
            if widget.seat.id == seat_id:
                widget.seat.app_name = app_name
                # 更新标题
                widget.title_label.setText(f"【{widget.seat.name}→{app_name}】")
                break

        # 刷新预览
        self.refresh_preview()
        self.status_bar.showMessage(f"席位 {seat_id} APP名称已设为 {app_name}")
    
    def update_seat_stats(self):
        """更新席位统计"""
        seats = self.db.get_all_seats()
        available_count = sum(1 for s in seats if s['available'])
        total_capacity = sum(s['persons_count'] for s in seats if s['available'])
        new_text = f"可用席位: {available_count}个, 总容量: {total_capacity}人"
        print(f"[DEBUG] update_seat_stats: {new_text}")
        self.seat_stats_label.setText(new_text)
        self.seat_stats_label.repaint()  # 强制重绘
    
    def load_data(self):
        """加载人员数据"""
        filter_level = self.filter_combo.currentText()
        
        if filter_level == "全部":
            rows = self.db.get_all_persons()
        else:
            all_rows = self.db.get_all_persons()
            rows = [r for r in all_rows if r['level'] == filter_level]
        
        self.person_table.setRowCount(len(rows))
        
        for i, row in enumerate(rows):
            row = dict(row)  # sqlite3.Row转dict
            name_item = QTableWidgetItem(row['name'])
            name_item.setTextAlignment(Qt.AlignCenter)
            self.person_table.setItem(i, 0, name_item)

            level_item = QTableWidgetItem(row['level'])
            level_item.setTextAlignment(Qt.AlignCenter)
            if row['level'] == "C1":
                level_item.setBackground(QColor(231, 76, 60))
                level_item.setForeground(QColor(255, 255, 255))
            elif row['level'] == "C2":
                level_item.setBackground(QColor(243, 156, 18))
            elif row['level'] == "C3":
                level_item.setBackground(QColor(39, 174, 96))
                level_item.setForeground(QColor(255, 255, 255))
            elif row['level'] == "I":
                level_item.setBackground(QColor(52, 152, 219))
                level_item.setForeground(QColor(255, 255, 255))
            elif row['level'] == "S":
                level_item.setBackground(QColor(155, 89, 182))
                level_item.setForeground(QColor(255, 255, 255))
            self.person_table.setItem(i, 1, level_item)

            # 分数
            score_item = QTableWidgetItem(str(row['score']))
            score_item.setTextAlignment(Qt.AlignCenter)
            self.person_table.setItem(i, 2, score_item)

            # 参与排班
            active_text = "✓ 是" if row['active'] else "✗ 否"
            active_item = QTableWidgetItem(active_text)
            active_item.setTextAlignment(Qt.AlignCenter)
            if not row['active']:
                active_item.setForeground(QColor(149, 165, 166))
            self.person_table.setItem(i, 3, active_item)

            # 生效规则 - 获取该人员的启用规则并计算实际得分
            # 只统计规则本身也启用(ACTIVE=1)的规则
            person_rules = self.db.get_person_rules(row['id'])
            total_modifier = 0.0
            enabled_rules = []
            for r in person_rules:
                r_dict = dict(r)
                if r_dict.get('enabled') == 1 and r_dict.get('active') == 1:
                    enabled_rules.append(r_dict['name'])
                    total_modifier += r_dict.get('score_modifier', 0)

            effective_score = row['score'] + total_modifier
            rules_text = f"{effective_score:.2f}"
            if enabled_rules:
                rules_text += f" ({', '.join(enabled_rules)})"
            else:
                rules_text += " (无规则)"

            rules_item = QTableWidgetItem(rules_text)
            rules_item.setTextAlignment(Qt.AlignCenter)
            if total_modifier > 0:
                rules_item.setForeground(QColor(39, 174, 96))
            elif total_modifier < 0:
                rules_item.setForeground(QColor(231, 76, 60))
            else:
                rules_item.setForeground(QColor(149, 165, 166))
            self.person_table.setItem(i, 4, rules_item)

        self.person_table.setColumnWidth(0, 120)
        self.person_table.setColumnWidth(1, 100)
        self.person_table.setColumnWidth(2, 60)
        self.person_table.setColumnWidth(3, 80)
        self.person_table.setColumnWidth(4, 200)
        
        self.update_stats()
        
        if hasattr(self, 'seat_stats_label'):
            self.update_seat_stats()
    
    def update_stats(self):
        """更新统计信息"""
        stats = self.db.get_statistics()
        self.stats_label.setText(f"总计: {stats['total']}人 (参与: {stats['active']}人)")
        self.c1_count_label.setText(f"C1: {stats['c1']}人")
        self.c2_count_label.setText(f"C2: {stats['c2']}人")
        self.c3_count_label.setText(f"C3: {stats['c3']}人")
        self.i_count_label.setText(f"I: {stats['i']}人")
        self.s_count_label.setText(f"S: {stats['s']}人")
    
    def on_person_double_clicked(self, index):
        """双击人员表格 - 打开规则选择对话框"""
        person_id = self.get_selected_person_id()
        if person_id:
            dialog = PersonRuleSelectDialog(self, person_id, self.db)
            dialog.exec_()
            self.load_data()

    def on_person_cell_clicked(self, row, col):
        """点击人员表格单元格 - 点击参与排班列时切换状态"""
        if col == 3:  # 参与排班列
            person_id = self.get_selected_person_id()
            if person_id:
                person = self.db.get_person_by_id(person_id)
                person_dict = dict(person)
                new_active = 0 if person_dict['active'] else 1
                self.db.update_person(person_id, person_dict['name'], person_dict['level'], new_active)
                self.load_data()
        # 注意：锁定列(col==4)是显示信息，不允许在人员管理中编辑
        # 锁定状态只在排班时通过手动选择来设置

    def on_search(self, text):
        """搜索功能"""
        if text:
            rows = self.db.search_persons(text)
            self.person_table.setRowCount(len(rows))
            
            for i, row in enumerate(rows):
                name_item = QTableWidgetItem(row['name'])
                name_item.setTextAlignment(Qt.AlignCenter)
                self.person_table.setItem(i, 0, name_item)

                level_item = QTableWidgetItem(row['level'])
                level_item.setTextAlignment(Qt.AlignCenter)
                if row['level'] == "C1":
                    level_item.setBackground(QColor(231, 76, 60))
                    level_item.setForeground(QColor(255, 255, 255))
                elif row['level'] == "C2":
                    level_item.setBackground(QColor(243, 156, 18))
                elif row['level'] == "C3":
                    level_item.setBackground(QColor(39, 174, 96))
                    level_item.setForeground(QColor(255, 255, 255))
                elif row['level'] == "I":
                    level_item.setBackground(QColor(52, 152, 219))
                    level_item.setForeground(QColor(255, 255, 255))
                elif row['level'] == "S":
                    level_item.setBackground(QColor(155, 89, 182))
                    level_item.setForeground(QColor(255, 255, 255))
                self.person_table.setItem(i, 1, level_item)

                # 分数
                score_item = QTableWidgetItem(str(row['score']))
                score_item.setTextAlignment(Qt.AlignCenter)
                self.person_table.setItem(i, 2, score_item)

            # 参与排班
            active_text = "✓ 是" if row['active'] else "✗ 否"
            active_item = QTableWidgetItem(active_text)
            active_item.setTextAlignment(Qt.AlignCenter)
            if not row['active']:
                active_item.setForeground(QColor(149, 165, 166))
            self.person_table.setItem(i, 3, active_item)

            # 生效规则 - 只统计规则本身也启用(active=1)的规则
            person_rules = self.db.get_person_rules(row['id'])
            total_modifier = 0.0
            enabled_rules = []
            for r in person_rules:
                r_dict = dict(r)
                if r_dict.get('enabled') == 1 and r_dict.get('active') == 1:
                    enabled_rules.append(r_dict['name'])
                    total_modifier += r_dict.get('score_modifier', 0)

            effective_score = row['score'] + total_modifier
            rules_text = f"{effective_score:.2f}"
            if enabled_rules:
                rules_text += f" ({', '.join(enabled_rules)})"
            else:
                rules_text += " (无规则)"

            rules_item = QTableWidgetItem(rules_text)
            rules_item.setTextAlignment(Qt.AlignCenter)
            if total_modifier > 0:
                rules_item.setForeground(QColor(39, 174, 96))
            elif total_modifier < 0:
                rules_item.setForeground(QColor(231, 76, 60))
            else:
                rules_item.setForeground(QColor(149, 165, 166))
            self.person_table.setItem(i, 4, rules_item)
        else:
            self.load_data()
    
    def get_selected_person_id(self):
        """获取选中的人员ID"""
        row = self.person_table.currentRow()
        if row < 0:
            return None
        name = self.person_table.item(row, 0).text()
        person = self.db.get_person_by_name(name)
        return person['id'] if person else None

    def get_selected_person(self):
        """获取选中的人员"""
        row = self.person_table.currentRow()
        if row < 0:
            return None
        name = self.person_table.item(row, 0).text()
        return self.db.get_person_by_name(name)
    
    def add_person(self):
        """添加人员"""
        dialog = PersonEditDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            success, msg = self.db.add_person(
                data['name'], data['level'], data['active'],
                data['phone'], data['email'], data['department'], data['notes']
            )
            if success:
                CustomMessageBox.information(self, "成功", msg)
                self.load_data()
                self.status_bar.showMessage(f"已添加: {data['name']}")
            else:
                CustomMessageBox.warning(self, "失败", msg)
    
    def edit_person(self):
        """编辑人员信息"""
        person_id = self.get_selected_person_id()
        if person_id is None:
            CustomMessageBox.warning(self, "请选择", "请先选择要编辑的人员")
            return
        
        person = self.db.get_person_by_id(person_id)
        dialog = PersonEditDialog(self, dict(person))
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            success, msg = self.db.update_person(
                person_id, data['name'], data['level'], data['active']
            )
            if success:
                CustomMessageBox.information(self, "成功", msg)
                self.load_data()
                self.status_bar.showMessage(f"已更新: {data['name']}")
            else:
                CustomMessageBox.warning(self, "失败", msg)
    
    def view_person_detail(self):
        """查看人员详情"""
        person_id = self.get_selected_person_id()
        if person_id is None:
            CustomMessageBox.warning(self, "请选择", "请先选择要查看的人员")
            return
        
        person = self.db.get_person_by_id(person_id)
        dialog = PersonDetailDialog(self, dict(person))
        dialog.exec_()
    
    def delete_person(self):
        """删除人员"""
        person_id = self.get_selected_person_id()
        if person_id is None:
            CustomMessageBox.warning(self, "请选择", "请先选择要删除的人员")
            return
        
        person = self.db.get_person_by_id(person_id)
        if CustomMessageBox.question(self, "确认删除",
            f"确定要删除 {person['name']} ({person['level']}) 吗?\n此操作不可恢复!"):
            success, msg = self.db.delete_person(person_id)
            if success:
                CustomMessageBox.information(self, "成功", msg)
                self.load_data()
            else:
                CustomMessageBox.warning(self, "失败", msg)

    def select_person_rules(self):
        """为人员选择生效的规则"""
        person_id = self.get_selected_person_id()
        if person_id is None:
            CustomMessageBox.warning(self, "请选择", "请先选择人员")
            return

        person = self.db.get_person_by_id(person_id)
        dialog = PersonRuleSelectDialog(self, person_id, self.db)
        dialog.exec_()

    def toggle_active(self):
        """切换参与排班状态"""
        person_id = self.get_selected_person_id()
        if person_id is None:
            CustomMessageBox.warning(self, "请选择", "请先选择人员")
            return
        
        person = self.db.get_person_by_id(person_id)
        new_status = 0 if person['active'] else 1
        success, msg = self.db.update_person(
            person_id, person['name'], person['level'], new_status,
            person['phone'] or "", person['email'] or "",
            person['department'] or "", person['notes'] or ""
        )
        
        if success:
            status_text = "参与排班" if new_status else "不参与排班"
            CustomMessageBox.information(self, "成功", f"{person['name']} 已设置为: {status_text}")
            self.load_data()
        else:
            CustomMessageBox.warning(self, "失败", msg)
    
    def run_schedule(self):
        """执行排班：先收集用户选择，再自动安排剩余人员"""
        
        # 获取可用席位
        available_seats = self.db.get_available_seats()
        
        if not available_seats:
            CustomMessageBox.warning(self, "席位未启用", "请至少启用一个席位")
            return
        
        active_persons = self.db.get_active_persons()
        
        if not active_persons:
            CustomMessageBox.warning(self, "人员不足", "没有参与排班的人员")
            return
        
        # 创建人员映射 - 将sqlite3.Row转换为字典
        person_map = {p['name']: dict(p) for p in active_persons}
        
        # 收集用户已选择的人员
        user_selections = {}  # {seat_id: [person_name, ...]}
        all_selected_names = []
        
        for widget in self.seat_result_widgets:
            if not widget.seat.available:
                continue
            
            selected = widget.get_selected_persons()
            if selected:
                user_selections[widget.seat.id] = selected
                all_selected_names.extend(selected)
        
        # 检查是否有重复选择
        if len(all_selected_names) != len(set(all_selected_names)):
            CustomMessageBox.warning(self, "选择错误", "同一个人不能被多次选择！")
            return
        
        # 验证用户选择是否符合规则
        for seat_id, selected_names in user_selections.items():
            seat_name = next((s['name'] for s in available_seats if s['id'] == seat_id), f"席位{seat_id}")
            
            c1_count = sum(1 for name in selected_names if person_map.get(name, {}).get('level') == 'C1')
            if c1_count > 1:
                CustomMessageBox.warning(self, "规则错误",
                    f"{seat_name}: C1人员不能超过1个！当前选择{c1_count}个C1")
                return

            # 检查C1和C2同席
            has_c1 = any(person_map.get(name, {}).get('level') == 'C1' for name in selected_names)
            has_c2 = any(person_map.get(name, {}).get('level') == 'C2' for name in selected_names)
            if has_c1 and has_c2:
                seat_name = next((s['name'] for s in available_seats if s['id'] == seat_id), f"席位{seat_id}")
                CustomMessageBox.warning(self, "规则错误",
                    f"{seat_name}: C1和C2不能同席！")
                return
        
        # 获取剩余未安排的人员
        remaining_persons = [p for p in active_persons if p['name'] not in all_selected_names]
        
        if not remaining_persons:
            # 用户已选择完所有人，直接显示结果
            for widget in self.seat_result_widgets:
                if widget.seat.id in user_selections:
                    names = user_selections[widget.seat.id]
                    persons = [person_map[n] for n in names]
                    # 显示已选择的人员
                    for i, name in enumerate(names):
                        if i < len(widget.person_combos):
                            idx = widget.person_combos[i].findData(name)
                            if idx >= 0:
                                widget.person_combos[i].setCurrentIndex(idx)
            CustomMessageBox.information(self, "完成", "所有人员已安排完成！")
            self.status_bar.showMessage("排班完成")
            return
        
        # 将剩余人员转为对象（包含规则分数修改和历史数据）
        remaining_objs = []
        for p in remaining_persons:
            p_dict = dict(p)  # sqlite3.Row转dict
            score_mod = self.db.get_person_score_modifier(p_dict['id'])
            total_hours = self.db.get_person_total_hours(p_dict['id'])
            seat_history_list = self.db.get_person_seat_history(p_dict['id'])
            seat_history = {r['seat_id']: r['count'] for r in seat_history_list}
            remaining_objs.append(Person(p_dict['id'], p_dict['name'], p_dict['level'], p_dict['score'], True, score_mod, seat_history, total_hours, False))  # 不使用数据库的locked字段

        # 创建席位对象
        seats = [
            Seat(s['id'], s['name'], s.get('app_name'), bool(s['available']), s['persons_count'], s['required_score'])
            for s in available_seats
        ]

        # 用户已选择的席位预先填充（手动选择的人员标记为手动指定）
        # 注意：这里不修改数据库，只在当次排班中临时标记
        pre_filled = {}
        for seat_id, selected_names in user_selections.items():
            pre_filled[seat_id] = []
            for name in selected_names:
                p = person_map.get(name)
                if p:
                    score_mod = self.db.get_person_score_modifier(p['id'])
                    total_hours = self.db.get_person_total_hours(p['id'])
                    seat_history_list = self.db.get_person_seat_history(p['id'])
                    seat_history = {r['seat_id']: r['count'] for r in seat_history_list}
                    # 用户手动选择的人员标记为手动指定(selfixed=True)，不参与系统安排
                    pre_filled[seat_id].append(Person(p['id'], p['name'], p['level'], p['score'], True, score_mod, seat_history, total_hours, True))
        
        self.status_bar.showMessage("正在排班...")
        QApplication.processEvents()
        
        # 创建调度器并生成排班（考虑用户已选择的人员）
        scheduler = ShiftScheduler(remaining_objs, seats)
        seats_dict, message = scheduler.generate_schedule_with_prefill(pre_filled)
        
        if seats_dict is None:
            CustomMessageBox.warning(self, "排班失败", message)
            self.status_bar.showMessage("排班失败")
        else:
            # 合并用户选择和自动安排的结果
            for widget in self.seat_result_widgets:
                seat_id = widget.seat.id
                
                # 先设置用户选择
                if seat_id in user_selections:
                    widget.set_selected_persons(user_selections[seat_id])
                
                # 再添加自动安排的人员
                if seat_id in seats_dict:
                    auto_persons = seats_dict[seat_id]
                    current_selected = widget.get_selected_persons()
                    
                    # 找到空位，添加自动安排的人员
                    for p in auto_persons:
                        if p.name not in current_selected:
                            # 找一个空的下拉框
                            for combo in widget.person_combos:
                                if not combo.currentData():
                                    idx = combo.findData(p.name)
                                    if idx >= 0:
                                        combo.setCurrentIndex(idx)
                                        break
            
            # 统计安排结果（seats_dict已包含所有人员：用户预填+自动安排，无需重复相加）
            total安排 = sum(len(v) for v in seats_dict.values())
            未安排 = len(active_persons) - total安排
            print(f"[DEBUG] placed_by_auto={sum(len(v) for v in seats_dict.values())}, user_selected={len(all_selected_names)}, total安排={total安排}, active={len(active_persons)}, 未安排={未安排}")
            
            if 未安排 > 0:
                CustomMessageBox.information(self, "排班完成",
                    f"已安排: {total安排}人\n未安排: {未安排}人\n\n{message}")
            else:
                CustomMessageBox.information(self, "排班完成", "所有人员已安排完成！")
            
            self.status_bar.showMessage(f"排班完成 - 已安排{total安排}人")

    def export_to_excel(self):
        """导出排班结果到Excel - 格式与样表一致
        每个席位3人ABC轮换:
        - 0850-1040: A+B (A管制席, B助理席)
        - 1030-1200: A+C (C替换B)
        - 1150-1240: B+C (B回来替换A)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from PyQt5.QtWidgets import QFileDialog

            file_path, _ = QFileDialog.getSaveFileName(
                self, "导出排班表", "", "Excel文件 (*.xlsx)"
            )
            if not file_path:
                return
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'

            wb = Workbook()
            ws = wb.active
            ws.title = "排班表"

            # 样式
            title_font = Font(bold=True, size=14)
            header_font = Font(bold=True, size=12)
            center = Alignment(horizontal='center', vertical='center', wrap_text=True)
            thin = Side(style='thin')
            thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
            blue_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
            orange_fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")

            today = datetime.now().strftime("%Y-%m-%d")
            assignments = []

            # 收集席位人员: 席位一→APP01, 席位二→APP02, ..., 席位五→APP05
            seat_map = {}
            seat_ids = {}
            for widget in self.seat_result_widgets:
                if not widget.seat.available:
                    continue
                selected_persons = widget.get_selected_persons()
                seat_map[widget.seat.name] = selected_persons
                seat_ids[widget.seat.name] = widget.seat.id

            # 时段定义 (与样表一致)
            # 早班3个时段(ABC轮换), 下午班4个时段, 晚班4个时段
            shifts = {
                '早班': {
                    'fill': blue_fill,
                    'seats': [('席位一', 'APP01'), ('席位二', 'APP02')],
                    'times': [
                        # (管制席时段, 管制席人员, 助理席时段, 助理席人员)
                        # 0850-1040: A+B; 1030-1200: A+C; 1150-1240: B+C
                        ('0850-1040', 'A', '0840-1000', 'B'),
                        ('1030-1200', 'A', '0950-1120', 'C'),
                        ('1150-1240', 'B', '1110-1230', 'C'),
                    ]
                },
                '下午班': {
                    'fill': green_fill,
                    'seats': [('席位三', 'APP03'), ('席位四', 'APP04')],
                    'times': [
                        ('1230-1400', 'B', '1220-1320', 'C'),
                        ('1350-1550', 'C', '1310-1500', 'A'),
                        ('1540-1730', 'A', '1450-1650', 'B'),
                        ('1720-1810', 'B', '1640-1800', 'C'),
                    ]
                },
                '晚班': {
                    'fill': orange_fill,
                    'seats': [('席位五', 'APP05'), ('席位一', 'APP01')],
                    'times': [
                        ('1800-1940', 'C', '1750-1900', 'A'),
                        ('1930-2100', 'A', '1850-2020', 'B'),
                        ('2050-2220', 'B', '2010-2140', 'C'),
                        ('2130-2210', 'C', '2120-2220', 'A'),
                    ]
                },
            }

            def get_person_by_pos(persons, pos):
                """根据位置(A/B/C)获取人员名"""
                if pos == 'A':
                    return persons[0] if len(persons) > 0 else 'A'
                elif pos == 'B':
                    return persons[1] if len(persons) > 1 else 'B'
                elif pos == 'C':
                    return persons[2] if len(persons) > 2 else 'C'
                return ''

            row = 1
            for shift_name, shift_data in shifts.items():
                fill = shift_data['fill']
                seats = shift_data['seats']
                times = shift_data['times']

                # 班次名称
                ws.cell(row, 1, shift_name).font = title_font
                ws.cell(row, 1).fill = fill

                # 两个席位标题
                ws.cell(row, 2, seats[0][1]).font = header_font
                ws.merge_cells(f'B{row}:E{row}')
                ws.cell(row, 2).fill = fill
                ws.cell(row, 2).alignment = center

                ws.cell(row, 6, seats[1][1]).font = header_font
                ws.merge_cells(f'F{row}:I{row}')
                ws.cell(row, 6).fill = fill
                ws.cell(row, 6).alignment = center

                row += 1

                # 表头
                headers = ['', '执勤时段', '管制席', '执勤时段', '助理席', '执勤时段', '管制席', '执勤时段', '助理席']
                for col, h in enumerate(headers, 1):
                    c = ws.cell(row, col, h)
                    c.font = header_font
                    c.alignment = center
                    c.border = thin_border
                    c.fill = fill
                row += 1

                # 时段数据
                seat1_name, seat2_name = seats[0][0], seats[1][0]
                seat1_persons = seat_map.get(seat1_name, [])
                seat2_persons = seat_map.get(seat2_name, [])

                for times_row in times:
                    t1, pos1, t2, pos2 = times_row

                    # 席位1
                    ws.cell(row, 2, t1).border = thin_border
                    ws.cell(row, 2).alignment = center
                    ws.cell(row, 3, get_person_by_pos(seat1_persons, pos1)).border = thin_border
                    ws.cell(row, 3).alignment = center
                    ws.cell(row, 4, t2).border = thin_border
                    ws.cell(row, 4).alignment = center
                    ws.cell(row, 5, get_person_by_pos(seat1_persons, pos2)).border = thin_border
                    ws.cell(row, 5).alignment = center

                    # 席位2
                    ws.cell(row, 6, t1).border = thin_border
                    ws.cell(row, 6).alignment = center
                    ws.cell(row, 7, get_person_by_pos(seat2_persons, pos1)).border = thin_border
                    ws.cell(row, 7).alignment = center
                    ws.cell(row, 8, t2).border = thin_border
                    ws.cell(row, 8).alignment = center
                    ws.cell(row, 9, get_person_by_pos(seat2_persons, pos2)).border = thin_border
                    ws.cell(row, 9).alignment = center

                    row += 1

                # 记录到历史 (少于2人按210分钟)
                for seat_name, persons in [(seat1_name, seat1_persons), (seat2_name, seat2_persons)]:
                    duration = 210 if len(persons) < 2 else 0
                    seat_id = seat_ids.get(seat_name)
                    for p in persons:
                        person = self.db.get_person_by_name(p)
                        if person:
                            assignments.append((today, person['id'], seat_id, f'{seat_name}-{shift_name}', duration))

                row += 1  # 班次间空行

            # 列宽
            ws.column_dimensions['A'].width = 10
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                ws.column_dimensions[col].width = 12

            wb.save(file_path)
            self.db.save_schedule(today, assignments)

            CustomMessageBox.information(self, "导出成功", f"排班表已导出到:\n{file_path}")
            self.status_bar.showMessage(f"已导出: {file_path}")

        except ImportError:
            CustomMessageBox.warning(self, "导出失败", "请先安装 openpyxl 库：\npip install openpyxl")
        except Exception as e:
            CustomMessageBox.warning(self, "导出失败", str(e))

    def closeEvent(self, event):
        """关闭时关闭数据库"""
        self.db.close()
        event.accept()


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")

    window = MainWindow()
    window.show()
    
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()